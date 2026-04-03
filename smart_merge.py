"""
Smart Merge — модуль сохранения пользовательских правок при перегенерации xlsx.

Логика:
  1. Читает старый xlsx (если существует)
  2. Сравнивает значения ячеек с последним парсингом из SQLite
  3. Определяет: что изменил пользователь, какие квартиры новые/проданные
  4. Сохраняет пользовательские листы

Результат — MergeResult, который передаётся в exporter.
"""
from __future__ import annotations

import sqlite3
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from parsers.apartments_base import ApartmentItem, logger

# Листы, создаваемые парсером — всё остальное = пользовательские
PARSER_SHEETS = {
    "Квартиры", "Все данные", "Средние цены",
    "Квартирография", "Диаграмма", "Информация о домах",
}

# Колонки листа «Все данные» и их маппинг на поля ApartmentItem
# Индексы 1-based, как в xlsx
FLAT_COL_MAP = {
    1: "city",
    2: "developer",
    3: "complex_name",
    4: "building",
    5: "rooms_label",
    6: "count",        # вычисляемое, не трекаем
    7: "apartment_number",
    8: "floor",
    9: "area",
    10: "living_area",
    11: "price",
    12: "price_per_meter",
    13: "url",
    14: "order",       # порядковый, не трекаем
    15: "item_id",     # скрытый столбец ID
}

# Колонки, которые мы отслеживаем на правки пользователя
TRACKABLE_COLS = {1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12}


@dataclass
class UserEdit:
    """Одна правка пользователя в ячейке."""
    column_idx: int
    column_name: str
    user_value: Any        # что написал пользователь
    parser_value: Any      # что было при последнем парсинге
    new_parser_value: Any  # что парсер хочет записать сейчас


@dataclass
class MergeResult:
    """Результат smart merge."""
    # item_id → {col_idx → UserEdit}
    user_edits: dict[str, dict[int, UserEdit]] = field(default_factory=dict)
    # ID квартир, которые пропали (проданы) — СЕЙЧАС, первый раз → красный
    sold_ids: set[str] = field(default_factory=set)
    # ID квартир, которые были проданы РАНЬШЕ (уже 2+ парсинг) → без цвета, с комментарием
    prev_sold_ids: set[str] = field(default_factory=set)
    # ID квартир, которые новые — СЕЙЧАС, первый раз → зелёный
    new_ids: set[str] = field(default_factory=set)
    # ID квартир, которые были новыми РАНЬШЕ (уже 2+ парсинг) → без цвета, с комментарием
    prev_new_ids: set[str] = field(default_factory=set)
    # Пользовательские листы (название → данные)
    user_sheets_data: list[dict] = field(default_factory=list)
    # Старый workbook для копирования листов
    old_workbook: Any = None


# ─── SQLite: таблица последних записанных значений ──────────

def init_last_values_table(conn: sqlite3.Connection) -> None:
    """Создать таблицы для хранения значений и статусов."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS apartment_last_values (
            site TEXT NOT NULL,
            item_id TEXT NOT NULL,
            column_idx INTEGER NOT NULL,
            value TEXT,
            PRIMARY KEY (site, item_id, column_idx)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS apartment_merge_status (
            site TEXT NOT NULL,
            item_id TEXT NOT NULL,
            status TEXT NOT NULL,
            marked_at TEXT NOT NULL,
            PRIMARY KEY (site, item_id)
        )
    """)
    conn.commit()


def _get_prev_statuses(conn: sqlite3.Connection, site: str) -> dict[str, tuple[str, str]]:
    """Получить статусы из прошлого парсинга. Returns: {item_id: (status, marked_at)}"""
    rows = conn.execute(
        "SELECT item_id, status, marked_at FROM apartment_merge_status WHERE site = ?",
        (site,),
    ).fetchall()
    return {r[0]: (r[1], r[2]) for r in rows}


def save_merge_statuses(
    conn: sqlite3.Connection, site: str,
    new_ids: set[str], sold_ids: set[str],
) -> None:
    """Сохранить статусы после текущего парсинга.

    Логика:
    - Новые 'new' → записать
    - Новые 'sold' → записать
    - Бывшие 'new' которые уже не новые → удалить (цвет убран)
    - Бывшие 'sold' которые вернулись в продажу → удалить
    - Бывшие 'sold' которые всё ещё проданы → удалить (цвет убран, комментарий остался в xlsx)
    """
    now = datetime.now().strftime("%d.%m.%Y")

    # Очищаем старые статусы
    conn.execute("DELETE FROM apartment_merge_status WHERE site = ?", (site,))

    # Записываем только ТЕКУЩИЕ новые и проданные (для следующего парсинга)
    rows = []
    for item_id in new_ids:
        rows.append((site, item_id, "new", now))
    for item_id in sold_ids:
        rows.append((site, item_id, "sold", now))

    if rows:
        conn.executemany(
            "INSERT INTO apartment_merge_status (site, item_id, status, marked_at) VALUES (?, ?, ?, ?)",
            rows,
        )

    conn.commit()
    logger.info("Статусы: %d новых, %d проданных", len(new_ids), len(sold_ids))


def save_written_values(
    conn: sqlite3.Connection, site: str, items: list[ApartmentItem],
) -> None:
    """Сохранить текущие значения, записанные парсером в xlsx."""
    rows = []
    for item in items:
        values = _item_to_flat_values(item)
        for col_idx in TRACKABLE_COLS:
            val = values.get(col_idx)
            rows.append((site, item.item_id, col_idx, _to_str(val)))

    conn.executemany(
        """INSERT OR REPLACE INTO apartment_last_values
           (site, item_id, column_idx, value) VALUES (?, ?, ?, ?)""",
        rows,
    )
    conn.commit()
    logger.info("Сохранено %d значений в apartment_last_values", len(rows))


def get_last_written_values(
    conn: sqlite3.Connection, site: str,
) -> dict[str, dict[int, str]]:
    """Получить последние записанные парсером значения.

    Returns: {item_id: {col_idx: value_str}}
    """
    rows = conn.execute(
        "SELECT item_id, column_idx, value FROM apartment_last_values WHERE site = ?",
        (site,),
    ).fetchall()

    result: dict[str, dict[int, str]] = {}
    for item_id, col_idx, value in rows:
        result.setdefault(item_id, {})[col_idx] = value

    return result


# ─── Чтение старого xlsx ────────────────────────────────────

def _read_old_flat_sheet(
    xlsx_path: Path,
) -> tuple[dict[str, dict[int, Any]], Any]:
    """Прочитать лист «Все данные» из старого файла.

    Returns:
        (data_map, workbook)
        data_map: {item_id: {col_idx: cell_value}}
    """
    wb = load_workbook(str(xlsx_path), data_only=True)

    if "Все данные" not in wb.sheetnames:
        return {}, wb

    ws = wb["Все данные"]
    data_map: dict[str, dict[int, Any]] = {}

    # Найти столбец ID (последний)
    max_col = ws.max_column
    id_col = None
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header == "ID":
            id_col = col
            break

    if id_col is None:
        logger.warning("Столбец ID не найден в старом файле, пропускаем merge")
        return {}, wb

    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=id_col).value
        if not item_id:
            continue

        row_data = {}
        for col_idx in TRACKABLE_COLS:
            row_data[col_idx] = ws.cell(row=row, column=col_idx).value

        data_map[str(item_id)] = row_data

    logger.info("Прочитано %d квартир из старого файла", len(data_map))
    return data_map, wb


# ─── Детекция правок пользователя ───────────────────────────

def _detect_user_edits(
    old_xlsx_data: dict[str, dict[int, Any]],
    last_written: dict[str, dict[int, str]],
    new_items_map: dict[str, ApartmentItem],
) -> dict[str, dict[int, UserEdit]]:
    """Определить, какие ячейки изменил пользователь.

    Логика: если значение в xlsx ≠ значение из SQLite (последний парсинг),
    значит пользователь правил.
    """
    edits: dict[str, dict[int, UserEdit]] = {}

    for item_id, xlsx_row in old_xlsx_data.items():
        written_row = last_written.get(item_id, {})
        new_item = new_items_map.get(item_id)

        for col_idx in TRACKABLE_COLS:
            xlsx_val = xlsx_row.get(col_idx)
            written_val_str = written_row.get(col_idx)

            if written_val_str is None:
                # Нет данных о том, что записывал парсер — не можем сравнить
                continue

            xlsx_val_str = _to_str(xlsx_val)

            if xlsx_val_str != written_val_str:
                # Пользователь изменил!
                new_parser_val = None
                if new_item:
                    new_values = _item_to_flat_values(new_item)
                    new_parser_val = new_values.get(col_idx)

                col_name = FLAT_COL_MAP.get(col_idx, f"col_{col_idx}")
                edits.setdefault(item_id, {})[col_idx] = UserEdit(
                    column_idx=col_idx,
                    column_name=col_name,
                    user_value=xlsx_val,
                    parser_value=written_val_str,
                    new_parser_value=new_parser_val,
                )

    if edits:
        total_edits = sum(len(v) for v in edits.values())
        logger.info(
            "Обнаружено %d пользовательских правок в %d квартирах",
            total_edits, len(edits),
        )

    return edits


# ─── Основная функция ───────────────────────────────────────

def smart_merge(
    new_items: list[ApartmentItem],
    xlsx_path: Path,
    conn: sqlite3.Connection,
    site: str,
) -> MergeResult:
    """Выполнить smart merge: определить правки, новые, проданные квартиры.

    Args:
        new_items: свежие данные из парсера
        xlsx_path: путь к существующему xlsx (может не существовать)
        conn: подключение к SQLite
        site: ключ сайта ('domrf', 'pik', ...)

    Returns:
        MergeResult с правками, статусами и пользовательскими листами
    """
    init_last_values_table(conn)

    result = MergeResult()
    new_items_map = {item.item_id: item for item in new_items}
    current_ids = set(new_items_map.keys())

    # Получаем статусы прошлого парсинга
    prev_statuses = _get_prev_statuses(conn, site)

    if not xlsx_path.exists():
        logger.info("Старый файл не найден — первый запуск")
        # Не помечаем все как новые при первом запуске
        return result

    # Читаем старый файл
    old_xlsx_data, old_wb = _read_old_flat_sheet(xlsx_path)
    result.old_workbook = old_wb

    if not old_xlsx_data:
        logger.info("Старый файл пуст или без столбца ID")
        return result

    old_ids = set(old_xlsx_data.keys())

    # ── Определяем статусы ──

    # Новые: есть в парсинге, нет в старом файле, И не были sold раньше
    really_new = current_ids - old_ids
    # Убираем те, что были "sold" но вернулись в продажу
    returned_from_sold = {iid for iid in really_new
                          if iid in prev_statuses and prev_statuses[iid][0] == "sold"}
    really_new -= returned_from_sold

    if returned_from_sold:
        logger.info("Вернулись в продажу: %d", len(returned_from_sold))

    # Бывшие новые (были 'new' в прошлом парсинге, сейчас уже не новые → убрать цвет)
    result.prev_new_ids = {iid for iid, (st, _) in prev_statuses.items()
                           if st == "new" and iid not in really_new}

    # Текущие новые → зелёный цвет
    result.new_ids = really_new
    if result.new_ids:
        logger.info("Новых квартир: %d (зелёный)", len(result.new_ids))
    if result.prev_new_ids:
        logger.info("Ранее новых (цвет убран, комментарий остался): %d", len(result.prev_new_ids))

    # Проданные: есть в старом, нет в текущем парсинге
    currently_gone = old_ids - current_ids
    # Из них: те, что были sold в прошлом парсинге → prev_sold (без цвета)
    result.prev_sold_ids = {iid for iid in currently_gone
                            if iid in prev_statuses and prev_statuses[iid][0] == "sold"}
    # Впервые пропавшие → красный
    result.sold_ids = currently_gone - result.prev_sold_ids

    if result.sold_ids:
        logger.info("Проданных квартир (новые, красный): %d", len(result.sold_ids))
    if result.prev_sold_ids:
        logger.info("Ранее проданных (цвет убран, комментарий остался): %d", len(result.prev_sold_ids))

    # Получаем последние записанные парсером значения
    last_written = get_last_written_values(conn, site)

    # Детектируем правки пользователя
    result.user_edits = _detect_user_edits(
        old_xlsx_data, last_written, new_items_map,
    )

    # Собираем имена пользовательских листов
    for sheet_name in old_wb.sheetnames:
        if sheet_name not in PARSER_SHEETS:
            result.user_sheets_data.append({"name": sheet_name})
            logger.info("Найден пользовательский лист: %s", sheet_name)

    return result


# ─── Утилиты ────────────────────────────────────────────────

def _to_str(val: Any) -> str:
    """Привести значение к строке для сравнения."""
    if val is None:
        return ""
    if isinstance(val, float):
        # Округляем чтобы избежать проблем с плавающей точкой
        return str(round(val, 2))
    return str(val)


def _item_to_flat_values(item: ApartmentItem) -> dict[int, Any]:
    """Преобразовать ApartmentItem в словарь {col_idx: value} для flat sheet."""
    dev = getattr(item, "developer", None) or ""
    living = getattr(item, "living_area", None)

    building_display = item.building
    if "||" in item.building:
        building_display = item.building.split("||", 1)[0].strip()

    try:
        number_val = int(item.apartment_number) if item.apartment_number else ""
    except ValueError:
        number_val = item.apartment_number or ""

    return {
        1: item.city,
        2: dev,
        3: item.complex_name,
        4: building_display,
        5: item.rooms_label,
        7: number_val,
        8: item.floor,
        9: item.area,
        10: living if living is not None else "",
        11: item.price if item.price is not None else "",
        12: item.price_per_meter if item.price_per_meter is not None else "",
    }


def copy_user_sheets(target_wb, source_wb, sheet_names: list[str]) -> None:
    """Копировать пользовательские листы из старого workbook в новый."""
    for name in sheet_names:
        if name not in source_wb.sheetnames:
            continue

        src_ws = source_wb[name]
        tgt_ws = target_wb.create_sheet(name)

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = tgt_ws.cell(
                    row=cell.row, column=cell.column, value=cell.value,
                )
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                if cell.comment:
                    new_cell.comment = copy(cell.comment)
                if cell.hyperlink:
                    new_cell.hyperlink = copy(cell.hyperlink)

        # Копируем ширины колонок
        for col_letter, dim in src_ws.column_dimensions.items():
            tgt_ws.column_dimensions[col_letter].width = dim.width

        # Копируем высоты строк
        for row_num, dim in src_ws.row_dimensions.items():
            tgt_ws.row_dimensions[row_num].height = dim.height

        # Копируем merged cells
        for merged in src_ws.merged_cells.ranges:
            tgt_ws.merge_cells(str(merged))

        logger.info("Скопирован пользовательский лист: %s", name)
