#!/usr/bin/env python3
"""
Тестовый запуск парсера квартир ДОМ.РФ — только новые объекты.

Использует configs/domrf_apartments_test.yaml (11 новых объектов).
Результат: output/apartments_DomRF_test.xlsx

Использование:
    python runners/run_apartments/run_domrf_test.py
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from parsers.apartments_base import (
    init_db, save_items, backup_db, validate_items,
    get_all_known_ids, calc_avg_prices, rooms_label, logger,
)
from parsers.domrf_apartments import DomRfApartmentParser, ObjectInfo
from exporter_apartments import export_apartments_xlsx

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter as _Counter
from kvartirografia import add_kvartirografia_sheets
from smart_merge import smart_merge, save_written_values, save_merge_statuses, copy_user_sheets
from parsers.apartments_base import OUTPUT_DIR
import re


_QUARTER_END_DATES = {
    "I": "31 марта",
    "II": "30 июня",
    "III": "30 сентября",
    "IV": "31 декабря",
}

_QUARTER_RE = re.compile(r'(I{1,3}V?)\s*(?:квартал|кв\.?)\s*(\d{4})', re.IGNORECASE)


def _quarter_comment(text: str) -> str | None:
    """Расшифровка квартала → конкретная дата конца квартала."""
    if not text:
        return None
    m = _QUARTER_RE.search(text)
    if not m:
        return None
    roman = m.group(1).upper()
    year = m.group(2)
    end_date = _QUARTER_END_DATES.get(roman)
    if not end_date:
        return None
    return f"{end_date} {year} года"




def _add_object_info_sheet(wb, object_infos: list[ObjectInfo]) -> None:
    """Добавить лист «Информация о домах» с данными из шапок страниц."""
    ws = wb.create_sheet("Информация о домах")

    headers = [
        "Объект", "ЖК", "Застройщик",
        "Ввод в эксплуатацию", "Выдача ключей",
        "Средняя цена за 1 м²", "Распроданность",
        "Всего квартир", "Продано квартир",
    ]
    base_url = "https://xn--80az8a.xn--d1aqf.xn--p1ai"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Сортировка: по застройщику, потом ЖК, потом object_id
    sorted_infos = sorted(object_infos, key=lambda o: (
        o.developer.lower(), o.complex_name.lower(), o.object_id,
    ))

    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    for i, info in enumerate(sorted_infos):
        row = i + 2

        # Очистка значений-дефисов (на сайте "-" означает "нет данных")
        def clean(val):
            if isinstance(val, str) and val.strip() in ("-", "–", "—"):
                return ""
            return val

        row_data = [
            info.object_id,
            info.complex_name,
            info.developer,
            clean(info.commissioning),
            clean(info.keys_date),
            clean(info.avg_price_per_meter),
            clean(info.sold_percent),
            info.total_apartments or "",
            info.sold_apartments or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # Первый столбец — ссылка на объект
        obj_cell = ws.cell(row=row, column=1)
        obj_url = f"{base_url}/сервисы/каталог-новостроек/объект/{info.object_id}"
        obj_cell.hyperlink = obj_url
        obj_cell.font = link_font

        # Примечание с расшифровкой квартала (только для ввода в эксплуатацию)
        hint = _quarter_comment(info.commissioning)
        if hint:
            c = ws.cell(row=row, column=4)
            c.comment = Comment(hint, "Parser")
            c.comment.width = 200
            c.comment.height = 30

    # Ширина колонок
    widths = [12, 25, 18, 20, 18, 22, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    last_row = len(sorted_infos) + 1
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


def _parse_args():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--cdp", type=int, default=None, metavar="PORT",
                    help="Подключиться к Chrome через CDP (порт remote-debugging)")
    return ap.parse_args()


async def main() -> int:
    args = _parse_args()

    logger.info("=" * 50)
    logger.info("ТЕСТ: парсер квартир ДОМ.РФ — только новые объекты")
    logger.info("=" * 50)

    config_path = Path(__file__).resolve().parent.parent.parent / "configs" / "domrf_apartments_test.yaml"

    backup_db()
    conn = init_db()

    try:
        parser = DomRfApartmentParser(config_path=config_path, cdp_port=args.cdp)
        items, object_infos = await parser.parse_all()

        if not items:
            logger.error("Парсер не вернул ни одной квартиры!")
            return 1

        warnings = validate_items(items)
        if warnings:
            logger.warning("Обнаружено %d предупреждений валидации", len(warnings))

        previously_known = get_all_known_ids(conn, "domrf")

        updated = save_items(conn, items)
        logger.info("Обновлено записей в БД: %d", updated)

        # Smart merge — определяем правки пользователя, новые/проданные
        xlsx_filename = "apartments_DomRF_test.xlsx"
        xlsx_path = OUTPUT_DIR / xlsx_filename
        merge_result = smart_merge(items, xlsx_path, conn, "domrf")

        output_path = export_apartments_xlsx(
            items, conn,
            filename=xlsx_filename,
            previously_known=previously_known,
            merge_result=merge_result,
        )

        # Добавляем листы
        wb = load_workbook(str(output_path))
        add_kvartirografia_sheets(wb, items)
        _add_object_info_sheet(wb, object_infos)

        # Копируем пользовательские листы из старого файла
        if merge_result.old_workbook and merge_result.user_sheets_data:
            user_sheet_names = [s["name"] for s in merge_result.user_sheets_data]
            copy_user_sheets(wb, merge_result.old_workbook, user_sheet_names)

        wb.save(str(output_path))

        # Сохраняем записанные значения и статусы для будущих сравнений
        save_written_values(conn, "domrf", items)
        save_merge_statuses(conn, "domrf", merge_result.new_ids, merge_result.sold_ids)

        logger.info("Файл готов: %s", output_path)

        # Статистика
        stats = calc_avg_prices(items)
        logger.info("Статистика:")
        logger.info("  Всего квартир: %d", len(items))
        for r, data in stats["by_rooms"].items():
            if data["avg_price"] > 0:
                logger.info(
                    "    %s: %d шт., ср. цена: %s ₽, ср. цена/м²: %s ₽",
                    rooms_label(r), data["count"],
                    f"{data['avg_price']:,.0f}", f"{data['avg_ppm']:,.0f}",
                )
            else:
                logger.info("    %s: %d шт.", rooms_label(r), data["count"])

        logger.info("  Объектов с информацией о доме: %d", len(object_infos))

        return 0

    except Exception as exc:
        logger.exception("Критическая ошибка: %s", exc)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
