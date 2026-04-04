"""
Генератор единого XLSX-файла расчёта ЕВА.

Компиляция данных из:
  - apartments_DomRF.xlsx   (квартиры с дом.рф: площади, жилая/нежилая, кол-во)
  - storehouses_DomRF.xlsx  (кладовки с дом.рф: общее кол-во)
  - apartments_[DEV].xlsx   (квартиры застройщика: цены)
  - storehouses_[DEV].xlsx  (кладовки застройщика: цены, остаток)
  + «Информация о домах» из apartments_DomRF (квартал сдачи → дни)

Два листа:
  1. «кладовки (застройщик)» — строка на каждую кладовку
  2. «жк» — строка на каждый корпус
"""
from __future__ import annotations

import logging
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("eva")

PROJECT_DIR = Path(__file__).resolve().parent.parent
CONFIGS_DIR = PROJECT_DIR / "configs"
APT_DIR = PROJECT_DIR / "apartments"
STORE_DIR = PROJECT_DIR / "output"

# ─── Стили ───────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
NO_BORDER = Border()  # без линий по умолчанию
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
MANUAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SCORE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="2F5496", size=11)
GRAY_FONT = Font(name="Calibri", size=11, color="BFBFBF")

SEGMENT_BOTTOM = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="4472C4"),
)

# Правая жирная граница (разделитель блоков столбцов)
BLOCK_RIGHT = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="medium", color="4472C4"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Столбцы-разделители блоков на листе «жк»
# Последний столбец каждого блока:
# 18 (ср. цена м²) | 23 (квартирография) | 28 (ср. площадь) | 33 (ср. нежилая)
# Столбцы с жирной ПРАВОЙ границей на листе «жк»
# 4=корпус, 7=ссылка застр., 8=первый этап, 12=удобство доступа,
# 16=кладовых ДОМ.РФ, 23=4к+ квартирография, 28=ср.площадь 4к+,
# 33=ср.нежилая 4к+, 34=исх.порядок
# Столбцы с жирной ПРАВОЙ границей
JK_BLOCK_DIVIDERS = {4, 7, 8, 12, 16, 23, 28, 33, 34}
# Столбцы с жирной ЛЕВОЙ границей = столбцы ПОСЛЕ разделителей + первый этап
JK_BLOCK_LEFT_DIVIDERS = {5, 8, 9, 13, 17, 24, 29, 34}

DOMRF_BASE = "https://xn--80az8a.xn--d1aqf.xn--p1ai"

# Маппинг site_key → человекочитаемое имя застройщика
DEVELOPER_NAMES = {
    "pik": "ПИК",
    "akbarsdom": "Ак Барс Дом",
    "glorax": "GloraX",
    "smu88": "СМУ-88",
    "unistroy": "Унистрой",
    "domrf": "ДОМ.РФ",
}


# ─── Алиасы и нормализация ─────────────────────────────────

_COMPLEX_ALIASES: dict[str, str] = {}  # заполняется из eva.yaml при загрузке
_BUILDING_ALIASES: dict[str, dict[str, str]] = {}  # ЖК → {паттерн: корпус DomRF}


def _apply_alias(complex_name: str) -> str:
    """Привести имя ЖК к каноническому через алиасы."""
    return _COMPLEX_ALIASES.get(complex_name, complex_name)


def _apply_building_alias(complex_name: str, building: str) -> str:
    """Привести корпус застройщика к корпусу DomRF через маппинг."""
    cn = _apply_alias(complex_name)
    aliases = _BUILDING_ALIASES.get(cn, {})
    if not aliases:
        return building
    for pattern, target in aliases.items():
        if pattern.startswith("^"):
            # Prefix match
            prefix = pattern[1:]
            if building.startswith(prefix):
                return target
        else:
            # Contains match (с учётом спецсимволов в Адали)
            if pattern in building:
                return target
    return building


def _norm(s: str) -> str:
    """Нормализовать строку для матчинга."""
    if not s:
        return ""
    s = s.lower().strip()
    # "Корпус 10" → "10", "корпус 1" → "1"
    s = re.sub(r'^корпус\s*', '', s)
    # Убрать дефисы, пробелы, подчёркивания, слеши
    s = re.sub(r'[\s\-–—_/]', '', s)
    return s


def _match_key(city: str, complex_name: str, building: str) -> tuple:
    return (_norm(city), _norm(_apply_alias(complex_name)), _norm(building))


# ─── Загрузка данных из БД ───────────────────────────────

def _load_apartments(conn: sqlite3.Connection, site_filter: str | None = None) -> list[dict]:
    """Загрузить последние записи квартир."""
    query = """
        SELECT a.site, a.city, a.complex_name, a.building, a.item_id,
               a.rooms, a.floor, a.area, a.price, a.price_per_meter,
               a.living_area, a.url, a.object_id
        FROM apartment_prices a
        INNER JOIN (
            SELECT site, item_id, MAX(parsed_at) AS max_pa
            FROM apartment_prices
            GROUP BY site, item_id
        ) latest ON a.site = latest.site AND a.item_id = latest.item_id
                 AND a.parsed_at = latest.max_pa
    """
    if site_filter:
        query += " WHERE a.site = ?"
        rows = conn.execute(query, (site_filter,)).fetchall()
    else:
        rows = conn.execute(query).fetchall()
    cols = ["site", "city", "complex_name", "building", "item_id",
            "rooms", "floor", "area", "price", "price_per_meter",
            "living_area", "url", "object_id"]
    return [dict(zip(cols, r)) for r in rows]


def _load_storehouses(conn: sqlite3.Connection, site_filter: str | None = None) -> list[dict]:
    """Загрузить последние записи кладовок."""
    query = """
        SELECT p.site, p.city, p.complex_name, p.building, p.item_id,
               p.item_number, p.area, p.price, p.price_per_meter, p.url
        FROM prices p
        INNER JOIN (
            SELECT site, item_id, MAX(parsed_at) AS max_pa
            FROM prices
            GROUP BY site, item_id
        ) latest ON p.site = latest.site AND p.item_id = latest.item_id
                 AND p.parsed_at = latest.max_pa
    """
    if site_filter:
        query += " WHERE p.site = ?"
        rows = conn.execute(query, (site_filter,)).fetchall()
    else:
        rows = conn.execute(query).fetchall()
    cols = ["site", "city", "complex_name", "building", "item_id",
            "item_number", "area", "price", "price_per_meter", "url"]
    return [dict(zip(cols, r)) for r in rows]


# ─── Загрузка ObjectInfo из xlsx ─────────────────────────

@dataclass
class ObjectInfo:
    object_id: int
    complex_name: str
    developer: str
    commissioning: str = ""
    total_apartments: int = 0


def load_object_infos(xlsx_path: Path | None = None) -> list[ObjectInfo]:
    """Прочитать 'Информация о домах' из apartments_DomRF.xlsx."""
    if xlsx_path is None:
        xlsx_path = APT_DIR / "apartments_DomRF.xlsx"
    if not xlsx_path.exists():
        logger.warning("Нет файла %s — commissioning будет пустым", xlsx_path)
        return []

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Информация о домах" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Информация о домах"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        result.append(ObjectInfo(
            object_id=int(row[0]) if row[0] else 0,
            complex_name=str(row[1] or ""),
            developer=str(row[2] or ""),
            commissioning=str(row[3] or ""),
            total_apartments=int(row[7]) if row[7] else 0,
        ))
    wb.close()
    logger.info("Загружено %d ObjectInfo из %s", len(result), xlsx_path.name)
    return result


def load_domrf_living_areas(xlsx_path: Path | None = None) -> dict[tuple, dict]:
    """
    Прочитать жилую площадь из apartments_DomRF.xlsx → 'Все данные'.
    Возвращает два уровня:
      - per-building: {(norm_city, norm_complex, norm_building): {rooms: (avg_area, avg_living)}}
      - complex fallback: {(norm_city, norm_complex): {rooms: (avg_area, avg_living)}}
    Объединено в один dict; ключ из 3 элементов = per-building, из 2 = complex.
    """
    if xlsx_path is None:
        xlsx_path = APT_DIR / "apartments_DomRF.xlsx"
    if not xlsx_path.exists():
        return {}

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Все данные" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Все данные"]
    # Cols: 1=Город, 2=Застройщик, 3=ЖК, 4=Корпус, 5=Тип, 9=Площадь, 10=Жилая площадь
    room_map = {"Студия": 0, "1-комн.": 1, "2-комн.": 2, "3-комн.": 3, "4-комн.": 4}

    def _building_base_la(bld: str) -> str:
        return bld.split("||")[0].strip() if "||" in bld else bld.strip()

    # Собираем: per-building и per-complex
    raw_bld: dict[tuple, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    raw_complex: dict[tuple, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        city = str(row[0] or "")
        cn = str(row[2] or "")
        bld = _building_base_la(str(row[3] or ""))
        room_type = room_map.get(str(row[4] or ""))
        if room_type is None:
            continue
        area = row[8] if row[8] else None
        living = row[9] if row[9] else None
        if area and living:
            ck = _complex_key(city, cn)
            bk = (ck[0], ck[1], _norm(bld))
            raw_bld[bk][room_type].append((float(area), float(living)))
            raw_complex[ck][room_type].append((float(area), float(living)))

    # Агрегация
    result: dict[tuple, dict] = {}

    # Per-building
    for bk, rooms_data in raw_bld.items():
        agg = {}
        for rooms, pairs in rooms_data.items():
            areas = [p[0] for p in pairs]
            livings = [p[1] for p in pairs]
            agg[rooms] = (round(sum(areas) / len(areas), 1),
                          round(sum(livings) / len(livings), 1))
        result[bk] = agg

    # Complex fallback (ключ из 2 элементов)
    for ck, rooms_data in raw_complex.items():
        agg = {}
        for rooms, pairs in rooms_data.items():
            areas = [p[0] for p in pairs]
            livings = [p[1] for p in pairs]
            agg[rooms] = (round(sum(areas) / len(areas), 1),
                          round(sum(livings) / len(livings), 1))
        result[ck] = agg

    wb.close()
    bld_count = sum(1 for k in result if len(k) == 3)
    complex_count = sum(1 for k in result if len(k) == 2)
    logger.info("Загружено жилых площадей: %d корпусов + %d комплексов", bld_count, complex_count)
    return result


def load_developer_urls() -> dict[str, str]:
    """
    Загрузить base_url сайтов застройщиков из yaml-конфигов.
    Возвращает: {developer_name: base_url}
    """
    result: dict[str, str] = {}

    for cfg_name in ["pik", "akbarsdom", "glorax", "smu88", "unistroy"]:
        cfg_path = CONFIGS_DIR / f"{cfg_name}.yaml"
        if not cfg_path.exists():
            continue
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)

        base_url = cfg.get("base_url", cfg.get("url", ""))
        if base_url:
            # Привязываем к человекочитаемому имени
            dev_name = DEVELOPER_NAMES.get(cfg_name, cfg_name)
            result[dev_name] = base_url

    return result


# ─── Расчёт дней до сдачи ────────────────────────────────

def _avg_ppm(items: list[dict]) -> float:
    ppms = [a["price_per_meter"] for a in items if a.get("price_per_meter")]
    return round(sum(ppms) / len(ppms)) if ppms else 0


def _calc_priority_ppm(apts: list[dict]) -> float:
    """
    Средняя цена за м² по приоритету (п.1-2, п.4):
    1) Только 1к
    2) Среднее между студиями и 2к
    4) Все данные, что есть
    П.3 (соседний корпус) применяется отдельно в _apply_neighbor_ppm.
    """
    # П.1: только 1к
    one_k = [a for a in apts if a["rooms"] == 1]
    if one_k:
        avg = _avg_ppm(one_k)
        if avg:
            return avg

    # П.2: среднее между студиями и 2к
    studios = [a for a in apts if a["rooms"] == 0]
    two_k = [a for a in apts if a["rooms"] == 2]
    if studios and two_k:
        avg_s = _avg_ppm(studios)
        avg_2 = _avg_ppm(two_k)
        if avg_s and avg_2:
            return round((avg_s + avg_2) / 2)

    # П.4: всё что есть
    return _avg_ppm(apts)


def _apply_neighbor_ppm(buildings_list: list) -> None:
    """
    П.3: если у корпуса ppm=0, взять ppm по п.1 от соседнего корпуса
    того же ЖК. Сосед: сначала шаг назад, если невозможно — шаг вперёд.
    """
    from itertools import groupby

    # Группируем по ЖК
    def ck(b):
        return (_norm(b.city), _norm(b.complex_name))

    by_complex: dict[tuple, list] = defaultdict(list)
    for b in buildings_list:
        by_complex[ck(b)].append(b)

    for complex_key, blds in by_complex.items():
        if len(blds) < 2:
            continue

        for i, b in enumerate(blds):
            if b.dev_avg_apt_ppm > 0:
                continue

            # Шаг назад
            neighbor_ppm = 0
            if i > 0 and blds[i - 1].dev_avg_apt_ppm > 0:
                neighbor_ppm = blds[i - 1].dev_avg_apt_ppm
            # Шаг вперёд
            elif i + 1 < len(blds) and blds[i + 1].dev_avg_apt_ppm > 0:
                neighbor_ppm = blds[i + 1].dev_avg_apt_ppm

            if neighbor_ppm:
                b.dev_avg_apt_ppm = neighbor_ppm


_QUARTER_RE = re.compile(r'(I{1,3}V?)\s*(?:квартал|кв\.?)\s*(\d{4})', re.IGNORECASE)
_QUARTER_END = {"I": (3, 31), "II": (6, 30), "III": (9, 30), "IV": (12, 31)}


def _days_until(commissioning: str) -> int | None:
    """Посчитать дни от сегодня до конца квартала сдачи."""
    if not commissioning:
        return None
    m = _QUARTER_RE.search(commissioning)
    if not m:
        return None
    roman = m.group(1).upper()
    year = int(m.group(2))
    month, day = _QUARTER_END.get(roman, (12, 31))
    try:
        target = date(year, month, day)
    except ValueError:
        return None
    delta = (target - date.today()).days
    return delta


# ─── Конфиг разбалловки ──────────────────────────────────

def load_eva_config() -> dict:
    global _COMPLEX_ALIASES, _BUILDING_ALIASES
    with open(CONFIGS_DIR / "eva.yaml", "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    _COMPLEX_ALIASES = cfg.get("complex_aliases", {})
    _BUILDING_ALIASES = cfg.get("building_aliases", {})
    return cfg


def _score_by_thresholds(value: float, thresholds: list[dict]) -> int:
    for t in thresholds:
        if value < t["max"]:
            return t["points"]
    return thresholds[-1]["points"] if thresholds else 0


# ─── Агрегация данных ────────────────────────────────────

@dataclass
class BuildingAgg:
    """Агрегированные данные одного корпуса."""
    city: str = ""
    developer: str = ""
    complex_name: str = ""
    building: str = ""

    # ObjectInfo (дом.рф)
    object_ids: list = field(default_factory=list)
    commissioning: str = ""
    days_until: int | None = None

    # Кол-во с дом.рф (per-building для квартир, per-COMPLEX для кладовок)
    domrf_apt_count: int = 0
    domrf_store_count: int = 0    # на уровне ВСЕГО ЖК

    # Кол-во у застройщика (остаток в продаже)
    dev_store_count: int = 0

    # Средняя цена м² квартиры (от застройщика)
    dev_avg_apt_ppm: float = 0

    # Квартирография (кол-во по типам, дом.рф, per-building)
    rooms_count: dict = field(default_factory=dict)
    avg_area: dict = field(default_factory=dict)
    avg_non_living: dict = field(default_factory=dict)

    # Кладовки застройщика (для листа 1)
    dev_storehouses: list = field(default_factory=list)
    # Кладовки дом.рф (fallback, per-complex)
    domrf_storehouses: list = field(default_factory=list)

    # Ссылки
    domrf_link: str = ""
    dev_link: str = ""


def _complex_key(city: str, complex_name: str) -> tuple:
    """Ключ на уровне ЖК (без корпуса)."""
    return (_norm(city), _norm(_apply_alias(complex_name)))


def _aggregate(
    domrf_apts: list[dict],
    domrf_stores: list[dict],
    dev_apts: list[dict],
    dev_stores: list[dict],
    object_infos: list[ObjectInfo],
) -> list[BuildingAgg]:
    """
    Агрегация:
    - Данные дом.рф (квартиры + кладовки) → на уровне КОМПЛЕКСА (город + ЖК),
      т.к. building = API-формат ("1||подъезд 2"), не совпадает с застройщиком.
    - Данные застройщика → по корпусам.
    - Строки берутся из конфига domrf_apartments.yaml (уникальные корпуса)
      + корпуса застройщиков.
    - Каждый корпус получает данные дом.рф от своего ЖК.
    """

    oi_by_id = {oi.object_id: oi for oi in object_infos}
    domrf_config = _load_domrf_config()

    # ─── Шаг 1: Агрегация дом.рф ───
    # Группируем по (city, complex, building_base) для per-building подсчёта.
    # building_base — часть до "||" (напр. "3.1||подъезд 1" → "3.1").

    def _building_base(bld: str) -> str:
        """Извлечь корпус без подъезда: '3.1||подъезд 1' → '3.1'."""
        return bld.split("||")[0].strip() if "||" in bld else bld.strip()

    @dataclass
    class ComplexAgg:
        apt_count: int = 0
        store_count: int = 0
        rooms_count: dict = field(default_factory=dict)
        avg_area: dict = field(default_factory=dict)
        avg_non_living: dict = field(default_factory=dict)
        storehouses: list = field(default_factory=list)

    # Маппинг конфигурационных корпусов: object_id → building из конфига
    # Используется для привязки DomRF-данных (building "1" из БД) к корпусу "М1/ПК-1"
    config_bld_by_objid: dict[int, str] = {}
    # Обратный маппинг: (norm_complex, norm_api_building) → norm_config_building
    # Нужен для старых данных в БД, где building_override не был применён
    api_bld_to_config: dict[tuple, str] = {}
    for link in domrf_config.get("links", []):
        obj_id = link.get("object_id")
        bld = link.get("building", "")
        cn = link.get("complex_name", "")
        if obj_id and bld:
            config_bld_by_objid[obj_id] = bld

    def _resolve_building(raw_building: str, obj_id: int | None) -> str:
        """Привести building из БД к нормализованному имени config-корпуса.

        Приоритет:
        1. object_id → config_building (через маппинг из конфига)
        2. Если building содержит буквы — уже конфигурационное имя
        3. Fallback: API-формат как есть
        """
        # Через object_id — самый надёжный способ
        if obj_id and obj_id in config_bld_by_objid:
            return config_bld_by_objid[obj_id]
        bld_base = _building_base(raw_building)
        return bld_base

    # Per-building агрегация квартир дом.рф
    # Ключ: (norm_city, norm_complex, norm_building)
    # Используем _resolve_building для ремапа API building → config building
    domrf_apt_by_building: dict[tuple, list[dict]] = defaultdict(list)
    for apt in domrf_apts:
        ck = _complex_key(apt["city"], apt["complex_name"])
        bld_resolved = _resolve_building(apt.get("building", ""), apt.get("object_id"))
        bk = (ck[0], ck[1], _norm(bld_resolved))
        domrf_apt_by_building[bk].append(apt)

    # Per-building агрегация кладовок дом.рф
    domrf_store_by_building: dict[tuple, list[dict]] = defaultdict(list)
    for st in domrf_stores:
        ck = _complex_key(st["city"], st["complex_name"])
        bld_resolved = _resolve_building(st.get("building", ""), st.get("object_id"))
        bk = (ck[0], ck[1], _norm(bld_resolved))
        domrf_store_by_building[bk].append(st)

    # Расчёт агрегатов per-building
    building_agg_data: dict[tuple, ComplexAgg] = {}

    for bk, apts in domrf_apt_by_building.items():
        ca = ComplexAgg(apt_count=len(apts))

        rooms_groups: dict[int, list[dict]] = defaultdict(list)
        for a in apts:
            rooms_groups[a["rooms"]].append(a)

        for rooms in range(5):
            group = rooms_groups.get(rooms, [])
            ca.rooms_count[rooms] = len(group)

            areas = [a["area"] for a in group if a.get("area")]
            ca.avg_area[rooms] = round(sum(areas) / len(areas), 1) if areas else 0

            livings = [a["living_area"] for a in group if a.get("living_area")]
            if livings and areas:
                avg_total = sum(areas) / len(areas)
                avg_living = sum(livings) / len(livings)
                ca.avg_non_living[rooms] = round(avg_total - avg_living, 1)
            else:
                ca.avg_non_living[rooms] = 0

        building_agg_data[bk] = ca

    for bk, stores in domrf_store_by_building.items():
        if bk not in building_agg_data:
            building_agg_data[bk] = ComplexAgg()
        building_agg_data[bk].store_count = len(stores)
        building_agg_data[bk].storehouses = stores

    # Также сохраняем complex-level агрегацию (для fallback)
    domrf_store_by_complex: dict[tuple, list[dict]] = defaultdict(list)
    for st in domrf_stores:
        ck = _complex_key(st["city"], st["complex_name"])
        domrf_store_by_complex[ck].append(st)

    # complex-level: объединяем ВСЕ квартиры по комплексу для fallback квартирографии
    domrf_apt_by_complex: dict[tuple, list[dict]] = defaultdict(list)
    for apt in domrf_apts:
        ck = _complex_key(apt["city"], apt["complex_name"])
        domrf_apt_by_complex[ck].append(apt)

    complex_agg: dict[tuple, ComplexAgg] = {}
    for bk, ba in building_agg_data.items():
        ck = (bk[0], bk[1])
        if ck not in complex_agg:
            complex_agg[ck] = ComplexAgg()
        ca = complex_agg[ck]
        ca.apt_count += ba.apt_count
        ca.store_count += ba.store_count
        ca.storehouses.extend(ba.storehouses)

    # Вычисляем квартирографию на complex-level (для fallback когда per-building не матчится)
    for ck, apts in domrf_apt_by_complex.items():
        if ck not in complex_agg:
            complex_agg[ck] = ComplexAgg(apt_count=len(apts))
        ca = complex_agg[ck]
        if not ca.apt_count:
            ca.apt_count = len(apts)

        rooms_groups: dict[int, list[dict]] = defaultdict(list)
        for a in apts:
            rooms_groups[a["rooms"]].append(a)

        for rooms in range(5):
            group = rooms_groups.get(rooms, [])
            ca.rooms_count[rooms] = len(group)
            areas = [a["area"] for a in group if a.get("area")]
            ca.avg_area[rooms] = round(sum(areas) / len(areas), 1) if areas else 0
            livings = [a["living_area"] for a in group if a.get("living_area")]
            if livings and areas:
                avg_total = sum(areas) / len(areas)
                avg_living = sum(livings) / len(livings)
                ca.avg_non_living[rooms] = round(avg_total - avg_living, 1)
            else:
                ca.avg_non_living[rooms] = 0

    for ck, stores in domrf_store_by_complex.items():
        if ck not in complex_agg:
            complex_agg[ck] = ComplexAgg()
            complex_agg[ck].store_count = len(stores)
            complex_agg[ck].storehouses = stores

    # ─── Шаг 2: Определяем список корпусов (строк) ───
    # Источники: конфиг domrf + застройщики
    buildings: dict[tuple, BuildingAgg] = {}

    def _is_better_name(new: str, old: str) -> bool:
        """Человекочитаемое имя лучше ключа (pik → ПИК, akbarsdom → Ак Барс Дом)."""
        if not new:
            return False
        if not old:
            return True
        # Кириллица или заглавные буквы/пробелы → лучше
        has_cyrillic = bool(re.search(r'[а-яА-ЯёЁ]', new))
        has_space = ' ' in new
        old_is_key = old.isascii() and old.islower()
        return (has_cyrillic or has_space) and old_is_key

    def _get_or_create(city, complex_name, building, developer="") -> BuildingAgg:
        key = _match_key(city, complex_name, building)
        if key not in buildings:
            buildings[key] = BuildingAgg(
                city=city, complex_name=complex_name, building=building,
                developer=developer,
            )
        b = buildings[key]
        if _is_better_name(developer, b.developer):
            b.developer = developer
        elif developer and not b.developer:
            b.developer = developer
        # Предпочитаем более полное имя ЖК (Нокса-Парк лучше Нокса парк? нет, берём первое)
        return b

    # Корпуса из конфига domrf (нормальные названия: М1/ПК-1, Корпус 1, etc.)
    config_complexes: set[tuple] = set()  # для трекинга комплексов с конфигом
    for link in domrf_config.get("links", []):
        city = link.get("city", "")
        cn = link.get("complex_name", "")
        bld = link.get("building", "")
        dev = link.get("developer", "")
        obj_id = link.get("object_id")
        if not cn:
            continue

        b = _get_or_create(city, cn, bld, developer=dev)
        config_complexes.add(_complex_key(city, cn))

        if obj_id:
            if obj_id not in b.object_ids:
                b.object_ids.append(obj_id)
            b.domrf_link = (
                f"{DOMRF_BASE}/сервисы/каталог-новостроек/объект/{obj_id}"
            )
            oi = oi_by_id.get(obj_id)
            if oi:
                if not b.commissioning:
                    b.commissioning = oi.commissioning
                    b.days_until = _days_until(oi.commissioning)
                if oi.developer:
                    b.developer = oi.developer
                # НЕ используем ObjectInfo.total_apartments как domrf_apt_count —
                # это ненадёжный источник (может быть per-permit, не per-building).
                # domrf_apt_count будет заполнен из реальных данных в Шаге 3.

    # Корпуса из застройщиков
    # Логика: если корпус застройщика СОВПАДАЕТ с DomRF корпусом (после нормализации)
    # → привязать данные к существующей строке. Если НЕ совпадает → создать новую строку.
    # Кладовки застройщика также считаются на уровне комплекса для "остаток".

    # Индекс DomRF корпусов по комплексу для быстрого поиска
    config_buildings_by_complex: dict[tuple, dict[str, tuple]] = defaultdict(dict)
    for key, b in buildings.items():
        ck = _complex_key(b.city, b.complex_name)
        if ck in config_complexes:
            config_buildings_by_complex[ck][_norm(b.building)] = key

    complex_dev_stores: dict[tuple, list[dict]] = defaultdict(list)
    # Кладовки застройщика без корпуса — для назначения позже
    complex_dev_storehouses: dict[tuple, list[dict]] = defaultdict(list)

    for st in dev_stores:
        cn = _apply_alias(st["complex_name"])
        ck = _complex_key(st["city"], cn)

        if ck in config_complexes:
            complex_dev_stores[ck].append(st)
            # Привязать к совпадающему корпусу (после alias + нормализации)
            # Используем _building_base, чтобы "1||секция 1" → "1" совпало с конфиг "1"
            resolved_bld = _apply_building_alias(cn, st["building"])
            norm_bld = _norm(_building_base(resolved_bld))
            matched_key = config_buildings_by_complex.get(ck, {}).get(norm_bld)
            if matched_key and matched_key in buildings:
                buildings[matched_key].dev_storehouses.append(st)
            elif not norm_bld:
                # Пустой корпус (напр. smu88) — сохранить на уровне комплекса
                complex_dev_storehouses[ck].append(st)
            else:
                # Корпус не совпал — создать новую строку
                b = _get_or_create(st["city"], cn, st["building"],
                                   developer=st.get("site", ""))
                b.dev_storehouses.append(st)
        else:
            b = _get_or_create(st["city"], cn, st["building"],
                               developer=st.get("site", ""))
            b.dev_storehouses.append(st)

    # "Остаток кладовок" — считаем из ВСЕХ dev-кладовок по комплексу (включая пустой building)
    # Это присвоение будет перезаписано ниже — финальное присвоение в конце

    dev_apt_groups: dict[tuple, list[dict]] = defaultdict(list)
    for apt in dev_apts:
        cn = _apply_alias(apt["complex_name"])
        ck = _complex_key(apt["city"], cn)

        if ck in config_complexes:
            # Проверяем: совпадает ли корпус (после alias + нормализации)?
            # Используем _building_base, чтобы "УБ-1||секция 1" → "УБ-1" совпало с конфиг "УБ1"
            resolved_bld = _apply_building_alias(cn, apt["building"])
            norm_bld = _norm(_building_base(resolved_bld))
            matched_key = config_buildings_by_complex.get(ck, {}).get(norm_bld)
            if matched_key:
                dev_apt_groups[matched_key].append(apt)
            else:
                # Уникальный корпус застройщика — создать строку, если нет ||
                if "||" not in apt["building"] and apt["building"]:
                    key = _match_key(apt["city"], cn, apt["building"])
                    dev_apt_groups[key].append(apt)
                    _get_or_create(apt["city"], cn, apt["building"],
                                   developer=apt.get("site", ""))
                else:
                    # Секции/подъезды — агрегировать на уровне комплекса
                    key = ("_complex_dev", ck[0], ck[1])
                    dev_apt_groups[key].append(apt)
        else:
            key = _match_key(apt["city"], cn, apt["building"])
            dev_apt_groups[key].append(apt)
            _get_or_create(apt["city"], cn, apt["building"],
                           developer=apt.get("site", ""))

    for key, apts in dev_apt_groups.items():
        avg_ppm = _calc_priority_ppm(apts)

        if key[0] == "_complex_dev":
            ck = (key[1], key[2])
            for b in buildings.values():
                if _complex_key(b.city, b.complex_name) == ck:
                    b.dev_avg_apt_ppm = avg_ppm
        elif key in buildings:
            buildings[key].dev_avg_apt_ppm = avg_ppm

    # ─── Шаг 3: Привязать квартирографию и кладовки дом.рф к корпусам ───
    # domrf_apt_count заполняется ТОЛЬКО из реальных спарсенных данных.
    for b in buildings.values():
        ck = _complex_key(b.city, b.complex_name)
        bk = (ck[0], ck[1], _norm(b.building))
        ba = building_agg_data.get(bk)

        if ba:
            # Per-building данные найдены — идеальный вариант
            b.domrf_apt_count = ba.apt_count
            if ba.store_count:
                b.domrf_store_count = ba.store_count
            b.rooms_count = dict(ba.rooms_count)
            b.avg_area = dict(ba.avg_area)
            b.avg_non_living = dict(ba.avg_non_living)
        else:
            # Per-building не найдено — попробовать через object_id маппинг.
            # Если у корпуса есть object_ids, собрать данные из всех building_agg_data,
            # привязанных к этим object_ids.
            matched_any = False
            if b.object_ids and config_bld_by_objid:
                # Собираем ВСЕ building_agg_data записи для данного комплекса,
                # и проверяем, нет ли данных с API-форматом building
                for other_bk, other_ba in building_agg_data.items():
                    if other_bk[0] == ck[0] and other_bk[1] == ck[1]:
                        # Это тот же комплекс — проверяем, не подходит ли этот building
                        # API building "1" может соответствовать config "М1/ПК-1"
                        # К сожалению, без object_id в БД мы не можем точно определить.
                        # Пока оставляем fallback на complex_agg.
                        pass

            # Fallback: complex-level квартирография и средние площади.
            # НО: domrf_apt_count и domrf_store_count НЕ берём из complex_agg,
            # т.к. это привело бы к присвоению данных ВСЕГО ЖК одному корпусу.
            ca = complex_agg.get(ck)
            if ca:
                # domrf_apt_count: используем complex-level ТОЛЬКО если в комплексе
                # ровно один корпус (тогда complex == building)
                num_buildings_in_complex = sum(
                    1 for k, ob in buildings.items()
                    if _complex_key(ob.city, ob.complex_name) == ck
                )
                if num_buildings_in_complex == 1:
                    if not b.rooms_count:
                        b.rooms_count = dict(ca.rooms_count)
                    if not b.avg_area:
                        b.avg_area = dict(ca.avg_area)
                    if not b.avg_non_living:
                        b.avg_non_living = dict(ca.avg_non_living)
                    if not b.domrf_apt_count:
                        b.domrf_apt_count = ca.apt_count
                    if not b.domrf_store_count:
                        b.domrf_store_count = ca.store_count
                else:
                    # Несколько корпусов: complex-level квартирографию
                    # присваиваем ТОЛЬКО если есть per-building apt_count
                    # (иначе получим данные ВСЕГО ЖК в строке одного корпуса)
                    if b.domrf_apt_count:
                        if not b.rooms_count:
                            b.rooms_count = dict(ca.rooms_count)
                        if not b.avg_area:
                            b.avg_area = dict(ca.avg_area)
                        if not b.avg_non_living:
                            b.avg_non_living = dict(ca.avg_non_living)
                    elif ca.apt_count:
                        logger.warning(
                            "  ⚠ %s / %s — per-building данные не найдены, "
                            "complex-level apt_count=%d НЕ присвоен (несколько корпусов)",
                            b.complex_name, b.building, ca.apt_count,
                        )

    # ─── Шаг 3.5: Для ЖК БЕЗ DomRF — агрегировать из данных застройщика ───
    # Квартирография, средняя площадь, кол-во квартир из dev apartments
    for key, apts in dev_apt_groups.items():
        if key[0] == "_complex_dev":
            continue  # это DomRF-ЖК, уже обработаны
        if key not in buildings:
            continue
        b = buildings[key]
        ck = _complex_key(b.city, b.complex_name)
        if ck in config_complexes:
            continue  # DomRF-ЖК, данные уже из DomRF

        # Агрегируем квартиры застройщика по корпусу
        b.domrf_apt_count = len(apts)  # используем как "всего квартир"
        rooms_groups: dict[int, list[dict]] = defaultdict(list)
        for a in apts:
            rooms_groups[a["rooms"]].append(a)

        for rooms in range(5):
            group = rooms_groups.get(rooms, [])
            b.rooms_count[rooms] = len(group)
            areas = [a["area"] for a in group if a.get("area")]
            b.avg_area[rooms] = round(sum(areas) / len(areas), 1) if areas else 0
            # Нежилая площадь — только из DomRF, тут не заполняем

    # ─── Шаг 4: Кладовки дом.рф fallback ───
    # Назначается ПОСЛЕ Шага 5b (удаления мусорных строк), чтобы не создавать
    # лишние строки. Выполняется в Шаге 4b ниже.

    # ─── Шаг 5a: Сохранить данные ДОМ.РФ ПЕРЕД удалением мусорных строк ───
    # Ссылки, commissioning, object_ids и другие DomRF-данные могут быть
    # только у конфиг-строк (в т.ч. с пустым building). Сохраняем их
    # на уровне комплекса ДО удаления, чтобы потом пропагировать.
    complex_domrf_link: dict[tuple, str] = {}
    complex_commissioning: dict[tuple, tuple] = {}  # (commissioning, days_until)
    complex_object_ids: dict[tuple, list] = defaultdict(list)

    for b in buildings.values():
        ck = _complex_key(b.city, b.complex_name)
        if b.domrf_link and ck not in complex_domrf_link:
            complex_domrf_link[ck] = b.domrf_link
        if b.commissioning and ck not in complex_commissioning:
            complex_commissioning[ck] = (b.commissioning, b.days_until)
        for oid in b.object_ids:
            if oid not in complex_object_ids[ck]:
                complex_object_ids[ck].append(oid)

    # ─── Шаг 5b: Убрать мусорные строки ───
    # "1||подъезд 1" → мусор если у комплекса есть нормальные корпуса из конфига
    # Пустой building → мусор если у ЖК есть другие корпуса
    keys_to_remove = []
    for key, b in buildings.items():
        ck = _complex_key(b.city, b.complex_name)
        if "||" in b.building and ck in config_complexes:
            keys_to_remove.append(key)
        elif not b.building.strip():
            # Пустой корпус — убрать если у ЖК есть другие корпуса
            has_others = any(
                k != key and _complex_key(ob.city, ob.complex_name) == ck
                and ob.building.strip()
                for k, ob in buildings.items()
            )
            if has_others:
                keys_to_remove.append(key)
    for key in keys_to_remove:
        del buildings[key]

    # ─── Шаг 4b: Кладовки ДОМ.РФ и комплекс-level dev кладовки ───
    # Выполняется ПОСЛЕ удаления мусорных строк.

    # Пересчитать complex_has_dev после Step 5b
    complex_has_dev: dict[tuple, bool] = defaultdict(bool)
    for b in buildings.values():
        if b.dev_storehouses:
            complex_has_dev[_complex_key(b.city, b.complex_name)] = True

    # Назначить комплекс-level кладовки застройщика (smu88 с пустым building)
    for ck, stores in complex_dev_storehouses.items():
        ck_buildings = [b for b in buildings.values()
                        if _complex_key(b.city, b.complex_name) == ck]
        if not ck_buildings:
            continue
        if len(ck_buildings) == 1:
            ck_buildings[0].dev_storehouses.extend(stores)
        else:
            # Несколько корпусов — добавить к первому
            # (лучше не дублировать по всем)
            ck_buildings[0].dev_storehouses.extend(stores)
        complex_has_dev[ck] = True

    # Назначить кладовки ДОМ.РФ для комплексов без dev кладовок.
    # Распределяем per-building через _resolve_building (object_id → config building).
    for ck, ca in complex_agg.items():
        if complex_has_dev.get(ck, False) or not ca.storehouses:
            continue
        # Группируем кладовки по resolved building
        by_bld: dict[str, list[dict]] = defaultdict(list)
        for st in ca.storehouses:
            bld_resolved = _resolve_building(st.get("building", ""), st.get("object_id"))
            by_bld[_norm(bld_resolved)].append(st)

        ck_buildings = [b for b in buildings.values()
                        if _complex_key(b.city, b.complex_name) == ck]
        ck_bld_map = {_norm(b.building): b for b in ck_buildings}

        for bld_key, bld_stores in by_bld.items():
            if bld_key in ck_bld_map:
                b = ck_bld_map[bld_key]
                b.domrf_storehouses = bld_stores
                b.domrf_store_count = len(bld_stores)
            else:
                # Нет конфиг-корпуса — создать строку
                city = bld_stores[0]["city"]
                cn = bld_stores[0]["complex_name"]
                bld_name = _building_base(bld_stores[0]["building"])
                b = _get_or_create(city, cn, bld_name)
                b.domrf_storehouses = bld_stores
                b.domrf_store_count = len(bld_stores)

    # ─── Шаг 6: Пропагация данных на уровне комплекса ───
    # Средняя цена м² застройщика — на уровне комплекса
    complex_dev_ppm: dict[tuple, float] = {}
    for b in buildings.values():
        if b.dev_avg_apt_ppm:
            ck = _complex_key(b.city, b.complex_name)
            if ck not in complex_dev_ppm:
                complex_dev_ppm[ck] = b.dev_avg_apt_ppm

    # Собрать все квартиры застройщика по комплексу для пересчёта ppm
    complex_all_apts: dict[tuple, list[dict]] = defaultdict(list)
    for key, apts in dev_apt_groups.items():
        b_tmp = buildings.get(key)
        if b_tmp:
            ck = _complex_key(b_tmp.city, b_tmp.complex_name)
            complex_all_apts[ck].extend(apts)

    for ck, apts in complex_all_apts.items():
        ppm = _calc_priority_ppm(apts)
        if ppm:
            complex_dev_ppm[ck] = ppm

    # Применить complex-level данные ко всем корпусам
    # (включая ссылки и commissioning, сохранённые в Шаге 5a)
    for b in buildings.values():
        ck = _complex_key(b.city, b.complex_name)
        if not b.dev_avg_apt_ppm and ck in complex_dev_ppm:
            b.dev_avg_apt_ppm = complex_dev_ppm[ck]
        if not b.domrf_link and ck in complex_domrf_link:
            b.domrf_link = complex_domrf_link[ck]
        if not b.commissioning and ck in complex_commissioning:
            b.commissioning, b.days_until = complex_commissioning[ck]
        if not b.object_ids and ck in complex_object_ids:
            b.object_ids = list(complex_object_ids[ck])

    # П.3: для корпусов с ppm=0 — взять от соседнего корпуса
    _apply_neighbor_ppm(list(buildings.values()))

    # dev_store_count — из реальных привязанных кладовок
    for b in buildings.values():
        if b.dev_storehouses:
            b.dev_store_count = len(b.dev_storehouses)

    # domrf_store_count — complex-level fallback ТОЛЬКО для ЖК с одним корпусом.
    # Для multi-building ЖК присвоение всех кладовок одному корпусу — ошибка.
    complex_domrf_store_count: dict[tuple, int] = {}
    for ck_key, ca in complex_agg.items():
        if ca.store_count:
            complex_domrf_store_count[ck_key] = ca.store_count
    for b in buildings.values():
        if not b.domrf_store_count:
            ck = _complex_key(b.city, b.complex_name)
            if ck in complex_domrf_store_count:
                num_buildings = sum(
                    1 for ob in buildings.values()
                    if _complex_key(ob.city, ob.complex_name) == ck
                )
                if num_buildings == 1:
                    b.domrf_store_count = complex_domrf_store_count[ck]

    # Нормализация имён застройщиков
    for b in buildings.values():
        if b.developer in DEVELOPER_NAMES:
            b.developer = DEVELOPER_NAMES[b.developer]

    # ─── Валидация агрегации ───
    for b in buildings.values():
        if b.domrf_apt_count > 2000:
            logger.warning(
                "  ⚠ ВАЛИДАЦИЯ: %s / %s — %d квартир (> 2000, подозрительно)",
                b.complex_name, b.building, b.domrf_apt_count,
            )
        if b.domrf_store_count > 200:
            logger.warning(
                "  ⚠ ВАЛИДАЦИЯ: %s / %s — %d кладовок ДОМ.РФ (> 200)",
                b.complex_name, b.building, b.domrf_store_count,
            )
        total_rooms = sum(b.rooms_count.values()) if b.rooms_count else 0
        if b.domrf_apt_count > 0 and total_rooms == 0:
            logger.warning(
                "  ⚠ ВАЛИДАЦИЯ: %s / %s — domrf_apt_count=%d, но квартирография пуста!",
                b.complex_name, b.building, b.domrf_apt_count,
            )
        if total_rooms > 0 and b.domrf_apt_count == 0:
            logger.warning(
                "  ⚠ ВАЛИДАЦИЯ: %s / %s — квартирография=%d, но domrf_apt_count=0",
                b.complex_name, b.building, total_rooms,
            )

    # Сортировка
    result = sorted(buildings.values(),
                    key=lambda b: (b.city, b.developer, b.complex_name, b.building))
    return result


def _load_domrf_config() -> dict:
    path = CONFIGS_DIR / "domrf_apartments.yaml"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ─── Скоринг ─────────────────────────────────────────────

def calc_first_stage(bd: BuildingAgg, scoring: dict) -> int:
    """Первый этап: критерии 1–10 (7-10 зависят от ручных данных → пока 0)."""
    pts = 0

    # 1. Срок ввода (дни)
    if bd.days_until is not None:
        if bd.days_until < 0:
            # Дом уже сдан: 20 баллов, убывает на 20/180 за каждый день в прошлом
            pts += max(0, 20 + bd.days_until * (20 / 180))
        elif bd.days_until < 365:
            pts += 20 - bd.days_until * (10 / 365)
        elif bd.days_until < 730:
            pts += 10 - (bd.days_until - 365) * (5 / 365)
        # else: 0

    # 2. Соотношение квартир / кладовых
    store_total = bd.dev_store_count or bd.domrf_store_count
    if store_total:
        ratio = bd.domrf_apt_count / store_total
        pts += _score_by_thresholds(ratio, scoring["apartment_storehouse_ratio"])

    # 3. Соотношение цены м²
    storehouses = bd.dev_storehouses or bd.domrf_storehouses
    store_ppms = [s["price_per_meter"] for s in storehouses if s.get("price_per_meter")]
    if store_ppms and bd.dev_avg_apt_ppm:
        price_ratio = bd.dev_avg_apt_ppm / (sum(store_ppms) / len(store_ppms))
        pts += _score_by_thresholds(price_ratio, scoring["price_ratio"])

    # 4. Квартирография
    # Используем сумму rooms_count как total (реальное кол-во спарсенных квартир),
    # а не domrf_apt_count (может быть из ObjectInfo / complex-level)
    total = sum(bd.rooms_count.values()) if bd.rooms_count else 0
    if not total:
        total = bd.domrf_apt_count if bd.domrf_apt_count else 1
    kvart = scoring["kvartirografia"]
    s1k = (bd.rooms_count.get(0, 0) + bd.rooms_count.get(1, 0)) / total * 100
    two_k = bd.rooms_count.get(2, 0) / total * 100
    three_4k = (bd.rooms_count.get(3, 0) + bd.rooms_count.get(4, 0)) / total * 100
    pts += _score_by_thresholds(s1k, kvart["studio_1k"])
    pts += _score_by_thresholds(two_k, kvart["two_k"])
    pts += _score_by_thresholds(three_4k, kvart["three_4k"])

    # 5. Средняя площадь
    area_cfg = scoring["avg_area"]
    room_keys = [(0, "studio"), (1, "one_k"), (2, "two_k"), (3, "three_k"), (4, "four_k")]
    for rooms, cfg_key in room_keys:
        avg = bd.avg_area.get(rooms, 0)
        if avg > 0:
            pts += _score_by_thresholds(avg, area_cfg[cfg_key])

    # 6. Средняя нежилая площадь
    nl_cfg = scoring["avg_non_living"]
    for rooms, cfg_key in room_keys:
        avg_nl = bd.avg_non_living.get(rooms, 0)
        if avg_nl > 0:
            pts += _score_by_thresholds(avg_nl, nl_cfg[cfg_key])

    # 7-10 — ручные, пока 0
    return pts


def calc_second_stage(store: dict, scoring: dict) -> int:
    """Второй этап: критерии 11-13 (площадь, стоимость, цена/м² кладовки)."""
    pts = 0
    area = store.get("area") or 0
    price = store.get("price") or 0
    ppm = store.get("price_per_meter") or 0

    if area > 0:
        pts += _score_by_thresholds(area, scoring["storehouse_area"])
    if price > 0:
        pts += _score_by_thresholds(price, scoring["storehouse_price"])
    if ppm > 0:
        pts += _score_by_thresholds(ppm, scoring["storehouse_ppm"])
    return pts


# ─── Утилиты xlsx ────────────────────────────────────────

def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="4472C4"))


def _apply_segment_border(ws, row: int, max_col: int):
    """Применить жирную нижнюю границу ко всей строке (конец сегмента)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        # Сохранить вертикальные разделители, добавить горизонтальную
        old = cell.border
        cell.border = Border(
            left=old.left, right=old.right,
            bottom=Side(style="medium", color="4472C4"),
        )


def _apply_vertical_dividers(ws, start_row: int, end_row: int,
                             right_cols: set, left_cols: set):
    """Применить сплошные вертикальные разделители ко ВСЕМ строкам диапазона."""
    thick = Side(style="medium", color="4472C4")
    for row in range(start_row, end_row + 1):
        for col in right_cols:
            cell = ws.cell(row=row, column=col)
            old = cell.border
            cell.border = Border(
                left=old.left, right=thick,
                top=old.top, bottom=old.bottom,
            )
        for col in left_cols:
            cell = ws.cell(row=row, column=col)
            old = cell.border
            cell.border = Border(
                left=thick, right=old.right,
                top=old.top, bottom=old.bottom,
            )


# ─── Цвета условного форматирования ──────────────────────

# 6 уровней: синий → зелёный → светло-зелёный → жёлтый → оранжевый → коричневый
CF_BLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CF_GREEN = PatternFill(start_color="6B9F3B", end_color="6B9F3B", fill_type="solid")
CF_LTGREEN = PatternFill(start_color="C6DFAB", end_color="C6DFAB", fill_type="solid")
CF_YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
CF_ORANGE = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
CF_BROWN = PatternFill(start_color="8B6914", end_color="8B6914", fill_type="solid")

CF_BLUE_FONT = Font(color="FFFFFF", bold=True)
CF_GREEN_FONT = Font(color="FFFFFF", bold=True)
CF_BROWN_FONT = Font(color="FFFFFF", bold=True)
CF_DEFAULT_FONT = Font(color="000000")


def _add_conditional_formatting(ws, col_letter: str, last_row: int, rules: list):
    """
    Добавить условное форматирование к столбцу.
    rules: [(operator, formula_or_values, fill, font), ...]
    Порядок: от САМОГО СПЕЦИФИЧНОГО к общему (Excel применяет первое совпадение).
    """
    cell_range = f"{col_letter}3:{col_letter}{last_row}"
    for rule_def in rules:
        op, vals, fill, font = rule_def
        if op == "formula":
            ws.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=vals, fill=fill, font=font)
            )
        else:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator=op, formula=vals, fill=fill, font=font)
            )


def _apply_color_rules(ws, last_row: int, col_map: dict):
    """
    Применить цветовые правила ко всем столбцам.
    col_map: {"E": "days", "M": "ratio_apt_store", ...}
    """
    for col_letter, rule_type in col_map.items():
        rules = _get_color_rules(rule_type, col_letter)
        _add_conditional_formatting(ws, col_letter, last_row, rules)


def _get_color_rules(rule_type: str, col: str) -> list:
    """Вернуть правила условного форматирования для типа столбца."""

    if rule_type == "days":
        # Дней до сдачи: <183 синий, 183-365 зелёный, 365-547 свет-зел, 547-730 жёлтый, >730 коричневый
        return [
            ("lessThan", ["183"], CF_BLUE, CF_BLUE_FONT),
            ("lessThan", ["365"], CF_GREEN, CF_GREEN_FONT),
            ("lessThan", ["547"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("lessThan", ["730"], CF_YELLOW, CF_DEFAULT_FONT),
            ("greaterThanOrEqual", ["730"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "ratio_apt_store":
        # Соотнош. квартир/кладовых: >13 синий, 10-13 зелёный, 7-10 свет-зел, 5-7 жёлтый, <5 коричневый
        return [
            ("greaterThan", ["13"], CF_BLUE, CF_BLUE_FONT),
            ("greaterThanOrEqual", ["10"], CF_GREEN, CF_GREEN_FONT),
            ("greaterThanOrEqual", ["7"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("greaterThanOrEqual", ["5"], CF_YELLOW, CF_DEFAULT_FONT),
            ("lessThan", ["5"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "ratio_price":
        # Соотнош. цены м²: >5 синий, 4-5 зелёный, 3-4 свет-зел, <3 коричневый
        return [
            ("greaterThan", ["5"], CF_BLUE, CF_BLUE_FONT),
            ("greaterThanOrEqual", ["4"], CF_GREEN, CF_GREEN_FONT),
            ("greaterThanOrEqual", ["3"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("lessThan", ["3"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "ratio_balcony":
        # Соотнош. квартир к балконам: >2 синий, 1.5-2 зелёный, 1-1.5 жёлтый, <1 коричневый
        return [
            ("greaterThan", ["2"], CF_BLUE, CF_BLUE_FONT),
            ("greaterThanOrEqual", ["1.5"], CF_GREEN, CF_GREEN_FONT),
            ("greaterThanOrEqual", ["1"], CF_YELLOW, CF_DEFAULT_FONT),
            ("lessThan", ["1"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "store_area":
        # Площадь кладовки: >6 синий, 5-6 зелёный, 4-5 свет-зел, 3-4 жёлтый, <2.5 коричневый
        return [
            ("greaterThanOrEqual", ["6"], CF_BLUE, CF_BLUE_FONT),
            ("greaterThanOrEqual", ["5"], CF_GREEN, CF_GREEN_FONT),
            ("greaterThanOrEqual", ["4"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("greaterThanOrEqual", ["3"], CF_YELLOW, CF_DEFAULT_FONT),
            ("lessThan", ["2.5"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "store_price":
        # Стоимость: <300к синий, 300-350 зелёный, 350-400 свет-зел, 400-500 жёлтый, 500-650 оранж, >650 коричневый
        return [
            ("lessThan", ["300000"], CF_BLUE, CF_BLUE_FONT),
            ("lessThan", ["350000"], CF_GREEN, CF_GREEN_FONT),
            ("lessThan", ["400000"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("lessThan", ["500000"], CF_YELLOW, CF_DEFAULT_FONT),
            ("lessThan", ["650000"], CF_ORANGE, CF_DEFAULT_FONT),
            ("greaterThanOrEqual", ["650000"], CF_BROWN, CF_BROWN_FONT),
        ]

    elif rule_type == "store_ppm":
        # Цена/м²: <65к синий, 65-70 зелёный, 70-85 свет-зел, 85-100 жёлтый, >100 коричневый
        return [
            ("lessThan", ["65000"], CF_BLUE, CF_BLUE_FONT),
            ("lessThan", ["70000"], CF_GREEN, CF_GREEN_FONT),
            ("lessThan", ["85000"], CF_LTGREEN, CF_DEFAULT_FONT),
            ("lessThan", ["100000"], CF_YELLOW, CF_DEFAULT_FONT),
            ("greaterThanOrEqual", ["100000"], CF_BROWN, CF_BROWN_FONT),
        ]

    return []


def _write_cell(ws, row, col, value, is_manual=False, is_score=False, is_link=False, is_gray=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGN
    # Без рамки по умолчанию — только вертикальные разделители (применяются отдельно)
    if is_manual:
        cell.fill = MANUAL_FILL
    elif is_score:
        cell.fill = SCORE_FILL
    if is_gray:
        cell.font = GRAY_FONT
    if is_link and value:
        cell.font = LINK_FONT
    return cell


# ─── Smart Merge (ручные данные) ─────────────────────────

def _read_manual_jk(old_path: Path) -> dict[tuple, dict]:
    """Прочитать ручные данные листа 'жк' из предыдущего файла."""
    if not old_path.exists():
        return {}
    wb = load_workbook(old_path, data_only=True)
    if "жк" not in wb.sheetnames:
        return {}
    ws = wb["жк"]
    manual = {}
    for r in range(3, ws.max_row + 1):
        city = ws.cell(r, 1).value
        cn = ws.cell(r, 3).value
        bld = ws.cell(r, 4).value
        if not city and not cn:
            continue
        key = _match_key(str(city or ""), str(cn or ""), str(bld or ""))
        data = {}
        # 9=балконы, 10=размер, 11=локация, 12=доступ
        # Если ячейка имеет примечание "автоматически" — это автозаполнение, не ручное
        for col_idx, name in [(9, "balcony_count"), (10, "balcony_size"),
                               (11, "location"), (12, "access")]:
            cell = ws.cell(r, col_idx)
            val = cell.value
            if val is None:
                continue
            # Пропускаем автозаполненные значения (с примечанием)
            comment = cell.comment
            if comment and "автоматически" in str(comment.text or ""):
                continue
            data[name] = val
        if data:
            manual[key] = data
    wb.close()
    return manual


# ─── ЛИСТ 1: кладовки (застройщик) ───────────────────────

STORE_HEADERS = [
    "Город",                                      # 1
    "Застройщик",                                  # 2
    "ЖК",                                          # 3
    "Корпус",                                       # 4
    "Дней до сдачи",                                # 5
    "Соотнош. квартир/кладовых",                    # 6
    "Соотнош. цены м² квартир/кладовых",            # 7
    "Соотнош. квартир к балконам",                   # 8 (формула из жк)
    "Первый этап",                                  # 9 (пока не заполнять)
    "Второй этап",                                  # 10 (пока не заполнять)
    "Общие баллы",                                  # 11 (пока не заполнять)
    "Площадь (м²)",                                 # 12
    "Стоимость (₽)",                                # 13
    "Цена/м² (₽)",                                  # 14
    "Номер кладовой",                               # 15
    "Ссылка",                                       # 16
    "Исх. порядок",                                 # 17
]

STORE_WIDTHS = [12, 16, 22, 14, 12, 18, 20, 20, 12, 12, 12, 10, 14, 14, 12, 12, 10]


def _fill_storehouses_sheet(ws, buildings: list[BuildingAgg]):
    """Лист 1: кладовки (застройщик). Без баллов — они будут добавлены формулами позже."""

    # Подзаголовки групп
    ws.merge_cells("A1:D1")
    ws.cell(1, 1, "Основные данные ЖК").fill = SUBHEADER_FILL
    ws.cell(1, 1).font = SUBHEADER_FONT
    ws.cell(1, 1).alignment = DATA_ALIGN
    ws.merge_cells("I1:K1")
    ws.cell(1, 9, "Баллы").fill = SUBHEADER_FILL
    ws.cell(1, 9).font = SUBHEADER_FONT
    ws.cell(1, 9).alignment = DATA_ALIGN
    ws.merge_cells("L1:Q1")
    ws.cell(1, 12, "Данные кладовой").fill = SUBHEADER_FILL
    ws.cell(1, 12).font = SUBHEADER_FONT
    ws.cell(1, 12).alignment = DATA_ALIGN

    for c, h in enumerate(STORE_HEADERS, 1):
        ws.cell(2, c, h)
    _style_header(ws, 2, len(STORE_HEADERS))
    for c, w in enumerate(STORE_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 3
    order = 1

    for bd in buildings:
        storehouses = bd.dev_storehouses if bd.dev_storehouses else bd.domrf_storehouses
        if not storehouses:
            continue

        # Соотношение цен: ср. цена м² квартиры (застр.) / ср. цена м² кладовки
        store_ppms = [s["price_per_meter"] for s in storehouses if s.get("price_per_meter")]
        avg_store_ppm = sum(store_ppms) / len(store_ppms) if store_ppms else 0
        price_ratio = (
            round(bd.dev_avg_apt_ppm / avg_store_ppm, 1)
            if avg_store_ppm and bd.dev_avg_apt_ppm else None
        )

        jk_row = bd._jk_row if hasattr(bd, '_jk_row') else None

        for store in storehouses:
            has_price = bool(store.get("price"))

            _write_cell(ws, row, 1, bd.city)
            _write_cell(ws, row, 2, bd.developer)
            _write_cell(ws, row, 3, bd.complex_name)
            _write_cell(ws, row, 4, bd.building)
            _write_cell(ws, row, 5, bd.days_until)
            # Соотнош. квартир/кладовых — ссылка на формулу с листа жк
            if jk_row:
                _write_cell(ws, row, 6, f'=жк!M{jk_row}')
            else:
                _write_cell(ws, row, 6, None)
            _write_cell(ws, row, 7, price_ratio)

            # 8: Соотнош. квартир к балконам — формула из листа жк
            if jk_row:
                formula = f'=IFERROR(жк!O{jk_row}/жк!I{jk_row},"")'
                _write_cell(ws, row, 8, formula)
            else:
                _write_cell(ws, row, 8, None)

            # 9: Первый этап — ссылка на формулу с листа жк
            if jk_row:
                _write_cell(ws, row, 9, f'=жк!H{jk_row}')
            else:
                _write_cell(ws, row, 9, 0)

            # 10: Второй этап — формула из модуля scoring.py
            from core.scoring import generate_second_stage_formula
            _write_cell(ws, row, 10, generate_second_stage_formula(row))

            # 11: Общие баллы = первый + второй
            _write_cell(ws, row, 11, f'=ROUND(I{row}+J{row},2)')

            _write_cell(ws, row, 12, store.get("area"))
            _write_cell(ws, row, 13, store.get("price") if has_price else "–")
            _write_cell(ws, row, 14, store.get("price_per_meter") if has_price else "–")
            _write_cell(ws, row, 15, store.get("item_number"))

            # Ссылка — приоритет на карточку конкретной кладовки, fallback на ЖК
            url = store.get("url") or bd.domrf_link
            cell_link = _write_cell(ws, row, 16, "Открыть" if url else None)
            if url:
                cell_link.hyperlink = url
                cell_link.font = LINK_FONT

            _write_cell(ws, row, 17, order, is_gray=True)
            order += 1
            row += 1

    # Жирная линия между сегментами (смена ЖК + корпус)
    prev_key = None
    for r in range(3, row):
        cn = ws.cell(r, 3).value
        bld = ws.cell(r, 4).value
        cur_key = (cn, bld)
        if prev_key is not None and cur_key != prev_key:
            _apply_segment_border(ws, r - 1, len(STORE_HEADERS))
        prev_key = cur_key

    # Вертикальные разделители для листа кладовок
    # Разделители: после корпуса(4), после балконов(8), после баллов(11),
    # после ссылки(16), после исх.порядка(17)
    store_right = {4, 8, 11, 16, 17}
    store_left = {5, 9, 12, 17}
    _apply_vertical_dividers(ws, 2, row - 1, store_right, store_left)

    # Условное форматирование (цвета по значениям)
    # Лист 1: E=5(дни), F=6(кв/клад), G=7(цена м²), H=8(кв/балк), L=12(площадь), M=13(стоимость), N=14(цена/м²)
    _apply_color_rules(ws, row - 1, {
        "E": "days",
        "F": "ratio_apt_store",
        "G": "ratio_price",
        "H": "ratio_balcony",
        "L": "store_area",
        "M": "store_price",
        "N": "store_ppm",
    })

    ws.freeze_panes = "E3"
    if row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(STORE_HEADERS))}{row - 1}"


# ─── ЛИСТ 2: жк ─────────────────────────────────────────

JK_HEADERS = [
    "Город",                                       # 1
    "Застройщик",                                   # 2
    "ЖК",                                           # 3
    "Корпус",                                        # 4
    "Дней до сдачи",                                 # 5
    "Ссылка на ЖК в ДОМ.РФ",                         # 6
    "Ссылка на сайт застройщика",                     # 7
    "Первый этап",                                   # 8 (не заполнять)
    "Кол-во балконов",                               # 9 (РУЧНОЙ)
    "Размер балконов",                               # 10 (РУЧНОЙ)
    "Локация ЖК",                                    # 11 (РУЧНОЙ)
    "Удобство доступа",                              # 12 (РУЧНОЙ)
    "Соотнош. квартир/кладовых (ДОМ.РФ)",             # 13
    "Соотнош. квартир к балконам",                    # 14 (РУЧНОЙ)
    "Кол-во квартир (всего, ДОМ.РФ)",                 # 15
    "Кол-во кладовых (всего, ДОМ.РФ)",                # 16
    "Остаток кладовок (застройщик)",                  # 17
    "Средняя цена м² квартиры (застройщик)",          # 18
    "Студии", "1к", "2к", "3к", "4к+",              # 19-23
    "Ср. площадь (ст.)", "Ср. площадь (1к)",        # 24-28
    "Ср. площадь (2к)", "Ср. площадь (3к)", "Ср. площадь (4к+)",
    "Ср. нежилая (ст.)", "Ср. нежилая (1к)",        # 29-33
    "Ср. нежилая (2к)", "Ср. нежилая (3к)", "Ср. нежилая (4к+)",
    "Исх. порядок",                                  # 34
]

MANUAL_JK_COLS = {9, 10, 11, 12}  # жёлтые (14 теперь формула)
EMPTY_JK_COLS = set()                  # нет пустых столбцов

JK_WIDTHS = (
    [12, 16, 22, 14, 12, 22, 22, 12]       # 1-8
    + [14, 14, 14, 14]                      # 9-12 (manual)
    + [20, 18, 18, 18, 16, 18]              # 13-18
    + [8, 8, 8, 8, 8]                       # 19-23 (kvartirografia)
    + [12, 12, 12, 12, 12]                  # 24-28 (avg area)
    + [12, 12, 12, 12, 12]                  # 29-33 (avg non-living)
    + [10]                                  # 34 (order)
)


def _fill_jk_sheet(
    ws,
    buildings: list[BuildingAgg],
    manual: dict[tuple, dict],
    dev_urls: dict[tuple, str],
    living_areas: dict[tuple, dict],
):
    """Лист 2: жк."""

    # Подзаголовки строка 1
    groups = [
        (1, 4, "Основные данные"),
        (9, 12, "Заполняет заказчик"),
        (13, 18, "Сводные данные"),
        (19, 23, "Квартирография (ДОМ.РФ)"),
        (24, 28, "Средняя площадь (ДОМ.РФ)"),
        (29, 33, "Средняя нежилая площадь (ДОМ.РФ)"),
    ]
    for start, end, title in groups:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, title)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = DATA_ALIGN

    # Заголовки строка 2
    for c, h in enumerate(JK_HEADERS, 1):
        ws.cell(2, c, h)
    _style_header(ws, 2, len(JK_HEADERS))

    for c, w in enumerate(JK_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Data validation для ручных столбцов
    dv_size = DataValidation(type="list", formula1='"–,С,Б,М"', allow_blank=True)
    dv_size.error = "Выберите: С, Б, М или – для сброса"
    dv_loc = DataValidation(type="list", formula1='"–,Центр,Норм,Окраина,Фу"', allow_blank=True)
    dv_loc.error = "Выберите: Центр, Норм, Окраина, Фу или – для сброса"
    dv_access = DataValidation(type="list", formula1='"–,Лифт,Улица,Лестница,Лифт+Улица"', allow_blank=True)
    dv_access.error = "Выберите: Лифт, Улица, Лестница, Лифт+Улица или – для сброса"
    ws.add_data_validation(dv_size)
    ws.add_data_validation(dv_loc)
    ws.add_data_validation(dv_access)

    for i, bd in enumerate(buildings):
        row = i + 3
        order = i + 1

        # Сохраняем номер строки на листе жк для формул на листе кладовок
        bd._jk_row = row

        key = _match_key(bd.city, bd.complex_name, bd.building)
        m = manual.get(key, {})
        ck = _complex_key(bd.city, bd.complex_name)

        # Количество квартир: domrf_apt_count, или sum(rooms_count) как fallback
        apt_total = bd.domrf_apt_count
        if not apt_total and bd.rooms_count:
            apt_total = sum(bd.rooms_count.values())
        # Соотношение квартир/кладовых — динамическая формула =O/P
        apt_store_ratio = f'=IF(OR(O{row}=0,P{row}=0),"",ROUND(O{row}/P{row},1))'

        # Ссылка застройщика из yaml (base_url по имени застройщика)
        dev_url = dev_urls.get(bd.developer)

        # Формулы из модуля scoring.py (единый источник истины)
        from core.scoring import generate_first_stage_formula, generate_balcony_ratio_formula
        balcony_ratio_formula = generate_balcony_ratio_formula(row)
        first_stage_formula = generate_first_stage_formula(row)

        values = {
            1: bd.city,
            2: bd.developer,
            3: bd.complex_name,
            4: bd.building,
            5: bd.days_until,
            6: bd.domrf_link or None,
            7: dev_url or None,
            8: first_stage_formula,
            9: m.get("balcony_count") or apt_total or 0,
            10: m.get("balcony_size") or "С",
            11: m.get("location"),
            12: m.get("access"),
            13: apt_store_ratio,
            14: balcony_ratio_formula,
            15: apt_total or 0,
            16: bd.domrf_store_count or 0,
            17: bd.dev_store_count or 0,
            18: round(bd.dev_avg_apt_ppm) if bd.dev_avg_apt_ppm else 0,
        }

        # 19-23: квартирография
        for rooms in range(5):
            cnt = bd.rooms_count.get(rooms, 0)
            values[19 + rooms] = cnt or 0

        # 24-28: средняя площадь (из дом.рф)
        # Приоритет: per-building данные из xlsx → complex fallback → bd.avg_area
        bk_la = (ck[0], ck[1], _norm(bd.building))
        la_data = living_areas.get(bk_la) or living_areas.get(ck, {})
        for rooms in range(5):
            avg = bd.avg_area.get(rooms, 0)
            # Предпочитаем данные из xlsx (более полные, включают жилую площадь)
            if la_data and rooms in la_data:
                avg = la_data[rooms][0]  # (avg_area, avg_living)
            values[24 + rooms] = round(avg, 1) if avg else 0

        # 29-33: средняя нежилая площадь = общая - жилая (из xlsx дом.рф)
        for rooms in range(5):
            if la_data and rooms in la_data:
                avg_area_val, avg_living_val = la_data[rooms]
                non_living = round(avg_area_val - avg_living_val, 1)
                values[29 + rooms] = non_living if non_living > 0 else 0
            else:
                avg_nl = bd.avg_non_living.get(rooms, 0)
                values[29 + rooms] = round(avg_nl, 1) if avg_nl else 0

        values[34] = order

        for col_idx, val in values.items():
            is_manual = col_idx in MANUAL_JK_COLS
            is_empty = col_idx in EMPTY_JK_COLS
            is_gray = col_idx == 34
            _write_cell(ws, row, col_idx, val, is_manual=is_manual or is_empty, is_gray=is_gray)

        # Ссылки
        if bd.domrf_link:
            cell = ws.cell(row, 6)
            cell.hyperlink = bd.domrf_link
            cell.font = LINK_FONT
            cell.value = "ДОМ.РФ"
        if dev_url:
            cell = ws.cell(row, 7)
            cell.hyperlink = dev_url
            cell.font = LINK_FONT
            cell.value = "Застройщик"

        # Примечания для автозаполненных ячеек
        if not m.get("balcony_count"):
            c = ws.cell(row, 9)
            c.comment = Comment("Значение было добавлено автоматически", "EVA")
            c.comment.width = 250
            c.comment.height = 30
        if not m.get("balcony_size"):
            c = ws.cell(row, 10)
            c.comment = Comment("Значение было добавлено автоматически", "EVA")
            c.comment.width = 250
            c.comment.height = 30

        # Dropdown валидация
        dv_size.add(ws.cell(row, 10))
        dv_loc.add(ws.cell(row, 11))
        dv_access.add(ws.cell(row, 12))

    # Жирная линия между сегментами (смена ЖК)
    last_data_row = len(buildings) + 2
    prev_cn = None
    for r in range(3, last_data_row + 1):
        cn = ws.cell(r, 3).value
        if prev_cn is not None and cn != prev_cn:
            _apply_segment_border(ws, r - 1, len(JK_HEADERS))
        prev_cn = cn

    # Сплошные вертикальные разделители по ВСЕМ строкам
    _apply_vertical_dividers(ws, 2, last_data_row, JK_BLOCK_DIVIDERS, JK_BLOCK_LEFT_DIVIDERS)

    # Условное форматирование (цвета по значениям)
    # Лист 2: E=5(дни), M=13(кв/клад), N=14(кв/балк)
    _apply_color_rules(ws, last_data_row, {
        "E": "days",
        "M": "ratio_apt_store",
        "N": "ratio_balcony",
    })

    ws.freeze_panes = "E3"
    if last_data_row > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(JK_HEADERS))}{last_data_row}"


# ─── Главная функция ─────────────────────────────────────

def export_eva_xlsx(
    conn_apt: sqlite3.Connection,
    conn_store: sqlite3.Connection,
    output_path: Path | None = None,
    filter_sites: list[str] | None = None,
) -> Path:
    """
    Сгенерировать единый xlsx расчёта ЕВА.

    Args:
        conn_apt: БД квартир (apartments_history.db)
        conn_store: БД кладовок (history.db)
        output_path: путь вывода (по умолчанию: расчет_ева.xlsx)
        filter_sites: фильтр сайтов для теста (None = все)
    """
    eva_config = load_eva_config()
    scoring = eva_config["scoring"]

    if output_path is None:
        output_path = PROJECT_DIR / eva_config["output_filename"]

    # ── 1. Загрузка ──
    logger.info("Загрузка данных...")

    # Квартиры и кладовки дом.рф
    domrf_apts = _load_apartments(conn_apt, "domrf")
    domrf_stores = _load_storehouses(conn_store, "domrf")

    # Квартиры и кладовки застройщиков (всё кроме domrf)
    all_apts = _load_apartments(conn_apt)
    all_stores = _load_storehouses(conn_store)
    dev_apts = [a for a in all_apts if a["site"] != "domrf"]
    dev_stores = [s for s in all_stores if s["site"] != "domrf"]

    if filter_sites:
        dev_apts = [a for a in dev_apts if a["site"] in filter_sites]
        dev_stores = [s for s in dev_stores if s["site"] in filter_sites]

    logger.info("Квартиры: %d (ДОМ.РФ) + %d (застройщики)", len(domrf_apts), len(dev_apts))
    logger.info("Кладовки: %d (ДОМ.РФ) + %d (застройщики)", len(domrf_stores), len(dev_stores))

    # ── 2. ObjectInfo ──
    object_infos = load_object_infos()

    # ── 3. Агрегация ──
    buildings = _aggregate(domrf_apts, domrf_stores, dev_apts, dev_stores, object_infos)
    logger.info("Корпусов: %d", len(buildings))

    # ── 4. Smart merge ручных данных ──
    manual = _read_manual_jk(output_path)
    if manual:
        logger.info("Загружено %d записей ручных данных из предыдущего файла", len(manual))

    # ── 5. Доп. данные ──
    dev_urls = load_developer_urls()
    logger.info("Загружено %d URL-ов застройщиков", len(dev_urls))

    living_areas = load_domrf_living_areas()

    # ── 6. Генерация XLSX ──
    wb = Workbook()

    # Назначить _jk_row заранее (нужно для формул на листе кладовок)
    for i, bd in enumerate(buildings):
        bd._jk_row = i + 3  # строка 3 = первая строка данных (после 2 строк заголовков)

    ws_store = wb.active
    ws_store.title = "кладовки (застройщик)"
    _fill_storehouses_sheet(ws_store, buildings)

    ws_jk = wb.create_sheet("жк")
    _fill_jk_sheet(ws_jk, buildings, manual, dev_urls, living_areas)

    # ── 6. Сохранение ──
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Файл сохранён: %s", output_path)
    return output_path
