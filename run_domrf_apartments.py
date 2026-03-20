#!/usr/bin/env python3
"""
Запуск парсера квартир ДОМ.РФ + экспорт в xlsx.

Использование:
    python run_domrf_apartments.py
"""
from __future__ import annotations

import asyncio
import sys

from parsers.apartments_base import (
    init_db, save_items, backup_db, validate_items,
    get_all_known_ids, calc_avg_prices, rooms_label, logger,
)
from parsers.domrf_apartments import DomRfApartmentParser, ObjectInfo
from exporter_apartments import export_apartments_xlsx

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re


_QUARTER_END_DATES = {
    "I": "31 марта",
    "II": "30 июня",
    "III": "30 сентября",
    "IV": "31 декабря",
}

_QUARTER_RE = re.compile(r'(I{1,3}V?)\s*квартал\s*(\d{4})', re.IGNORECASE)


def _quarter_comment(text: str) -> str | None:
    """Расшифровка квартала → конкретная дата конца квартала."""
    if not text:
        return None
    m = _QUARTER_RE.search(text)
    if not m:
        return None
    roman = m.group(1).upper()
    year = m.group(2)
    end_date = _QUARTER_END_DATES.get(roman)
    if not end_date:
        return None
    return f"{end_date} {year} года"


def _add_object_info_sheet(wb, object_infos: list[ObjectInfo]) -> None:
    """Добавить лист «Информация о домах» с данными из шапок страниц."""
    ws = wb.create_sheet("Информация о домах")

    headers = [
        "Объект (ID)", "ЖК", "Застройщик",
        "Ввод в эксплуатацию", "Выдача ключей",
        "Средняя цена за 1 м²", "Распроданность",
        "Всего квартир", "Продано квартир",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Сортировка: по застройщику, потом ЖК, потом object_id
    sorted_infos = sorted(object_infos, key=lambda o: (
        o.developer.lower(), o.complex_name.lower(), o.object_id,
    ))

    for i, info in enumerate(sorted_infos):
        row = i + 2
        row_data = [
            info.object_id,
            info.complex_name,
            info.developer,
            info.commissioning,
            info.keys_date,
            info.avg_price_per_meter,
            info.sold_percent,
            info.total_apartments or "",
            info.sold_apartments or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # Примечание с расшифровкой квартала (только для ввода в эксплуатацию)
        hint = _quarter_comment(info.commissioning)
        if hint:
            c = ws.cell(row=row, column=4)
            c.comment = Comment(hint, "Parser")
            c.comment.width = 200
            c.comment.height = 30

    # Ширина колонок
    widths = [12, 25, 18, 20, 18, 22, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    last_row = len(sorted_infos) + 1
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


async def main() -> int:
    logger.info("=" * 50)
    logger.info("Запуск парсера квартир ДОМ.РФ")
    logger.info("=" * 50)

    backup_db()
    conn = init_db()

    try:
        parser = DomRfApartmentParser()
        items, object_infos = await parser.parse_all()

        if not items:
            logger.error("Парсер не вернул ни одной квартиры!")
            return 1

        warnings = validate_items(items)
        if warnings:
            logger.warning("Обнаружено %d предупреждений валидации", len(warnings))

        previously_known = get_all_known_ids(conn, "domrf")

        updated = save_items(conn, items)
        logger.info("Обновлено записей в БД: %d", updated)

        output_path = export_apartments_xlsx(
            items, conn,
            filename="apartments_DomRF.xlsx",
            previously_known=previously_known,
        )

        # Добавляем лист с информацией о домах
        wb = load_workbook(str(output_path))
        _add_object_info_sheet(wb, object_infos)
        wb.save(str(output_path))

        logger.info("Файл готов: %s", output_path)

        # Статистика
        stats = calc_avg_prices(items)
        logger.info("Статистика:")
        logger.info("  Всего квартир: %d", len(items))
        for r, data in stats["by_rooms"].items():
            if data["avg_price"] > 0:
                logger.info(
                    "    %s: %d шт., ср. цена: %s ₽, ср. цена/м²: %s ₽",
                    rooms_label(r), data["count"],
                    f"{data['avg_price']:,.0f}", f"{data['avg_ppm']:,.0f}",
                )
            else:
                logger.info("    %s: %d шт.", rooms_label(r), data["count"])

        logger.info("  Объектов с информацией о доме: %d", len(object_infos))

        return 0

    except Exception as exc:
        logger.exception("Критическая ошибка: %s", exc)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
