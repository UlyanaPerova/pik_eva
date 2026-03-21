"""
Квартирография — лист с площадями по типам квартир + диаграмма.

Общий модуль для всех парсеров квартир.
Используется в run_*_apartments.py через add_kvartirografia_sheets().
"""
from __future__ import annotations

import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from parsers.apartments_base import ApartmentItem

# ── Константы ──────────────────────────────────────────

_ROOM_TYPES = [
    (0, "Студия"),
    (1, "1-комн."),
    (2, "2-комн."),
    (3, "3-комн."),
    (4, "4-комн."),
]

_CHART_COLORS = ["FFF9C4", "C5E1A5", "B3E5FC", "D3D3D3", "616161"]

_CHART_ROOM_TYPES = [
    (0, "Студия"),
    (1, "1 комнатные"),
    (2, "2 комнатные"),
    (3, "3 комнатные"),
    (4, "4+ комнатные"),
]

_KV_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_KV_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_KV_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_KV_SUBHEADER_FONT = Font(name="Calibri", bold=True, color="2F5496", size=10)
_KV_DATA_FONT = Font(name="Calibri", size=11)
_KV_DATA_ALIGN = Alignment(horizontal="center", vertical="center")
_KV_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_KV_GROUP_RIGHT = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="medium", color="4472C4"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_GROUP_SEPARATOR_COLS = {1, 4, 7, 10, 13, 16}


# ── Публичный API ──────────────────────────────────────

def add_kvartirografia_sheets(wb, items: list[ApartmentItem]) -> None:
    """Добавить листы «Квартирография» и «Диаграмма» в книгу."""
    _add_kvartirografia_sheet(wb, items)
    _add_kvartirografia_chart_sheet(wb, items)


# ── Лист «Квартирография» ─────────────────────────────

def _add_kvartirografia_sheet(wb, items: list[ApartmentItem]) -> None:
    """Лист с площадями по типам квартир в столбцах."""
    ws = wb.create_sheet("Квартирография")

    # Шапка: строка 1 — типы квартир (объединённые ячейки)
    jk_cell = ws.cell(row=1, column=1, value="ЖК")
    jk_cell.fill = _KV_HEADER_FILL
    jk_cell.font = _KV_HEADER_FONT
    jk_cell.alignment = Alignment(horizontal="center", vertical="center")
    jk_cell.border = _KV_THIN_BORDER
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for i, (_, label) in enumerate(_ROOM_TYPES):
        start_col = 2 + i * 3
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 2)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill = _KV_HEADER_FILL
        cell.font = _KV_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(start_col, start_col + 3):
            ws.cell(row=1, column=c).fill = _KV_HEADER_FILL
            border = _KV_GROUP_RIGHT if c in _GROUP_SEPARATOR_COLS else _KV_THIN_BORDER
            ws.cell(row=1, column=c).border = border

    # Строка 2 — подзаголовки
    sub_headers = ["Площадь\n(м²)", "Жилая\nплощадь\n(м²)", "Не жилая\nплощадь\n(м²)"]
    for i in range(5):
        start_col = 2 + i * 3
        for j, sh in enumerate(sub_headers):
            cell = ws.cell(row=2, column=start_col + j, value=sh)
            cell.fill = _KV_SUBHEADER_FILL
            cell.font = _KV_SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c = start_col + j
            cell.border = _KV_GROUP_RIGHT if c in _GROUP_SEPARATOR_COLS else _KV_THIN_BORDER

    # Данные: группировка по ЖК
    groups = defaultdict(list)
    for item in items:
        dev = (getattr(item, "developer", None) or "")
        groups[(dev, item.complex_name)].append(item)

    row = 3
    for (dev, complex_name) in sorted(groups.keys(), key=lambda k: (k[0].lower(), k[1].lower())):
        group_items = groups[(dev, complex_name)]

        by_rooms = defaultdict(list)
        for item in group_items:
            room_key = min(item.rooms, 4)
            by_rooms[room_key].append(item)

        for k in by_rooms:
            by_rooms[k].sort(key=lambda x: x.area)

        max_count = max((len(v) for v in by_rooms.values()), default=0)
        if max_count == 0:
            continue

        jk_label = f"{dev} — {complex_name}" if dev else complex_name
        ws.cell(row=row, column=1, value=jk_label).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=row, column=1).border = _KV_GROUP_RIGHT

        for r_offset in range(max_count):
            if r_offset > 0:
                ws.cell(row=row + r_offset, column=1).border = _KV_GROUP_RIGHT

            for type_idx, (rooms_num, _) in enumerate(_ROOM_TYPES):
                room_items = by_rooms.get(rooms_num, [])
                start_col = 2 + type_idx * 3

                if r_offset < len(room_items):
                    item = room_items[r_offset]
                    area = item.area
                    la = item.living_area or 0
                    non_living = round(area - la, 2) if la else ""

                    ws.cell(row=row + r_offset, column=start_col, value=area)
                    ws.cell(row=row + r_offset, column=start_col + 1,
                            value=la if la else "")
                    ws.cell(row=row + r_offset, column=start_col + 2,
                            value=non_living)

                for c in range(start_col, start_col + 3):
                    cell = ws.cell(row=row + r_offset, column=c)
                    cell.font = _KV_DATA_FONT
                    cell.alignment = _KV_DATA_ALIGN
                    cell.border = _KV_GROUP_RIGHT if c in _GROUP_SEPARATOR_COLS else _KV_THIN_BORDER
                    cell.number_format = '0.00'

        row += max_count + 1

    # Ширина колонок
    ws.column_dimensions["A"].width = 30
    for i in range(5):
        for j in range(3):
            col_letter = get_column_letter(2 + i * 3 + j)
            ws.column_dimensions[col_letter].width = 12


# ── Лист «Диаграмма» ──────────────────────────────────

def _kvartiry_plural(n: int) -> str:
    """Склонение слова «квартира» по числу."""
    last2 = n % 100
    last1 = n % 10
    if 11 <= last2 <= 19:
        return f"{n} квартир"
    if last1 == 1:
        return f"{n} квартира"
    if 2 <= last1 <= 4:
        return f"{n} квартиры"
    return f"{n} квартир"


def _compute_building_stats(group_items) -> dict:
    """Вычислить статистику по зданию из списка квартир."""
    entrances = set()
    floors = set()

    for it in group_items:
        floors.add(it.floor)
        if "||" in it.building:
            parts = it.building.split("||", 1)
            entrance_text = parts[1].strip()
            m = re.search(r'\d+', entrance_text)
            if m:
                entrances.add(int(m.group()))

    num_entrances = len(entrances) if entrances else 1
    max_floor = max(floors) if floors else 1
    min_floor = min(floors) if floors else 1

    num_floors = max_floor - min_floor + 1
    avg_per_floor = round(len(group_items) / num_floors / num_entrances) if num_floors > 0 else 0

    return {
        "num_entrances": num_entrances,
        "avg_per_floor": avg_per_floor,
        "first_floor_non_residential": min_floor > 1,
    }


def _add_kvartirografia_chart_sheet(wb, items: list[ApartmentItem]) -> None:
    """Лист «Диаграмма» — stacked bar + легенда + инфо-блок."""
    ws = wb.create_sheet("Диаграмма")

    # Ширина колонок
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 8

    hidden_font = Font(size=1, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    info_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    label_rpr = CharacterProperties(sz=1100, b=True, solidFill="555555")

    color_map = {rn: _CHART_COLORS[i] for i, (rn, _) in enumerate(_CHART_ROOM_TYPES)}

    # Агрегация по ВСЕМ квартирам
    room_counts = {}
    for rooms_num, _ in _CHART_ROOM_TYPES:
        room_counts[rooms_num] = sum(1 for it in items if min(it.rooms, 4) == rooms_num)

    total = sum(room_counts.values())
    if total == 0:
        return

    active_types = [(rn, label) for rn, label in _CHART_ROOM_TYPES if room_counts.get(rn, 0) > 0]
    percentages = [round(room_counts[rn] / total * 100) for rn, _ in active_types]

    row = 1

    # ============ ЗАГОЛОВОК ============
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Квартирография").font = Font(
        size=16, bold=True, color="000000")
    ws.row_dimensions[row].height = 35
    row += 1

    # ============ ДАННЫЕ ДЛЯ ДИАГРАММЫ (скрытые строки) ============
    label_row = row
    for i, (rn, label) in enumerate(active_types):
        cell = ws.cell(row=label_row, column=i + 2, value=label)
        cell.font = hidden_font
    ws.row_dimensions[label_row].height = 1
    row += 1

    data_row = row
    for i, pct in enumerate(percentages):
        cell = ws.cell(row=data_row, column=i + 2, value=pct)
        cell.font = hidden_font
        cell.number_format = '0"%"'
    ws.row_dimensions[data_row].height = 1
    row += 1

    # ============ ДИАГРАММА ============
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.style = 10
    chart.width = 22
    chart.height = 3
    chart.title = None
    chart.legend = None
    chart.x_axis.delete = True
    chart.y_axis.delete = True
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None

    for i, (rn, _) in enumerate(active_types):
        col = i + 2
        data_ref = Reference(ws, min_col=col, min_row=label_row, max_row=data_row)
        chart.add_data(data_ref, titles_from_data=True)
        series = chart.series[i]
        series.graphicalProperties.solidFill = color_map[rn]
        series.graphicalProperties.line.noFill = True

        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.showSerName = False
        series.dLbls.numFmt = '0"%"'
        series.dLbls.txPr = RichText(p=[Paragraph(
            pPr=ParagraphProperties(defRPr=label_rpr),
            endParaRPr=label_rpr,
        )])

        if percentages[i] < 3:
            series.dLbls.showVal = False

    chart.plot_area.graphicalProperties = None

    chart_row = row
    ws.add_chart(chart, f"A{chart_row}")
    for r in range(chart_row, chart_row + 4):
        ws.row_dimensions[r].height = 15
    row = chart_row + 6

    # ============ ЛЕГЕНДА ============
    ws.row_dimensions[row - 1].height = 15
    legend_start = row

    for i, (rn, label) in enumerate(active_types):
        count = room_counts[rn]
        color = color_map[rn]

        cell_a = ws.cell(row=row, column=1, value=" ")
        cell_a.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_a.border = thin_border

        ws.cell(row=row, column=2, value=label).font = Font(size=12, color="000000")
        ws.cell(row=row, column=3, value=_kvartiry_plural(count)).font = Font(size=12, color="808080")
        row += 1

    # ============ ИНФОРМАЦИОННЫЙ БЛОК ============
    stats = _compute_building_stats(items)
    info_data = [
        ("Среднее количество квартир на этаже", str(stats["avg_per_floor"])),
        ("Количество подъездов", str(stats["num_entrances"])),
        ("Первый этаж нежилой", "Да" if stats["first_floor_non_residential"] else "Нет"),
    ]

    for col in [5, 6]:
        ws.cell(row=legend_start - 1, column=col).border = Border(
            bottom=Side(style="thin", color="E0E0E0"))

    for ii, (lbl, val) in enumerate(info_data):
        info_row = legend_start + ii
        cell_e = ws.cell(row=info_row, column=5, value=lbl)
        cell_e.font = Font(size=11, color="666666")
        cell_e.border = info_border

        cell_f = ws.cell(row=info_row, column=6, value=val)
        cell_f.font = Font(size=14, bold=True, color="000000")
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.border = info_border
