"""
Общие стили, константы и утилиты для экспортёров кладовок и квартир.

Устраняет дублирование между exporter.py и exporter_apartments.py.
"""
from __future__ import annotations

import re

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Стили ────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

CITY_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CITY_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)

COMPLEX_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
COMPLEX_FONT = Font(name="Calibri", bold=True, color="2F5496", size=12)

DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
NEW_ITEM_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
COMPLEX_TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GRAND_TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

BUILDING_BOTTOM = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="808080"),
)


# ── Справочники сайтов ──────────────────────────────────

SITE_NAMES = {
    "pik": "ПИК",
    "akbarsdom": "Ак Бар Дом",
    "smu88": "СМУ-88",
    "glorax": "GloraX",
    "unistroy": "УниСтрой",
    "domrf": "ДОМ.РФ",
}

SITE_FILE_KEYS = {
    "pik": "PIK",
    "akbarsdom": "AkBarsDom",
    "smu88": "SMU88",
    "glorax": "GloraX",
    "unistroy": "Unistroy",
    "domrf": "DomRF",
}


# ── Утилиты ─────────────────────────────────────────────

def append_comment(cell, text: str, author: str = "Парсер") -> None:
    """Добавить текст к комментарию ячейки, не затирая существующий."""
    if cell.comment:
        existing = cell.comment.text or ""
        if text not in existing:
            cell.comment = Comment(f"{existing}\n{text}", author)
    else:
        cell.comment = Comment(text, author)


def natural_sort_key(text: str):
    """Натуральная сортировка: 'ПК-2' < 'ПК-10', 'Корпус 1.1' < 'Корпус 1.2'."""
    parts = re.split(r'(\d+(?:\.\d+)?)', text.lower())
    result = []
    for part in parts:
        try:
            result.append(float(part))
        except ValueError:
            result.append(part)
    return result


def add_new_item_comment(
    ws, row, col, item, previously_known, conn, baseline_ids,
    *,
    get_first_seen_fn,
    author: str = "Парсер",
    total_cols: int | None = None,
) -> None:
    """
    Примечание «Добавлена от [дата]» — если item нет в baseline.
    Зелёная заливка — только если item новый в ЭТОМ парсинге.

    Args:
        get_first_seen_fn: функция (conn, site, item_id) -> str | None
        author: автор примечания ('Парсер кладовок' или 'Парсер квартир')
    """
    is_first_run = len(previously_known) == 0
    is_new_this_parse = (not is_first_run) and (item.item_id not in previously_known)
    is_after_baseline = item.item_id not in baseline_ids

    if is_after_baseline:
        first_seen = get_first_seen_fn(conn, item.site, item.item_id)
        if first_seen:
            try:
                from datetime import datetime as _dt
                dt = _dt.fromisoformat(first_seen)
                date_str = dt.strftime("%d.%m.%Y %H:%M")
            except (ValueError, TypeError):
                date_str = first_seen[:16]
        else:
            from datetime import datetime as _dt
            date_str = _dt.now().strftime("%d.%m.%Y %H:%M")

        cell = ws.cell(row=row, column=col)
        append_comment(cell, f"Добавлена от {date_str}", author)
        cell.comment.width = 200
        cell.comment.height = 30

    if is_new_this_parse:
        ncols = total_cols or ws.max_column
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = NEW_ITEM_FILL


def add_price_comment(ws, row, col, conn, item, *, get_price_history_fn, author: str = "Парсер") -> None:
    """Примечание к ячейке «Цена» — скидка + история."""
    lines = []

    if item.discount_percent and item.original_price:
        lines.append(
            f"Скидка {item.discount_percent:.0f}%, "
            f"цена без скидки: {item.original_price:,.0f} ₽"
        )

    history = get_price_history_fn(conn, item.site, item.item_id)
    if len(history) > 1:
        lines.append("Предыдущие цены:")
        for price, ppm, orig_price, discount, date_str in history[1:]:
            date_short = date_str[:10]
            entry = f"• {price:,.0f} ₽ ({date_short})"
            if discount and orig_price:
                entry += f" [скидка {discount:.0f}%, без скидки: {orig_price:,.0f} ₽]"
            lines.append(entry)

    if lines:
        cell = ws.cell(row=row, column=col)
        append_comment(cell, "\n".join(lines), author)
        cell.comment.width = 350
        cell.comment.height = max(80, len(lines) * 20)


def add_ppm_comment(ws, row, col, conn, item, *, get_price_history_fn, author: str = "Парсер") -> None:
    """Примечание к ячейке «Цена/м²» — история."""
    history = get_price_history_fn(conn, item.site, item.item_id)
    if len(history) <= 1:
        return

    lines = ["Предыдущие цены/м²:"]
    for price, ppm, orig_price, discount, date_str in history[1:]:
        date_short = date_str[:10]
        lines.append(f"• {ppm:,.0f} ₽/м² ({date_short})")

    cell = ws.cell(row=row, column=col)
    append_comment(cell, "\n".join(lines), author)
    cell.comment.width = 300
    cell.comment.height = max(60, len(lines) * 18)
