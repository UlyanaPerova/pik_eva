"""
Генератор единого XLSX-файла расчёта ЕВА.

Компиляция данных из:
  - apartments_DomRF.xlsx   (квартиры с дом.рф: площади, жилая/нежилая, кол-во)
  - storehouses_DomRF.xlsx  (кладовки с дом.рф: общее кол-во)
  - apartments_[DEV].xlsx   (квартиры застройщика: цены)
  - storehouses_[DEV].xlsx  (кладовки застройщика: цены, остаток)
  + «Информация о домах» из apartments_DomRF (квартал сдачи → дни)

Два листа:
  1. «кладовки (застройщик)» — строка на каждую кладовку
  2. «жк» — строка на каждый корпус
"""
from __future__ import annotations

import logging
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("eva")

PROJECT_DIR = Path(__file__).resolve().parent
CONFIGS_DIR = PROJECT_DIR / "configs"
APT_DIR = PROJECT_DIR / "apartments"
STORE_DIR = PROJECT_DIR / "output"

# ─── Стили ───────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
NO_BORDER = Border()  # без линий по умолчанию
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
MANUAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SCORE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="2F5496", size=11)
GRAY_FONT = Font(name="Calibri", size=11, color="BFBFBF")

SEGMENT_BOTTOM = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="4472C4"),
)

# Правая жирная граница (разделитель блоков столбцов)
BLOCK_RIGHT = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="medium", color="4472C4"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Столбцы-разделители блоков на листе «жк»
# Последний столбец каждого блока:
# 18 (ср. цена м²) | 23 (квартирография) | 28 (ср. площадь) | 33 (ср. нежилая)
# Столбцы с жирной ПРАВОЙ границей на листе «жк»
# 4=корпус, 7=ссылка застр., 8=первый этап, 12=удобство доступа,
# 16=кладовых ДОМ.РФ, 23=4к+ квартирография, 28=ср.площадь 4к+,
# 33=ср.нежилая 4к+, 34=исх.порядок
# Столбцы с жирной ПРАВОЙ границей
JK_BLOCK_DIVIDERS = {4, 7, 8, 12, 16, 23, 28, 33, 34}
# Столбцы с жирной ЛЕВОЙ границей = столбцы ПОСЛЕ разделителей + первый этап
JK_BLOCK_LEFT_DIVIDERS = {5, 8, 9, 13, 17, 24, 29, 34}

DOMRF_BASE = "https://xn--80az8a.xn--d1aqf.xn--p1ai"

# Маппинг site_key → человекочитаемое имя застройщика
DEVELOPER_NAMES = {
    "pik": "ПИК",
    "akbarsdom": "Ак Барс Дом",
    "glorax": "GloraX",
    "smu88": "СМУ-88",
    "unistroy": "Унистрой",
    "domrf": "ДОМ.РФ",
}


# ─── Алиасы и нормализация ─────────────────────────────────

_COMPLEX_ALIASES: dict[str, str] = {}  # заполняется из eva.yaml при загрузке
_BUILDING_ALIASES: dict[str, dict[str, str]] = {}  # ЖК → {паттерн: корпус DomRF}


def _apply_alias(complex_name: str) -> str:
    """Привести имя ЖК к каноническому через алиасы."""
    return _COMPLEX_ALIASES.get(complex_name, complex_name)


def _apply_building_alias(complex_name: str, building: str) -> str:
    """Привести корпус застройщика к корпусу DomRF через маппинг."""
    cn = _apply_alias(complex_name)
    aliases = _BUILDING_ALIASES.get(cn, {})
    if not aliases:
        return building
    for pattern, target in aliases.items():
        if pattern.startswith("^"):
            # Prefix match
            prefix = pattern[1:]
            if building.startswith(prefix):
                return target
        else:
            # Contains match (с учётом спецсимволов в Адали)
            if pattern in building:
                return target
    return building


def _norm(s: str) -> str:
    """Нормализовать строку для матчинга."""
    if not s:
        return ""
    s = s.lower().strip()
    # "Корпус 10" → "10", "корпус 1" → "1"
    s = re.sub(r'^корпус\s*', '', s)
    # Убрать дефисы, пробелы, подчёркивания, слеши
    s = re.sub(r'[\s\-–—_/]', '', s)
    return s


def _match_key(city: str, complex_name: str, building: str) -> tuple:
    return (_norm(city), _norm(_apply_alias(complex_name)), _norm(building))


# ─── Загрузка данных из БД ───────────────────────────────

def _load_apartments(conn: sqlite3.Connection, site_filter: str | None = None) -> list[dict]:
    """Загрузить последние записи квартир."""
    query = """
        SELECT a.site, a.city, a.complex_name, a.building, a.item_id,
               a.rooms, a.floor, a.area, a.price, a.price_per_meter,
               a.living_area, a.url
        FROM apartment_prices a
        INNER JOIN (
            SELECT site, item_id, MAX(parsed_at) AS max_pa
            FROM apartment_prices
            GROUP BY site, item_id
        ) latest ON a.site = latest.site AND a.item_id = latest.item_id
                 AND a.parsed_at = latest.max_pa
    """
    if site_filter:
        query += f" WHERE a.site = '{site_filter}'"
    rows = conn.execute(query).fetchall()
    cols = ["site", "city", "complex_name", "building", "item_id",
            "rooms", "floor", "area", "price", "price_per_meter",
            "living_area", "url"]
    return [dict(zip(cols, r)) for r in rows]


def _load_storehouses(conn: sqlite3.Connection, site_filter: str | None = None) -> list[dict]:
    """Загрузить последние записи кладовок."""
    query = """
        SELECT p.site, p.city, p.complex_name, p.building, p.item_id,
               p.item_number, p.area, p.price, p.price_per_meter, p.url
        FROM prices p
        INNER JOIN (
            SELECT site, item_id, MAX(parsed_at) AS max_pa
            FROM prices
            GROUP BY site, item_id
        ) latest ON p.site = latest.site AND p.item_id = latest.item_id
                 AND p.parsed_at = latest.max_pa
    """
    if site_filter:
        query += f" WHERE p.site = '{site_filter}'"
    rows = conn.execute(query).fetchall()
    cols = ["site", "city", "complex_name", "building", "item_id",
            "item_number", "area", "price", "price_per_meter", "url"]
    return [dict(zip(cols, r)) for r in rows]


# ─── Загрузка ObjectInfo из xlsx ─────────────────────────

@dataclass
class ObjectInfo:
    object_id: int
    complex_name: str
    developer: str
    commissioning: str = ""
    total_apartments: int = 0


def load_object_infos(xlsx_path: Path | None = None) -> list[ObjectInfo]:
    """Прочитать 'Информация о домах' из apartments_DomRF.xlsx."""
    if xlsx_path is None:
        xlsx_path = APT_DIR / "apartments_DomRF.xlsx"
    if not xlsx_path.exists():
        logger.warning("Нет файла %s — commissioning будет пустым", xlsx_path)
        return []

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Информация о домах" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Информация о домах"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        result.append(ObjectInfo(
            object_id=int(row[0]) if row[0] else 0,
            complex_name=str(row[1] or ""),
            developer=str(row[2] or ""),
            commissioning=str(row[3] or ""),
            total_apartments=int(row[7]) if row[7] else 0,
        ))
    wb.close()
    logger.info("Загружено %d ObjectInfo из %s", len(result), xlsx_path.name)
    return result


def load_domrf_living_areas(xlsx_path: Path | None = None) -> dict[tuple, dict]:
    """
    Прочитать жилую площадь из apartments_DomRF.xlsx → 'Все данные'.
    Возвращает: {(norm_city, norm_complex): {rooms: (avg_area, avg_living)}}
    """
    if xlsx_path is None:
        xlsx_path = APT_DIR / "apartments_DomRF.xlsx"
    if not xlsx_path.exists():
        return {}

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "Все данные" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Все данные"]
    # Cols: 1=Город, 3=ЖК, 5=Тип, 9=Площадь, 10=Жилая площадь
    room_map = {"Студия": 0, "1-комн.": 1, "2-комн.": 2, "3-комн.": 3, "4-комн.": 4}

    # Собираем: (city, complex) → rooms → [(area, living)]
    raw: dict[tuple, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        city = str(row[0] or "")
        cn = str(row[2] or "")
        room_type = room_map.get(str(row[4] or ""))
        if room_type is None:
            continue
        area = row[8] if row[8] else None
        living = row[9] if row[9] else None
        if area and living:
            ck = _complex_key(city, cn)
            raw[ck][room_type].append((float(area), float(living)))

    # Агрегация
    result: dict[tuple, dict] = {}
    for ck, rooms_data in raw.items():
        agg = {}
        for rooms, pairs in rooms_data.items():
            areas = [p[0] for p in pairs]
            livings = [p[1] for p in pairs]
            avg_area = sum(areas) / len(areas)
            avg_living = sum(livings) / len(livings)
            agg[rooms] = (round(avg_area, 1), round(avg_living, 1))
        result[ck] = agg

    wb.close()
    logger.info("Загружено жилых площадей для %d комплексов", len(result))
    return result


def load_developer_urls() -> dict[str, str]:
    """
    Загрузить base_url сайтов застройщиков из yaml-конфигов.
    Возвращает: {developer_name: base_url}
    """
    result: dict[str, str] = {}

    for cfg_name in ["pik", "akbarsdom", "glorax", "smu88", "unistroy"]:
        cfg_path = CONFIGS_DIR / f"{cfg_name}.yaml"
        if not cfg_path.exists():
            continue
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)

        base_url = cfg.get("base_url", cfg.get("url", ""))
        if base_url:
            # Привязываем к человекочитаемому имени
            dev_name = DEVELOPER_NAMES.get(cfg_name, cfg_name)
            result[dev_name] = base_url

    return result


# ─── Расчёт дней до сдачи ────────────────────────────────

def _avg_ppm(items: list[dict]) -> float:
    ppms = [a["price_per_meter"] for a in items if a.get("price_per_meter")]
    return round(sum(ppms) / len(ppms)) if ppms else 0


def _calc_priority_ppm(apts: list[dict]) -> float:
    """
    Средняя цена за м² по приоритету (п.1-2, п.4):
    1) Только 1к
    2) Среднее между студиями и 2к
    4) Все данные, что есть
    П.3 (соседний корпус) применяется отдельно в _apply_neighbor_ppm.
    """
    # П.1: только 1к
    one_k = [a for a in apts if a["rooms"] == 1]
    if one_k:
        avg = _avg_ppm(one_k)
        if avg:
            return avg

    # П.2: среднее между студиями и 2к
    studios = [a for a in apts if a["rooms"] == 0]
    two_k = [a for a in apts if a["rooms"] == 2]
    if studios and two_k:
        avg_s = _avg_ppm(studios)
        avg_2 = _avg_ppm(two_k)
        if avg_s and avg_2:
            return round((avg_s + avg_2) / 2)

    # П.4: всё что есть
    return _avg_ppm(apts)


def _apply_neighbor_ppm(buildings_list: list) -> None:
    """
    П.3: если у корпуса ppm=0, взять ppm по п.1 от соседнего корпуса
    того же ЖК. Сосед: сначала шаг назад, если невозможно — шаг вперёд.
    """
    from itertools import groupby

    # Группируем по ЖК
    def ck(b):
        return (_norm(b.city), _norm(b.complex_name))

    by_complex: dict[tuple, list] = defaultdict(list)
    for b in buildings_list:
        by_complex[ck(b)].append(b)

    for complex_key, blds in by_complex.items():
        if len(blds) < 2:
            continue

        for i, b in enumerate(blds):
            if b.dev_avg_apt_ppm > 0:
                continue

            # Шаг назад
            neighbor_ppm = 0
            if i > 0 and blds[i - 1].dev_avg_apt_ppm > 0:
                neighbor_ppm = blds[i - 1].dev_avg_apt_ppm
            # Шаг вперёд
            elif i + 1 < len(blds) and blds[i + 1].dev_avg_apt_ppm > 0:
                neighbor_ppm = blds[i + 1].dev_avg_apt_ppm

            if neighbor_ppm:
                b.dev_avg_apt_ppm = neighbor_ppm


_QUARTER_RE = re.compile(r'(I{1,3}V?)\s*(?:квартал|кв\.?)\s*(\d{4})', re.IGNORECASE)
_QUARTER_END = {"I": (3, 31), "II": (6, 30), "III": (9, 30), "IV": (12, 31)}


def _days_until(commissioning: str) -> int | None:
    """Посчитать дни от сегодня до конца квартала сдачи."""
    if not commissioning:
        return None
    m = _QUARTER_RE.search(commissioning)
    if not m:
        return None
    roman = m.group(1).upper()
    year = int(m.group(2))
    month, day = _QUARTER_END.get(roman, (12, 31))
    try:
        target = date(year, month, day)
    except ValueError:
        return None
    delta = (target - date.today()).days
    return max(delta, 0)


# ─── Конфиг разбалловки ──────────────────────────────────

def load_eva_config() -> dict:
    global _COMPLEX_ALIASES, _BUILDING_ALIASES
    with open(CONFIGS_DIR / "eva.yaml", "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    _COMPLEX_ALIASES = cfg.get("complex_aliases", {})
    _BUILDING_ALIASES = cfg.get("building_aliases", {})
    return cfg


def _score_by_thresholds(value: float, thresholds: list[dict]) -> int:
    for t in thresholds:
        if value < t["max"]:
            return t["points"]
    return thresholds[-1]["points"] if thresholds else 0


# ─── Агрегация данных ────────────────────────────────────

@dataclass
class BuildingAgg:
    """Агрегированные данные одного корпуса."""
    city: str = ""
    developer: str = ""
    complex_name: str = ""
    building: str = ""

    # ObjectInfo (дом.рф)
    object_ids: list = field(default_factory=list)
    commissioning: str = ""
    days_until: int | None = None

    # Кол-во с дом.рф (per-building для квартир, per-COMPLEX для кладовок)
    domrf_apt_count: int = 0
    domrf_store_count: int = 0    # на уровне ВСЕГО ЖК

    # Кол-во у застройщика (остаток в продаже)
    dev_store_count: int = 0

    # Средняя цена м² квартиры (от застройщика)
    dev_avg_apt_ppm: float = 0

    # Квартирография (кол-во по типам, дом.рф, per-building)
    rooms_count: dict = field(default_factory=dict)
    avg_area: dict = field(default_factory=dict)
    avg_non_living: dict = field(default_factory=dict)

    # Кладовки застройщика (для листа 1)
    dev_storehouses: list = field(default_factory=list)
    # Кладовки дом.рф (fallback, per-complex)
    domrf_storehouses: list = field(default_factory=list)

    # Ссылки
    domrf_link: str = ""
    dev_link: str = ""


def _complex_key(city: str, complex_name: str) -> tuple:
    """Ключ на уровне ЖК (без корпуса)."""
    return (_norm(city), _norm(_apply_alias(complex_name)))


def _aggregate(
    domrf_apts: list[dict],
    domrf_stores: list[dict],
    dev_apts: list[dict],
    dev_stores: list[dict],
    object_infos: list[ObjectInfo],
) -> list[BuildingAgg]:
    """
    Агрегация:
    - Данные дом.рф (квартиры + кладовки) → на уровне КОМПЛЕКСА (город + ЖК),
      т.к. building = API-формат ("1||подъезд 2"), не совпадает с застройщиком.
    - Данные застройщика → по корпусам.
    - Строки берутся из конфига domrf_apartments.yaml (уникальные корпуса)
      + корпуса застройщиков.
    - Каждый корпус получает данные дом.рф от своего ЖК.
    """

    oi_by_id = {oi.object_id: oi for oi in object_infos}
    domrf_config = _load_domrf_config()

    # ─── Шаг 1: Агрегация дом.рф на уровне КОМПЛЕКСА ───

    # Квартиры дом.рф → по комплексу
    domrf_apt_by_complex: dict[tuple, list[dict]] = defaultdict(list)
    for apt in domrf_apts:
        ck = _complex_key(apt["city"], apt["complex_name"])
        domrf_apt_by_complex[ck].append(apt)

    # Кладовки дом.рф → по комплексу
    domrf_store_by_complex: dict[tuple, list[dict]] = defaultdict(list)
    for st in domrf_stores:
        ck = _complex_key(st["city"], st["complex_name"])
        domrf_store_by_complex[ck].append(st)

    # Расчёт агрегатов по комплексу
    @dataclass
    class ComplexAgg:
        apt_count: int = 0
        store_count: int = 0
        rooms_count: dict = field(default_factory=dict)
        avg_area: dict = field(default_factory=dict)
        avg_non_living: dict = field(default_factory=dict)
        storehouses: list = field(default_factory=list)

    complex_agg: dict[tuple, ComplexAgg] = {}

    for ck, apts in domrf_apt_by_complex.items():
        ca = ComplexAgg(apt_count=len(apts))

        rooms_groups: dict[int, list[dict]] = defaultdict(list)
        for a in apts:
            rooms_groups[a["rooms"]].append(a)

        for rooms in range(5):
            group = rooms_groups.get(rooms, [])
            ca.rooms_count[rooms] = len(group)

            areas = [a["area"] for a in group if a.get("area")]
            ca.avg_area[rooms] = round(sum(areas) / len(areas), 1) if areas else 0

            livings = [a["living_area"] for a in group if a.get("living_area")]
            if livings and areas:
                avg_total = sum(areas) / len(areas)
                avg_living = sum(livings) / len(livings)
                ca.avg_non_living[rooms] = round(avg_total - avg_living, 1)
            else:
                ca.avg_non_living[rooms] = 0

        complex_agg[ck] = ca

    for ck, stores in domrf_store_by_complex.items():
        if ck not in complex_agg:
            complex_agg[ck] = ComplexAgg()
        complex_agg[ck].store_count = len(stores)
        complex_agg[ck].storehouses = stores

    # ─── Шаг 2: Определяем список корпусов (строк) ───
    # Источники: конфиг domrf + застройщики
    buildings: dict[tuple, BuildingAgg] = {}

    def _is_better_name(new: str, old: str) -> bool:
        """Человекочитаемое имя лучше ключа (pik → ПИК, akbarsdom → Ак Барс Дом)."""
        if not new:
            return False
        if not old:
            return True
        # Кириллица или заглавные буквы/пробелы → лучше
        has_cyrillic = bool(re.search(r'[а-яА-ЯёЁ]', new))
        has_space = ' ' in new
        old_is_key = old.isascii() and old.islower()
        return (has_cyrillic or has_space) and old_is_key

    def _get_or_create(city, complex_name, building, developer="") -> BuildingAgg:
        key = _match_key(city, complex_name, building)
        if key not in buildings:
            buildings[key] = BuildingAgg(
                city=city, complex_name=complex_name, building=building,
                developer=developer,
            )
        b = buildings[key]
        if _is_better_name(developer, b.developer):
            b.developer = developer
        elif developer and not b.developer:
            b.developer = developer
        # Предпочитаем более полное имя ЖК (Нокса-Парк лучше Нокса парк? нет, берём первое)
        return b

    # Корпуса из конфига domrf (нормальные названия: М1/ПК-1, Корпус 1, etc.)
    config_complexes: set[tuple] = set()  # для трекинга комплексов с конфигом
    for link in domrf_config.get("links", []):
        city = link.get("city", "")
        cn = link.get("complex_name", "")
        bld = link.get("building", "")
        dev = link.get("developer", "")
        obj_id = link.get("object_id")
        if not cn:
            continue

        b = _get_or_create(city, cn, bld, developer=dev)
        config_complexes.add(_complex_key(city, cn))

        if obj_id:
            if obj_id not in b.object_ids:
                b.object_ids.append(obj_id)
            b.domrf_link = (
                f"{DOMRF_BASE}/сервисы/каталог-новостроек/объект/{obj_id}"
            )
            oi = oi_by_id.get(obj_id)
            if oi:
                if not b.commissioning:
                    b.commissioning = oi.commissioning
                    b.days_until = _days_until(oi.commissioning)
                if oi.developer:
                    b.developer = oi.developer

    # Корпуса из застройщиков
    # Логика: если корпус застройщика СОВПАДАЕТ с DomRF корпусом (после нормализации)
    # → привязать данные к существующей строке. Если НЕ совпадает → создать новую строку.
    # Кладовки застройщика также считаются на уровне комплекса для "остаток".

    # Индекс DomRF корпусов по комплексу для быстрого поиска
    config_buildings_by_complex: dict[tuple, dict[str, tuple]] = defaultdict(dict)
    for key, b in buildings.items():
        ck = _complex_key(b.city, b.complex_name)
        if ck in config_complexes:
            config_buildings_by_complex[ck][_norm(b.building)] = key

    complex_dev_stores: dict[tuple, list[dict]] = defaultdict(list)

    for st in dev_stores:
        cn = _apply_alias(st["complex_name"])
        ck = _complex_key(st["city"], cn)

        if ck in config_complexes:
            complex_dev_stores[ck].append(st)
            # Привязать к совпадающему корпусу (после alias + нормализации)
            resolved_bld = _apply_building_alias(cn, st["building"])
            norm_bld = _norm(resolved_bld)
            matched_key = config_buildings_by_complex.get(ck, {}).get(norm_bld)
            if matched_key and matched_key in buildings:
                buildings[matched_key].dev_storehouses.append(st)
            else:
                # Корпус не совпал — создать новую строку
                b = _get_or_create(st["city"], cn, st["building"],
                                   developer=st.get("site", ""))
                b.dev_storehouses.append(st)
        else:
            b = _get_or_create(st["city"], cn, st["building"],
                               developer=st.get("site", ""))
            b.dev_storehouses.append(st)

    # "Остаток кладовок" — на уровне комплекса для всех корпусов
    for ck, stores in complex_dev_stores.items():
        for b in buildings.values():
            if _complex_key(b.city, b.complex_name) == ck:
                b.dev_store_count = len(stores)

    dev_apt_groups: dict[tuple, list[dict]] = defaultdict(list)
    for apt in dev_apts:
        cn = _apply_alias(apt["complex_name"])
        ck = _complex_key(apt["city"], cn)

        if ck in config_complexes:
            # Проверяем: совпадает ли корпус (после alias + нормализации)?
            resolved_bld = _apply_building_alias(cn, apt["building"])
            norm_bld = _norm(resolved_bld)
            matched_key = config_buildings_by_complex.get(ck, {}).get(norm_bld)
            if matched_key:
                dev_apt_groups[matched_key].append(apt)
            else:
                # Уникальный корпус застройщика — создать строку, если нет ||
                if "||" not in apt["building"] and apt["building"]:
                    key = _match_key(apt["city"], cn, apt["building"])
                    dev_apt_groups[key].append(apt)
                    _get_or_create(apt["city"], cn, apt["building"],
                                   developer=apt.get("site", ""))
                else:
                    # Секции/подъезды — агрегировать на уровне комплекса
                    key = ("_complex_dev", ck[0], ck[1])
                    dev_apt_groups[key].append(apt)
        else:
            key = _match_key(apt["city"], cn, apt["building"])
            dev_apt_groups[key].append(apt)
            _get_or_create(apt["city"], cn, apt["building"],
                           developer=apt.get("site", ""))

    for key, apts in dev_apt_groups.items():
        avg_ppm = _calc_priority_ppm(apts)

        if key[0] == "_complex_dev":
            ck = (key[1], key[2])
            for b in buildings.values():
                if _complex_key(b.city, b.complex_name) == ck:
                    b.dev_avg_apt_ppm = avg_ppm
        elif key in buildings:
            buildings[key].dev_avg_apt_ppm = avg_ppm

    # ─── Шаг 3: Привязать данные дом.рф (на уровне комплекса) к корпусам ───
    for b in buildings.values():
        ck = _complex_key(b.city, b.complex_name)
        ca = complex_agg.get(ck)
        if ca:
            b.domrf_apt_count = ca.apt_count
            b.domrf_store_count = ca.store_count
            b.rooms_count = dict(ca.rooms_count)
            b.avg_area = dict(ca.avg_area)
            b.avg_non_living = dict(ca.avg_non_living)
        b.dev_store_count = len(b.dev_storehouses)

    # ─── Шаг 3.5: Для ЖК БЕЗ DomRF — агрегировать из данных застройщика ───
    # Квартирография, средняя площадь, кол-во квартир из dev apartments
    for key, apts in dev_apt_groups.items():
        if key[0] == "_complex_dev":
            continue  # это DomRF-ЖК, уже обработаны
        if key not in buildings:
            continue
        b = buildings[key]
        ck = _complex_key(b.city, b.complex_name)
        if ck in config_complexes:
            continue  # DomRF-ЖК, данные уже из DomRF

        # Агрегируем квартиры застройщика по корпусу
        b.domrf_apt_count = len(apts)  # используем как "всего квартир"
        rooms_groups: dict[int, list[dict]] = defaultdict(list)
        for a in apts:
            rooms_groups[a["rooms"]].append(a)

        for rooms in range(5):
            group = rooms_groups.get(rooms, [])
            b.rooms_count[rooms] = len(group)
            areas = [a["area"] for a in group if a.get("area")]
            b.avg_area[rooms] = round(sum(areas) / len(areas), 1) if areas else 0
            # Нежилая площадь — только из DomRF, тут не заполняем

    # ─── Шаг 4: Кладовки дом.рф fallback ───
    # Если у комплекса нет кладовок застройщика, используем дом.рф
    complex_has_dev: dict[tuple, bool] = defaultdict(bool)
    for b in buildings.values():
        if b.dev_storehouses:
            ck = _complex_key(b.city, b.complex_name)
            complex_has_dev[ck] = True

    for ck, ca in complex_agg.items():
        if not complex_has_dev.get(ck, False) and ca.storehouses:
            # Создаём "сводные" строки для кладовок дом.рф
            by_bld: dict[str, list[dict]] = defaultdict(list)
            for st in ca.storehouses:
                by_bld[st["building"]].append(st)
            for bld_name, bld_stores in by_bld.items():
                city = bld_stores[0]["city"]
                cn = bld_stores[0]["complex_name"]
                b = _get_or_create(city, cn, bld_name)
                b.domrf_storehouses = bld_stores

    # ─── Шаг 5: Убрать мусорные строки ───
    # "1||подъезд 1" → мусор если у комплекса есть нормальные корпуса из конфига
    # Пустой building → мусор если у комплекса есть другие корпуса
    keys_to_remove = []
    for key, b in buildings.items():
        ck = _complex_key(b.city, b.complex_name)
        if "||" in b.building and ck in config_complexes:
            keys_to_remove.append(key)
        elif not b.building.strip():
            # Пустой корпус — убрать если у ЖК есть другие корпуса
            has_others = any(
                k != key and _complex_key(ob.city, ob.complex_name) == ck
                and ob.building.strip()
                for k, ob in buildings.items()
            )
            if has_others:
                keys_to_remove.append(key)
    for key in keys_to_remove:
        del buildings[key]

    # ─── Шаг 6: Пропагация данных на уровне комплекса ───
    # Средняя цена м² застройщика — на уровне комплекса
    complex_dev_ppm: dict[tuple, float] = {}
    for b in buildings.values():
        if b.dev_avg_apt_ppm:
            ck = _complex_key(b.city, b.complex_name)
            if ck not in complex_dev_ppm:
                complex_dev_ppm[ck] = b.dev_avg_apt_ppm

    # Собрать все квартиры застройщика по комплексу для пересчёта ppm
    complex_all_apts: dict[tuple, list[dict]] = defaultdict(list)
    for key, apts in dev_apt_groups.items():
        b_tmp = buildings.get(key)
        if b_tmp:
            ck = _complex_key(b_tmp.city, b_tmp.complex_name)
            complex_all_apts[ck].extend(apts)

    for ck, apts in complex_all_apts.items():
        ppm = _calc_priority_ppm(apts)
        if ppm:
            complex_dev_ppm[ck] = ppm

    # Ссылка ДОМ.РФ — на уровне комплекса (первый object_id)
    complex_domrf_link: dict[tuple, str] = {}
    for b in buildings.values():
        if b.domrf_link:
            ck = _complex_key(b.city, b.complex_name)
            if ck not in complex_domrf_link:
                complex_domrf_link[ck] = b.domrf_link

    # Применить complex-level данные ко всем корпусам
    for b in buildings.values():
        ck = _complex_key(b.city, b.complex_name)
        if not b.dev_avg_apt_ppm and ck in complex_dev_ppm:
            b.dev_avg_apt_ppm = complex_dev_ppm[ck]
        if not b.domrf_link and ck in complex_domrf_link:
            b.domrf_link = complex_domrf_link[ck]

    # П.3: для корпусов с ppm=0 — взять от соседнего корпуса
    _apply_neighbor_ppm(list(buildings.values()))

    # Нормализация имён застройщиков
    for b in buildings.values():
        if b.developer in DEVELOPER_NAMES:
            b.developer = DEVELOPER_NAMES[b.developer]

    # Сортировка
    result = sorted(buildings.values(),
                    key=lambda b: (b.city, b.developer, b.complex_name, b.building))
    return result


def _load_domrf_config() -> dict:
    path = CONFIGS_DIR / "domrf_apartments.yaml"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ─── Скоринг ─────────────────────────────────────────────

def calc_first_stage(bd: BuildingAgg, scoring: dict) -> int:
    """Первый этап: критерии 1–10 (7-10 зависят от ручных данных → пока 0)."""
    pts = 0

    # 1. Срок ввода
    if bd.days_until is not None:
        months = bd.days_until / 30.0
        pts += _score_by_thresholds(months, scoring["deadline_months"])

    # 2. Соотношение квартир / кладовых
    store_total = bd.domrf_store_count or bd.dev_store_count
    if store_total:
        ratio = bd.domrf_apt_count / store_total
        pts += _score_by_thresholds(ratio, scoring["apartment_storehouse_ratio"])

    # 3. Соотношение цены м²
    storehouses = bd.dev_storehouses or bd.domrf_storehouses
    store_ppms = [s["price_per_meter"] for s in storehouses if s.get("price_per_meter")]
    if store_ppms and bd.dev_avg_apt_ppm:
        price_ratio = bd.dev_avg_apt_ppm / (sum(store_ppms) / len(store_ppms))
        pts += _score_by_thresholds(price_ratio, scoring["price_ratio"])

    # 4. Квартирография
    total = bd.domrf_apt_count or 1
    kvart = scoring["kvartirografia"]
    s1k = (bd.rooms_count.get(0, 0) + bd.rooms_count.get(1, 0)) / total * 100
    two_k = bd.rooms_count.get(2, 0) / total * 100
    three_4k = (bd.rooms_count.get(3, 0) + bd.rooms_count.get(4, 0)) / total * 100
    pts += _score_by_thresholds(s1k, kvart["studio_1k"])
    pts += _score_by_thresholds(two_k, kvart["two_k"])
    pts += _score_by_thresholds(three_4k, kvart["three_4k"])

    # 5. Средняя площадь
    area_cfg = scoring["avg_area"]
    room_keys = [(0, "studio"), (1, "one_k"), (2, "two_k"), (3, "three_k"), (4, "four_k")]
    for rooms, cfg_key in room_keys:
        avg = bd.avg_area.get(rooms, 0)
        if avg > 0:
            pts += _score_by_thresholds(avg, area_cfg[cfg_key])

    # 6. Средняя нежилая площадь
    nl_cfg = scoring["avg_non_living"]
    for rooms, cfg_key in room_keys:
        avg_nl = bd.avg_non_living.get(rooms, 0)
        if avg_nl > 0:
            pts += _score_by_thresholds(avg_nl, nl_cfg[cfg_key])

    # 7-10 — ручные, пока 0
    return pts


def calc_second_stage(store: dict, scoring: dict) -> int:
    """Второй этап: критерии 11-13 (площадь, стоимость, цена/м² кладовки)."""
    pts = 0
    area = store.get("area") or 0
    price = store.get("price") or 0
    ppm = store.get("price_per_meter") or 0

    if area > 0:
        pts += _score_by_thresholds(area, scoring["storehouse_area"])
    if price > 0:
        pts += _score_by_thresholds(price, scoring["storehouse_price"])
    if ppm > 0:
        pts += _score_by_thresholds(ppm, scoring["storehouse_ppm"])
    return pts


# ─── Утилиты xlsx ────────────────────────────────────────

def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="4472C4"))


def _apply_segment_border(ws, row: int, max_col: int):
    """Применить жирную нижнюю границу ко всей строке (конец сегмента)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        # Сохранить вертикальные разделители, добавить горизонтальную
        old = cell.border
        cell.border = Border(
            left=old.left, right=old.right,
            bottom=Side(style="medium", color="4472C4"),
        )


def _apply_vertical_dividers(ws, start_row: int, end_row: int,
                             right_cols: set, left_cols: set):
    """Применить сплошные вертикальные разделители ко ВСЕМ строкам диапазона."""
    thick = Side(style="medium", color="4472C4")
    for row in range(start_row, end_row + 1):
        for col in right_cols:
            cell = ws.cell(row=row, column=col)
            old = cell.border
            cell.border = Border(
                left=old.left, right=thick,
                top=old.top, bottom=old.bottom,
            )
        for col in left_cols:
            cell = ws.cell(row=row, column=col)
            old = cell.border
            cell.border = Border(
                left=thick, right=old.right,
                top=old.top, bottom=old.bottom,
            )


def _write_cell(ws, row, col, value, is_manual=False, is_score=False, is_link=False, is_gray=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGN
    # Без рамки по умолчанию — только вертикальные разделители (применяются отдельно)
    if is_manual:
        cell.fill = MANUAL_FILL
    elif is_score:
        cell.fill = SCORE_FILL
    if is_gray:
        cell.font = GRAY_FONT
    if is_link and value:
        cell.font = LINK_FONT
    return cell


# ─── Smart Merge (ручные данные) ─────────────────────────

def _read_manual_jk(old_path: Path) -> dict[tuple, dict]:
    """Прочитать ручные данные листа 'жк' из предыдущего файла."""
    if not old_path.exists():
        return {}
    wb = load_workbook(old_path, data_only=True)
    if "жк" not in wb.sheetnames:
        return {}
    ws = wb["жк"]
    manual = {}
    for r in range(3, ws.max_row + 1):
        city = ws.cell(r, 1).value
        cn = ws.cell(r, 3).value
        bld = ws.cell(r, 4).value
        if not city and not cn:
            continue
        key = _match_key(str(city or ""), str(cn or ""), str(bld or ""))
        data = {}
        # 9=балконы, 10=размер, 11=локация, 12=доступ, 14=балконы/квартиры
        for col_idx, name in [(9, "balcony_count"), (10, "balcony_size"),
                               (11, "location"), (12, "access"),
                               (14, "balcony_ratio")]:
            val = ws.cell(r, col_idx).value
            if val is not None:
                data[name] = val
        if data:
            manual[key] = data
    wb.close()
    return manual


# ─── ЛИСТ 1: кладовки (застройщик) ───────────────────────

STORE_HEADERS = [
    "Город",                                      # 1
    "Застройщик",                                  # 2
    "ЖК",                                          # 3
    "Корпус",                                       # 4
    "Дней до сдачи",                                # 5
    "Соотнош. квартир/кладовых",                    # 6
    "Соотнош. цены м² квартир/кладовых",            # 7
    "Соотнош. квартир к балконам",                   # 8 (формула из жк)
    "Первый этап",                                  # 9 (пока не заполнять)
    "Второй этап",                                  # 10 (пока не заполнять)
    "Общие баллы",                                  # 11 (пока не заполнять)
    "Площадь (м²)",                                 # 12
    "Стоимость (₽)",                                # 13
    "Цена/м² (₽)",                                  # 14
    "Номер кладовой",                               # 15
    "Ссылка",                                       # 16
    "Исх. порядок",                                 # 17
]

STORE_WIDTHS = [12, 16, 22, 14, 12, 18, 20, 20, 12, 12, 12, 10, 14, 14, 12, 12, 10]


def _fill_storehouses_sheet(ws, buildings: list[BuildingAgg]):
    """Лист 1: кладовки (застройщик). Без баллов — они будут добавлены формулами позже."""

    # Подзаголовки групп
    ws.merge_cells("A1:D1")
    ws.cell(1, 1, "Основные данные ЖК").fill = SUBHEADER_FILL
    ws.cell(1, 1).font = SUBHEADER_FONT
    ws.cell(1, 1).alignment = DATA_ALIGN
    ws.merge_cells("I1:K1")
    ws.cell(1, 9, "Баллы").fill = SUBHEADER_FILL
    ws.cell(1, 9).font = SUBHEADER_FONT
    ws.cell(1, 9).alignment = DATA_ALIGN
    ws.merge_cells("L1:Q1")
    ws.cell(1, 12, "Данные кладовой").fill = SUBHEADER_FILL
    ws.cell(1, 12).font = SUBHEADER_FONT
    ws.cell(1, 12).alignment = DATA_ALIGN

    for c, h in enumerate(STORE_HEADERS, 1):
        ws.cell(2, c, h)
    _style_header(ws, 2, len(STORE_HEADERS))
    for c, w in enumerate(STORE_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 3
    order = 1

    for bd in buildings:
        storehouses = bd.dev_storehouses if bd.dev_storehouses else bd.domrf_storehouses
        if not storehouses:
            continue

        store_total = bd.domrf_store_count or bd.dev_store_count
        apt_store_ratio = round(bd.domrf_apt_count / store_total, 1) if store_total else None

        # Соотношение цен: ср. цена м² квартиры (застр.) / ср. цена м² кладовки
        store_ppms = [s["price_per_meter"] for s in storehouses if s.get("price_per_meter")]
        avg_store_ppm = sum(store_ppms) / len(store_ppms) if store_ppms else 0
        price_ratio = (
            round(bd.dev_avg_apt_ppm / avg_store_ppm, 1)
            if avg_store_ppm and bd.dev_avg_apt_ppm else None
        )

        for store in storehouses:
            has_price = bool(store.get("price"))

            _write_cell(ws, row, 1, bd.city)
            _write_cell(ws, row, 2, bd.developer)
            _write_cell(ws, row, 3, bd.complex_name)
            _write_cell(ws, row, 4, bd.building)
            _write_cell(ws, row, 5, bd.days_until)
            _write_cell(ws, row, 6, apt_store_ratio)
            _write_cell(ws, row, 7, price_ratio)

            # 8: Соотнош. квартир к балконам — формула из листа жк
            # =IFERROR(жк!O{row_jk} / жк!I{row_jk}, "")
            # Находим строку корпуса на листе жк по порядку buildings
            jk_row = bd._jk_row if hasattr(bd, '_jk_row') else None
            if jk_row:
                formula = f'=IFERROR(жк!O{jk_row}/жк!I{jk_row},"")'
                _write_cell(ws, row, 8, formula)
            else:
                _write_cell(ws, row, 8, None)

            # 9-11: этапы баллов (пока пустые)
            _write_cell(ws, row, 9, None)
            _write_cell(ws, row, 10, None)
            _write_cell(ws, row, 11, None)

            _write_cell(ws, row, 12, store.get("area"))
            _write_cell(ws, row, 13, store.get("price") if has_price else "–")
            _write_cell(ws, row, 14, store.get("price_per_meter") if has_price else "–")
            _write_cell(ws, row, 15, store.get("item_number"))

            # Ссылка — приоритет ДОМ.РФ, fallback на застройщика
            url = bd.domrf_link or store.get("url")
            cell_link = _write_cell(ws, row, 16, "Открыть" if url else None)
            if url:
                cell_link.hyperlink = url
                cell_link.font = LINK_FONT

            _write_cell(ws, row, 17, order, is_gray=True)
            order += 1
            row += 1

    # Жирная линия между сегментами (смена ЖК + корпус)
    prev_key = None
    for r in range(3, row):
        cn = ws.cell(r, 3).value
        bld = ws.cell(r, 4).value
        cur_key = (cn, bld)
        if prev_key is not None and cur_key != prev_key:
            _apply_segment_border(ws, r - 1, len(STORE_HEADERS))
        prev_key = cur_key

    # Вертикальные разделители для листа кладовок
    # Разделители: после корпуса(4), после балконов(8), после баллов(11),
    # после ссылки(16), после исх.порядка(17)
    store_right = {4, 8, 11, 16, 17}
    store_left = {5, 9, 12, 17}
    _apply_vertical_dividers(ws, 2, row - 1, store_right, store_left)

    ws.freeze_panes = "E3"
    if row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(STORE_HEADERS))}{row - 1}"


# ─── ЛИСТ 2: жк ─────────────────────────────────────────

JK_HEADERS = [
    "Город",                                       # 1
    "Застройщик",                                   # 2
    "ЖК",                                           # 3
    "Корпус",                                        # 4
    "Дней до сдачи",                                 # 5
    "Ссылка на ЖК в ДОМ.РФ",                         # 6
    "Ссылка на сайт застройщика",                     # 7
    "Первый этап",                                   # 8 (не заполнять)
    "Кол-во балконов",                               # 9 (РУЧНОЙ)
    "Размер балконов",                               # 10 (РУЧНОЙ)
    "Локация ЖК",                                    # 11 (РУЧНОЙ)
    "Удобство доступа",                              # 12 (РУЧНОЙ)
    "Соотнош. квартир/кладовых (ДОМ.РФ)",             # 13
    "Соотнош. квартир к балконам",                    # 14 (РУЧНОЙ)
    "Кол-во квартир (всего, ДОМ.РФ)",                 # 15
    "Кол-во кладовых (всего, ДОМ.РФ)",                # 16
    "Остаток кладовок (застройщик)",                  # 17
    "Средняя цена м² квартиры (застройщик)",          # 18
    "Студии", "1к", "2к", "3к", "4к+",              # 19-23
    "Ср. площадь (ст.)", "Ср. площадь (1к)",        # 24-28
    "Ср. площадь (2к)", "Ср. площадь (3к)", "Ср. площадь (4к+)",
    "Ср. нежилая (ст.)", "Ср. нежилая (1к)",        # 29-33
    "Ср. нежилая (2к)", "Ср. нежилая (3к)", "Ср. нежилая (4к+)",
    "Исх. порядок",                                  # 34
]

MANUAL_JK_COLS = {9, 10, 11, 12}  # жёлтые (14 теперь формула)
EMPTY_JK_COLS = {8}                   # «первый этап» — не заполнять

JK_WIDTHS = (
    [12, 16, 22, 14, 12, 22, 22, 12]       # 1-8
    + [14, 14, 14, 14]                      # 9-12 (manual)
    + [20, 18, 18, 18, 16, 18]              # 13-18
    + [8, 8, 8, 8, 8]                       # 19-23 (kvartirografia)
    + [12, 12, 12, 12, 12]                  # 24-28 (avg area)
    + [12, 12, 12, 12, 12]                  # 29-33 (avg non-living)
    + [10]                                  # 34 (order)
)


def _fill_jk_sheet(
    ws,
    buildings: list[BuildingAgg],
    manual: dict[tuple, dict],
    dev_urls: dict[tuple, str],
    living_areas: dict[tuple, dict],
):
    """Лист 2: жк."""

    # Подзаголовки строка 1
    groups = [
        (1, 4, "Основные данные"),
        (9, 12, "Заполняет заказчик"),
        (13, 18, "Сводные данные"),
        (19, 23, "Квартирография (ДОМ.РФ)"),
        (24, 28, "Средняя площадь (ДОМ.РФ)"),
        (29, 33, "Средняя нежилая площадь (ДОМ.РФ)"),
    ]
    for start, end, title in groups:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, title)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = DATA_ALIGN

    # Заголовки строка 2
    for c, h in enumerate(JK_HEADERS, 1):
        ws.cell(2, c, h)
    _style_header(ws, 2, len(JK_HEADERS))

    for c, w in enumerate(JK_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Data validation для ручных столбцов
    dv_size = DataValidation(type="list", formula1='"–,С,Б,М"', allow_blank=True)
    dv_size.error = "Выберите: С, Б, М или – для сброса"
    dv_loc = DataValidation(type="list", formula1='"–,Центр,Норм,Окраина,Фу"', allow_blank=True)
    dv_loc.error = "Выберите: Центр, Норм, Окраина, Фу или – для сброса"
    dv_access = DataValidation(type="list", formula1='"–,Лифт,Улица,Лестница,Лифт+Улица"', allow_blank=True)
    dv_access.error = "Выберите: Лифт, Улица, Лестница, Лифт+Улица или – для сброса"
    ws.add_data_validation(dv_size)
    ws.add_data_validation(dv_loc)
    ws.add_data_validation(dv_access)

    for i, bd in enumerate(buildings):
        row = i + 3
        order = i + 1

        # Сохраняем номер строки на листе жк для формул на листе кладовок
        bd._jk_row = row

        key = _match_key(bd.city, bd.complex_name, bd.building)
        m = manual.get(key, {})
        ck = _complex_key(bd.city, bd.complex_name)

        store_total = bd.domrf_store_count or bd.dev_store_count
        apt_store_ratio = round(bd.domrf_apt_count / store_total, 1) if store_total else None

        # Ссылка застройщика из yaml (base_url по имени застройщика)
        dev_url = dev_urls.get(bd.developer)

        # Формула: кол-во квартир (O) / кол-во балконов (I), с защитой от ошибки
        balcony_ratio_formula = f'=IFERROR(O{row}/I{row},"")'

        values = {
            1: bd.city,
            2: bd.developer,
            3: bd.complex_name,
            4: bd.building,
            5: bd.days_until,
            6: bd.domrf_link or None,
            7: dev_url or None,
            8: None,  # первый этап — не заполнять
            9: m.get("balcony_count"),
            10: m.get("balcony_size"),
            11: m.get("location"),
            12: m.get("access"),
            13: apt_store_ratio,
            14: balcony_ratio_formula,
            15: bd.domrf_apt_count or 0,
            16: bd.domrf_store_count or 0,
            17: bd.dev_store_count or 0,
            18: round(bd.dev_avg_apt_ppm) if bd.dev_avg_apt_ppm else 0,
        }

        # 19-23: квартирография
        for rooms in range(5):
            cnt = bd.rooms_count.get(rooms, 0)
            values[19 + rooms] = cnt or 0

        # 24-28: средняя площадь (из дом.рф)
        la_data = living_areas.get(ck, {})
        for rooms in range(5):
            avg = bd.avg_area.get(rooms, 0)
            # Предпочитаем данные из xlsx (более полные)
            if ck in living_areas and rooms in la_data:
                avg = la_data[rooms][0]  # (avg_area, avg_living)
            values[24 + rooms] = round(avg, 1) if avg else 0

        # 29-33: средняя нежилая площадь = общая - жилая (из xlsx дом.рф)
        for rooms in range(5):
            if rooms in la_data:
                avg_area_val, avg_living_val = la_data[rooms]
                non_living = round(avg_area_val - avg_living_val, 1)
                values[29 + rooms] = non_living if non_living > 0 else 0
            else:
                avg_nl = bd.avg_non_living.get(rooms, 0)
                values[29 + rooms] = round(avg_nl, 1) if avg_nl else 0

        values[34] = order

        for col_idx, val in values.items():
            is_manual = col_idx in MANUAL_JK_COLS
            is_empty = col_idx in EMPTY_JK_COLS
            is_gray = col_idx == 34
            _write_cell(ws, row, col_idx, val, is_manual=is_manual or is_empty, is_gray=is_gray)

        # Ссылки
        if bd.domrf_link:
            cell = ws.cell(row, 6)
            cell.hyperlink = bd.domrf_link
            cell.font = LINK_FONT
            cell.value = "ДОМ.РФ"
        if dev_url:
            cell = ws.cell(row, 7)
            cell.hyperlink = dev_url
            cell.font = LINK_FONT
            cell.value = "Застройщик"

        # Dropdown валидация
        dv_size.add(ws.cell(row, 10))
        dv_loc.add(ws.cell(row, 11))
        dv_access.add(ws.cell(row, 12))

    # Жирная линия между сегментами (смена ЖК)
    last_data_row = len(buildings) + 2
    prev_cn = None
    for r in range(3, last_data_row + 1):
        cn = ws.cell(r, 3).value
        if prev_cn is not None and cn != prev_cn:
            _apply_segment_border(ws, r - 1, len(JK_HEADERS))
        prev_cn = cn

    # Сплошные вертикальные разделители по ВСЕМ строкам
    _apply_vertical_dividers(ws, 2, last_data_row, JK_BLOCK_DIVIDERS, JK_BLOCK_LEFT_DIVIDERS)

    ws.freeze_panes = "E3"
    if last_data_row > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(JK_HEADERS))}{last_data_row}"


# ─── Главная функция ─────────────────────────────────────

def export_eva_xlsx(
    conn_apt: sqlite3.Connection,
    conn_store: sqlite3.Connection,
    output_path: Path | None = None,
    filter_sites: list[str] | None = None,
) -> Path:
    """
    Сгенерировать единый xlsx расчёта ЕВА.

    Args:
        conn_apt: БД квартир (apartments_history.db)
        conn_store: БД кладовок (history.db)
        output_path: путь вывода (по умолчанию: расчет_ева.xlsx)
        filter_sites: фильтр сайтов для теста (None = все)
    """
    eva_config = load_eva_config()
    scoring = eva_config["scoring"]

    if output_path is None:
        output_path = PROJECT_DIR / eva_config["output_filename"]

    # ── 1. Загрузка ──
    logger.info("Загрузка данных...")

    # Квартиры и кладовки дом.рф
    domrf_apts = _load_apartments(conn_apt, "domrf")
    domrf_stores = _load_storehouses(conn_store, "domrf")

    # Квартиры и кладовки застройщиков (всё кроме domrf)
    all_apts = _load_apartments(conn_apt)
    all_stores = _load_storehouses(conn_store)
    dev_apts = [a for a in all_apts if a["site"] != "domrf"]
    dev_stores = [s for s in all_stores if s["site"] != "domrf"]

    if filter_sites:
        dev_apts = [a for a in dev_apts if a["site"] in filter_sites]
        dev_stores = [s for s in dev_stores if s["site"] in filter_sites]

    logger.info("Квартиры: %d (ДОМ.РФ) + %d (застройщики)", len(domrf_apts), len(dev_apts))
    logger.info("Кладовки: %d (ДОМ.РФ) + %d (застройщики)", len(domrf_stores), len(dev_stores))

    # ── 2. ObjectInfo ──
    object_infos = load_object_infos()

    # ── 3. Агрегация ──
    buildings = _aggregate(domrf_apts, domrf_stores, dev_apts, dev_stores, object_infos)
    logger.info("Корпусов: %d", len(buildings))

    # ── 4. Smart merge ручных данных ──
    manual = _read_manual_jk(output_path)
    if manual:
        logger.info("Загружено %d записей ручных данных из предыдущего файла", len(manual))

    # ── 5. Доп. данные ──
    dev_urls = load_developer_urls()
    logger.info("Загружено %d URL-ов застройщиков", len(dev_urls))

    living_areas = load_domrf_living_areas()

    # ── 6. Генерация XLSX ──
    wb = Workbook()

    # Назначить _jk_row заранее (нужно для формул на листе кладовок)
    for i, bd in enumerate(buildings):
        bd._jk_row = i + 3  # строка 3 = первая строка данных (после 2 строк заголовков)

    ws_store = wb.active
    ws_store.title = "кладовки (застройщик)"
    _fill_storehouses_sheet(ws_store, buildings)

    ws_jk = wb.create_sheet("жк")
    _fill_jk_sheet(ws_jk, buildings, manual, dev_urls, living_areas)

    # ── 6. Сохранение ──
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Файл сохранён: %s", output_path)
    return output_path
