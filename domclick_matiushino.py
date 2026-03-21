#!/usr/bin/env python3
"""Создание xlsx из TSV данных Домклика."""
import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parent
TSV_PATH = Path.home() / "Downloads" / "domclick_matiushino.tsv"
OUT_PATH = PROJECT_DIR / "output" / "apartments" / "domclick_matiushino.xlsx"

HEADERS_RU = {
    "id": "Ссылка",
    "rooms": "Комнат",
    "flatNumber": "№ кв.",
    "area": "Площадь, м²",
    "areaLiving": "Жилая, м²",
    "areaKitchen": "Кухня, м²",
    "floor": "Этаж",
    "balconiesNum": "Балконы",
    "loggiasNum": "Лоджии",
    "connectedBathrooms": "Совм. санузлы",
    "separatedBathrooms": "Разд. санузлы",
    "renovation": "Отделка",
    "windowView": "Вид из окна",
    "isEuroplan": "Европланировка",
    "price": "Цена, ₽",
    "pricePerM2": "Цена/м², ₽",
}

INT_COLS = {"rooms", "balconiesNum", "loggiasNum", "connectedBathrooms", "separatedBathrooms", "price", "pricePerM2"}
FLOAT_COLS = {"area", "areaLiving", "areaKitchen"}

wb = Workbook()
ws = wb.active
ws.title = "ЖК Матюшино парк"

# Read TSV
with open(TSV_PATH, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter="\t")
    fields = reader.fieldnames
    rows = list(reader)

# Header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for col_idx, field in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_idx, value=HEADERS_RU.get(field, field))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Data
data_font = Font(name="Calibri", size=11)
data_align = Alignment(horizontal="center", vertical="center")

link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")

for row_idx, row in enumerate(rows, 2):
    for col_idx, field in enumerate(fields, 1):
        val = row.get(field, "")

        if field == "id" and val:
            url = f"https://kazan.domclick.ru/card/sale__new_flat__{val}"
            cell = ws.cell(row=row_idx, column=col_idx, value="Открыть")
            cell.hyperlink = url
            cell.font = link_font
            cell.alignment = data_align
            cell.border = thin_border
            continue

        if field in INT_COLS and val:
            try: val = int(val)
            except ValueError: pass
        elif field in FLOAT_COLS and val:
            try: val = float(val)
            except ValueError: pass
        elif field == "isEuroplan":
            val = "Да" if val == "true" else "Нет"

        cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

# Column widths
widths = [14, 8, 8, 12, 10, 10, 8, 10, 10, 14, 14, 16, 14, 16, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Price format
for row_idx in range(2, len(rows) + 2):
    for col in [15, 16]:  # price, pricePerM2
        cell = ws.cell(row=row_idx, column=col)
        if cell.value:
            cell.number_format = '#,##0'

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT_PATH))
print(f"Saved: {OUT_PATH}")
print(f"Rows: {len(rows)}")
