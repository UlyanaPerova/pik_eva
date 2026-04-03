"""
Скрипт валидации данных: проверяет согласованность БД, XLSX и конфигов.

Использование:
    python3.13 validate_data.py              # проверить всё
    python3.13 validate_data.py --apartments  # только квартиры
    python3.13 validate_data.py --storehouses # только кладовки
    python3.13 validate_data.py --eva         # только EVA-калькулятор

Проверки:
    1. Конфиг ↔ БД: все ли object_id из конфига имеют данные в БД
    2. БД → XLSX: совпадает ли количество и содержание данных
    3. Внутренняя согласованность БД: дубли, аномалии, пропуски
    4. EVA: корректность агрегации (кол-во квартир/кладовок по корпусам)
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path

import yaml
from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parent
CONFIGS_DIR = PROJECT_DIR / "configs"
APT_DIR = PROJECT_DIR / "apartments"
STORE_DIR = PROJECT_DIR / "output"
DATA_DIR = PROJECT_DIR / "data"

# ─── Цвета для вывода ───────────────────────────────────

RED = "\033[91m"
YELLOW = "\033[93m"
GREEN = "\033[92m"
GRAY = "\033[90m"
RESET = "\033[0m"
BOLD = "\033[1m"

errors: list[str] = []
warnings: list[str] = []
infos: list[str] = []


def error(msg: str) -> None:
    errors.append(msg)
    print(f"  {RED}ОШИБКА{RESET}: {msg}")


def warn(msg: str) -> None:
    warnings.append(msg)
    print(f"  {YELLOW}ВНИМАНИЕ{RESET}: {msg}")


def info(msg: str) -> None:
    infos.append(msg)
    print(f"  {GREEN}OK{RESET}: {msg}")


# ═══════════════════════════════════════════════════════
#  1. ВАЛИДАЦИЯ КОНФИГОВ
# ═══════════════════════════════════════════════════════

def validate_configs() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  1. ВАЛИДАЦИЯ КОНФИГОВ{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    # domrf.yaml (кладовки)
    store_cfg_path = CONFIGS_DIR / "domrf.yaml"
    apt_cfg_path = CONFIGS_DIR / "domrf_apartments.yaml"

    store_ids: set[int] = set()
    apt_ids: set[int] = set()

    if store_cfg_path.exists():
        with open(store_cfg_path, encoding="utf-8") as f:
            store_cfg = yaml.safe_load(f)
        store_links = store_cfg.get("links", [])
        print(f"\n  domrf.yaml: {len(store_links)} ссылок")

        for i, link in enumerate(store_links):
            oid = link.get("object_id")
            cn = link.get("complex_name", "")
            if not oid:
                error(f"domrf.yaml link #{i+1}: отсутствует object_id")
                continue
            if not cn:
                error(f"domrf.yaml link #{i+1} (object_id={oid}): отсутствует complex_name")
            if not link.get("city"):
                warn(f"domrf.yaml object_id={oid} ({cn}): нет city, по умолчанию 'Казань'")
            store_ids.add(oid)
    else:
        error("Файл configs/domrf.yaml не найден")

    if apt_cfg_path.exists():
        with open(apt_cfg_path, encoding="utf-8") as f:
            apt_cfg = yaml.safe_load(f)
        apt_links = apt_cfg.get("links", [])
        print(f"  domrf_apartments.yaml: {len(apt_links)} ссылок")

        for i, link in enumerate(apt_links):
            oid = link.get("object_id")
            cn = link.get("complex_name", "")
            bld = link.get("building", "")
            if not oid:
                error(f"domrf_apartments.yaml link #{i+1}: отсутствует object_id")
                continue
            if not cn:
                error(f"domrf_apartments.yaml link #{i+1} (object_id={oid}): отсутствует complex_name")
            if not bld:
                warn(f"domrf_apartments.yaml object_id={oid} ({cn}): нет building — будет использован API fallback")
            apt_ids.add(oid)

        # Проверка: одинаковые ли наборы object_id
        only_store = store_ids - apt_ids
        only_apt = apt_ids - store_ids
        if only_store:
            warn(f"object_id только в domrf.yaml (нет в apartments): {only_store}")
        if only_apt:
            warn(f"object_id только в domrf_apartments.yaml (нет в storehouses): {only_apt}")
        if not only_store and not only_apt:
            info(f"Наборы object_id совпадают: {len(store_ids)} шт.")
    else:
        error("Файл configs/domrf_apartments.yaml не найден")


# ═══════════════════════════════════════════════════════
#  2. ВАЛИДАЦИЯ БД КВАРТИР
# ═══════════════════════════════════════════════════════

def validate_apartments_db() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  2. ВАЛИДАЦИЯ БД КВАРТИР{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    db_path = DATA_DIR / "apartments" / "apartments_history.db"
    if not db_path.exists():
        error(f"БД квартир не найдена: {db_path}")
        return

    conn = sqlite3.connect(str(db_path))

    # Общая статистика
    total = conn.execute("SELECT COUNT(*) FROM apartment_prices").fetchone()[0]
    distinct = conn.execute("SELECT COUNT(DISTINCT item_id) FROM apartment_prices").fetchone()[0]
    sites = conn.execute("SELECT DISTINCT site FROM apartment_prices").fetchall()
    site_list = [r[0] for r in sites]
    print(f"\n  Всего записей: {total}, уникальных item_id: {distinct}")
    print(f"  Сайты: {', '.join(site_list)}")

    # Per-site статистика
    for site in site_list:
        row = conn.execute(
            "SELECT COUNT(*), COUNT(DISTINCT item_id), COUNT(DISTINCT complex_name) "
            "FROM apartment_prices WHERE site = ?", (site,)
        ).fetchone()
        print(f"  {site}: {row[0]} записей, {row[1]} уникальных, {row[2]} ЖК")

    # Проверка: аномальные данные
    print(f"\n  {BOLD}Проверка аномалий:{RESET}")

    # Нулевая площадь
    zero_area = conn.execute(
        "SELECT COUNT(*) FROM apartment_prices WHERE area <= 0 OR area IS NULL"
    ).fetchone()[0]
    if zero_area:
        warn(f"{zero_area} квартир с нулевой/пустой площадью")
    else:
        info("Все квартиры имеют площадь > 0")

    # Подозрительно большая площадь
    big_area = conn.execute(
        "SELECT site, complex_name, building, item_id, area "
        "FROM apartment_prices WHERE area > 200 LIMIT 5"
    ).fetchall()
    if big_area:
        warn(f"{len(big_area)} квартир с площадью > 200 м²:")
        for r in big_area:
            print(f"    {r[0]}/{r[1]}/{r[2]} item={r[3]} area={r[4]}")

    # Нулевой rooms для не-студий
    zero_rooms = conn.execute(
        "SELECT COUNT(*) FROM apartment_prices WHERE rooms < 0 OR rooms > 6"
    ).fetchone()[0]
    if zero_rooms:
        warn(f"{zero_rooms} квартир с rooms < 0 или > 6")
    else:
        info("Все rooms в допустимом диапазоне (0-6)")

    # Дублирование item_id (одинаковые item_id с разными данными)
    dupes = conn.execute("""
        SELECT item_id, COUNT(DISTINCT site || '|' || complex_name) as cnt
        FROM apartment_prices
        GROUP BY item_id
        HAVING cnt > 1
        LIMIT 5
    """).fetchall()
    if dupes:
        warn(f"item_id встречается в разных ЖК/сайтах: {len(dupes)} шт.")
        for d in dupes:
            print(f"    item_id={d[0]} в {d[1]} разных (site, complex)")
    else:
        info("Нет конфликтующих item_id между сайтами")

    # Проверка: building содержит конфигурационное имя для domrf
    domrf_buildings = conn.execute("""
        SELECT DISTINCT building, complex_name, COUNT(*) as cnt
        FROM apartment_prices
        WHERE site = 'domrf'
        GROUP BY building, complex_name
        ORDER BY complex_name, building
    """).fetchall()

    api_format_count = 0
    config_format_count = 0
    import re
    for bld, cn, cnt in domrf_buildings:
        base = bld.split("||")[0].strip() if "||" in bld else bld.strip()
        # API-формат: только цифры и точки (напр. "1", "3.1")
        if re.match(r'^\d+(\.\d+)?$', base):
            api_format_count += cnt
        else:
            config_format_count += cnt

    if api_format_count > 0 and config_format_count > 0:
        warn(f"ДОМ.РФ: смешанные форматы корпусов — {api_format_count} записей в API-формате ('1', '2'), "
             f"{config_format_count} в конфиг-формате. Рекомендуется перепарсить для обновления building.")
    elif api_format_count > 0:
        warn(f"ДОМ.РФ: все {api_format_count} записей в API-формате ('1||подъезд N'). "
             f"Необходим повторный парсинг с текущим конфигом для корректной привязки по корпусам.")
    elif config_format_count > 0:
        info(f"ДОМ.РФ: все {config_format_count} записей в конфиг-формате (корректно)")

    conn.close()


# ═══════════════════════════════════════════════════════
#  3. ВАЛИДАЦИЯ БД КЛАДОВОК
# ═══════════════════════════════════════════════════════

def validate_storehouses_db() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  3. ВАЛИДАЦИЯ БД КЛАДОВОК{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    db_path = DATA_DIR / "history.db"
    if not db_path.exists():
        error(f"БД кладовок не найдена: {db_path}")
        return

    conn = sqlite3.connect(str(db_path))

    total = conn.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
    distinct = conn.execute("SELECT COUNT(DISTINCT item_id) FROM prices").fetchone()[0]
    sites = conn.execute("SELECT DISTINCT site FROM prices").fetchall()
    site_list = [r[0] for r in sites]
    print(f"\n  Всего записей: {total}, уникальных item_id: {distinct}")
    print(f"  Сайты: {', '.join(site_list)}")

    for site in site_list:
        row = conn.execute(
            "SELECT COUNT(*), COUNT(DISTINCT item_id), COUNT(DISTINCT complex_name) "
            "FROM prices WHERE site = ?", (site,)
        ).fetchone()
        print(f"  {site}: {row[0]} записей, {row[1]} уникальных, {row[2]} ЖК")

    print(f"\n  {BOLD}Проверка аномалий:{RESET}")

    # Нулевая площадь
    zero_area = conn.execute(
        "SELECT COUNT(*) FROM prices WHERE area <= 0 OR area IS NULL"
    ).fetchone()[0]
    if zero_area:
        warn(f"{zero_area} кладовок с нулевой/пустой площадью")
    else:
        info("Все кладовки имеют площадь > 0")

    # Подозрительно большая площадь
    big_area = conn.execute(
        "SELECT site, complex_name, building, item_id, area "
        "FROM prices WHERE area > 30 LIMIT 5"
    ).fetchall()
    if big_area:
        warn(f"Кладовки с площадью > 30 м² (может быть не кладовка):")
        for r in big_area:
            print(f"    {r[0]}/{r[1]}/{r[2]} item={r[3]} area={r[4]}")
    else:
        info("Нет кладовок с подозрительно большой площадью")

    # Нулевая цена (для не-domrf)
    zero_price = conn.execute(
        "SELECT site, COUNT(*) FROM prices "
        "WHERE (price <= 0 OR price IS NULL) AND site != 'domrf' "
        "GROUP BY site"
    ).fetchall()
    if zero_price:
        for site, cnt in zero_price:
            warn(f"{site}: {cnt} кладовок с нулевой ценой")
    else:
        info("Все кладовки застройщиков имеют цену > 0")

    # Рассогласованность price_per_meter = price / area
    ppm_mismatch = conn.execute("""
        SELECT site, complex_name, item_id, area, price, price_per_meter,
               ABS(price / area - price_per_meter) as diff
        FROM prices
        WHERE price > 0 AND area > 0 AND price_per_meter > 0
              AND ABS(price / area - price_per_meter) > 100
        LIMIT 5
    """).fetchall()
    if ppm_mismatch:
        warn(f"Рассогласованность price_per_meter vs price/area (разница > 100 ₽):")
        for r in ppm_mismatch:
            calc_ppm = r[4] / r[3]
            print(f"    {r[0]}/{r[1]} item={r[2]} area={r[3]} price={r[4]:.0f} "
                  f"ppm_stored={r[5]:.0f} ppm_calc={calc_ppm:.0f} diff={r[6]:.0f}")
    else:
        info("price_per_meter согласована с price/area")

    conn.close()


# ═══════════════════════════════════════════════════════
#  4. ВАЛИДАЦИЯ БД ↔ XLSX
# ═══════════════════════════════════════════════════════

def validate_db_vs_xlsx() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  4. ВАЛИДАЦИЯ БД ↔ XLSX{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    # Квартиры
    _check_apartments_xlsx()
    # Кладовки
    _check_storehouses_xlsx()


def _check_apartments_xlsx() -> None:
    db_path = DATA_DIR / "apartments" / "apartments_history.db"
    if not db_path.exists():
        return

    conn = sqlite3.connect(str(db_path))

    # Проверяем каждый XLSX-файл квартир
    xlsx_files = sorted(APT_DIR.glob("apartments_*.xlsx"))
    for xlsx_path in xlsx_files:
        if xlsx_path.name.startswith("~$"):
            continue
        site_name = xlsx_path.stem.replace("apartments_", "")
        print(f"\n  {BOLD}{xlsx_path.name}{RESET}:")

        # Определяем site key
        site_map = {
            "DomRF": "domrf", "PIK": "pik", "AkBarsDom": "akbarsdom",
            "GloraX": "glorax", "SMU88": "smu88", "Unistroy": "unistroy",
        }
        site_key = site_map.get(site_name)
        if not site_key:
            warn(f"Неизвестный сайт в имени файла: {site_name}")
            continue

        # Количество в БД (последние записи)
        db_count = conn.execute("""
            SELECT COUNT(DISTINCT item_id) FROM apartment_prices
            WHERE site = ?
        """, (site_key,)).fetchone()[0]

        # Количество в XLSX (лист "Все данные")
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as e:
            warn(f"Не удалось открыть {xlsx_path.name}: {e}")
            continue

        sheet_name = "Все данные"
        if sheet_name not in wb.sheetnames:
            warn(f"Нет листа '{sheet_name}' в {xlsx_path.name}")
            wb.close()
            continue

        ws = wb[sheet_name]
        xlsx_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # первый столбец не пустой
                xlsx_count += 1
        wb.close()

        if db_count == 0 and xlsx_count == 0:
            info(f"{site_key}: БД и XLSX пусты")
        elif abs(db_count - xlsx_count) == 0:
            info(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — совпадает")
        elif abs(db_count - xlsx_count) <= 5:
            info(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — почти совпадает (±{abs(db_count-xlsx_count)})")
        else:
            pct = abs(db_count - xlsx_count) / max(db_count, 1) * 100
            if pct > 10:
                error(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — "
                      f"расхождение {abs(db_count-xlsx_count)} ({pct:.0f}%)")
            else:
                warn(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — "
                     f"расхождение {abs(db_count-xlsx_count)} ({pct:.1f}%)")

    conn.close()


def _check_storehouses_xlsx() -> None:
    db_path = DATA_DIR / "history.db"
    if not db_path.exists():
        return

    conn = sqlite3.connect(str(db_path))

    xlsx_files = sorted(STORE_DIR.glob("storehouses_*.xlsx"))
    for xlsx_path in xlsx_files:
        if xlsx_path.name.startswith("~$"):
            continue
        site_name = xlsx_path.stem.replace("storehouses_", "")
        print(f"\n  {BOLD}{xlsx_path.name}{RESET}:")

        site_map = {
            "DomRF": "domrf", "PIK": "pik", "AkBarsDom": "akbarsdom",
            "GloraX": "glorax", "SMU88": "smu88", "UniStroy": "unistroy",
        }
        site_key = site_map.get(site_name)
        if not site_key:
            warn(f"Неизвестный сайт: {site_name}")
            continue

        db_count = conn.execute("""
            SELECT COUNT(DISTINCT item_id) FROM prices WHERE site = ?
        """, (site_key,)).fetchone()[0]

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as e:
            warn(f"Не удалось открыть {xlsx_path.name}: {e}")
            continue

        sheet_name = "Все данные"
        if sheet_name not in wb.sheetnames:
            warn(f"Нет листа '{sheet_name}' в {xlsx_path.name}")
            wb.close()
            continue

        ws = wb[sheet_name]
        xlsx_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                xlsx_count += 1
        wb.close()

        if db_count == 0 and xlsx_count == 0:
            info(f"{site_key}: БД и XLSX пусты")
        elif abs(db_count - xlsx_count) == 0:
            info(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — совпадает")
        elif abs(db_count - xlsx_count) <= 5:
            info(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — почти совпадает (±{abs(db_count-xlsx_count)})")
        else:
            pct = abs(db_count - xlsx_count) / max(db_count, 1) * 100
            if pct > 10:
                error(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — "
                      f"расхождение {abs(db_count-xlsx_count)} ({pct:.0f}%)")
            else:
                warn(f"{site_key}: БД={db_count}, XLSX={xlsx_count} — "
                     f"расхождение {abs(db_count-xlsx_count)} ({pct:.1f}%)")

    conn.close()


# ═══════════════════════════════════════════════════════
#  5. ВАЛИДАЦИЯ EVA-АГРЕГАЦИИ
# ═══════════════════════════════════════════════════════

def validate_eva() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  5. ВАЛИДАЦИЯ EVA-АГРЕГАЦИИ{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    try:
        import eva_calculator
        eva_calculator.load_eva_config()
    except Exception as e:
        error(f"Не удалось загрузить eva_calculator: {e}")
        return

    db_apt_path = DATA_DIR / "apartments" / "apartments_history.db"
    db_store_path = DATA_DIR / "history.db"

    if not db_apt_path.exists() or not db_store_path.exists():
        error("БД не найдены для EVA валидации")
        return

    conn_apt = sqlite3.connect(str(db_apt_path))
    conn_store = sqlite3.connect(str(db_store_path))

    domrf_apts = eva_calculator._load_apartments(conn_apt, "domrf")
    domrf_stores = eva_calculator._load_storehouses(conn_store, "domrf")
    dev_apts = eva_calculator._load_apartments(conn_apt)
    dev_stores = eva_calculator._load_storehouses(conn_store)
    ois = eva_calculator.load_object_infos()

    print(f"\n  Данные: {len(domrf_apts)} квартир ДОМ.РФ, {len(domrf_stores)} кладовок ДОМ.РФ")
    print(f"  Данные: {len(dev_apts)} всего квартир, {len(dev_stores)} всего кладовок")
    print(f"  ObjectInfo: {len(ois)} записей")

    buildings = eva_calculator._aggregate(
        domrf_apts, domrf_stores, dev_apts, dev_stores, ois
    )

    print(f"  Корпусов после агрегации: {len(buildings)}")
    print(f"\n  {BOLD}Проверка корпусов:{RESET}")

    issues = 0
    for b in buildings:
        total_rooms = sum(b.rooms_count.values()) if b.rooms_count else 0

        # domrf_apt_count подозрительно большой
        if b.domrf_apt_count > 2000:
            error(f"{b.complex_name} / {b.building}: domrf_apt_count={b.domrf_apt_count} (> 2000)")
            issues += 1

        # Квартирография есть, но = 0
        if total_rooms == 0 and b.domrf_apt_count > 0:
            warn(f"{b.complex_name} / {b.building}: квартирография пуста при apt_count={b.domrf_apt_count}")
            issues += 1

        # domrf_apt_count != sum(rooms_count)
        if total_rooms > 0 and b.domrf_apt_count > 0 and total_rooms != b.domrf_apt_count:
            diff = abs(total_rooms - b.domrf_apt_count)
            if diff > max(total_rooms, b.domrf_apt_count) * 0.1:
                warn(f"{b.complex_name} / {b.building}: domrf_apt_count={b.domrf_apt_count} ≠ sum(rooms)={total_rooms}")
                issues += 1

        # Нет ни квартир, ни кладовок
        if (b.domrf_apt_count == 0 and total_rooms == 0
                and b.domrf_store_count == 0 and b.dev_store_count == 0
                and not b.dev_storehouses and not b.domrf_storehouses):
            warn(f"{b.complex_name} / {b.building}: пустая строка — нет ни квартир, ни кладовок")
            issues += 1

        # dev_avg_apt_ppm = 0 (нет данных о ценах)
        if b.dev_avg_apt_ppm == 0 and b.domrf_apt_count > 0:
            # Это нормально для domrf-only комплексов, но стоит отметить
            pass

    if issues == 0:
        info("Все корпуса прошли валидацию")
    else:
        print(f"\n  {YELLOW}Найдено {issues} проблем в агрегации{RESET}")

    # Перекрёстная проверка: данные из БД vs данные в агрегации
    print(f"\n  {BOLD}Перекрёстная проверка БД ↔ EVA:{RESET}")

    # Сумма квартир по комплексам в БД
    db_apt_by_complex = defaultdict(int)
    for apt in domrf_apts:
        key = (apt["city"], apt["complex_name"])
        db_apt_by_complex[key] += 1

    # Для EVA: берём MAX rooms_count по корпусам (не сумму!),
    # т.к. при complex-level fallback все корпуса получают одинаковые rooms_count.
    # Если данные per-building, то каждый корпус имеет свой rooms_count → нужна сумма.
    eva_apt_by_complex = defaultdict(int)
    eva_buildings_by_complex: dict[tuple, list] = defaultdict(list)
    for b in buildings:
        key = (b.city, b.complex_name)
        eva_apt_by_complex[key] += b.domrf_apt_count
        rooms_total = sum(b.rooms_count.values()) if b.rooms_count else 0
        eva_buildings_by_complex[key].append(rooms_total)

    for key in sorted(db_apt_by_complex.keys()):
        db_cnt = db_apt_by_complex[key]
        eva_cnt = eva_apt_by_complex.get(key, 0)
        bld_rooms = eva_buildings_by_complex.get(key, [])
        cn = key[1]

        # Определяем: per-building или complex fallback?
        # Если все корпуса имеют одинаковый rooms_total — это fallback
        unique_rooms = set(bld_rooms)
        if len(unique_rooms) <= 1 and len(bld_rooms) > 1:
            # Complex-level fallback — rooms_count одинаковый для всех корпусов
            eva_rooms = bld_rooms[0] if bld_rooms else 0
            if eva_rooms == db_cnt:
                info(f"{cn}: БД={db_cnt}, EVA rooms={eva_rooms} (complex-level, {len(bld_rooms)} корпусов — нужен повторный парсинг)")
            elif eva_cnt == db_cnt:
                info(f"{cn}: БД={db_cnt}, EVA={eva_cnt} — совпадает")
            else:
                diff = abs(db_cnt - max(eva_cnt, eva_rooms))
                if diff > db_cnt * 0.1:
                    warn(f"{cn}: БД={db_cnt}, EVA apt={eva_cnt}, EVA rooms_each={eva_rooms} ({len(bld_rooms)} корпусов) — расхождение!")
                else:
                    info(f"{cn}: БД={db_cnt}, EVA apt={eva_cnt}, rooms={eva_rooms}")
        else:
            # Per-building — суммируем
            eva_rooms = sum(bld_rooms)
            if eva_cnt == db_cnt or eva_rooms == db_cnt:
                info(f"{cn}: БД={db_cnt}, EVA apt={eva_cnt}, rooms_sum={eva_rooms} — совпадает")
            else:
                diff = abs(db_cnt - max(eva_cnt, eva_rooms))
                if diff > db_cnt * 0.1:
                    warn(f"{cn}: БД={db_cnt}, EVA apt={eva_cnt}, rooms_sum={eva_rooms} — расхождение!")
                else:
                    info(f"{cn}: БД={db_cnt}, EVA apt={eva_cnt}, rooms_sum={eva_rooms}")

    conn_apt.close()
    conn_store.close()


# ═══════════════════════════════════════════════════════
#  6. КОНФИГ ↔ БД (есть ли данные для каждого object_id)
# ═══════════════════════════════════════════════════════

def validate_config_vs_db() -> None:
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  6. КОНФИГ ↔ БД (полнота данных){RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")

    # Загружаем конфиги
    apt_cfg_path = CONFIGS_DIR / "domrf_apartments.yaml"
    store_cfg_path = CONFIGS_DIR / "domrf.yaml"

    if not apt_cfg_path.exists():
        return

    with open(apt_cfg_path, encoding="utf-8") as f:
        apt_cfg = yaml.safe_load(f)

    # Группируем object_id по complex_name
    config_complexes: dict[str, list[int]] = defaultdict(list)
    for link in apt_cfg.get("links", []):
        cn = link.get("complex_name", "")
        oid = link.get("object_id")
        if cn and oid:
            config_complexes[cn].append(oid)

    # Проверяем наличие данных в БД
    db_path = DATA_DIR / "apartments" / "apartments_history.db"
    if not db_path.exists():
        error("БД квартир не найдена")
        return

    conn = sqlite3.connect(str(db_path))

    # Для каждого комплекса из конфига проверяем наличие данных в БД
    db_complexes = set()
    for row in conn.execute(
        "SELECT DISTINCT complex_name FROM apartment_prices WHERE site = 'domrf'"
    ).fetchall():
        db_complexes.add(row[0])

    print(f"\n  Комплексов в конфиге: {len(config_complexes)}")
    print(f"  Комплексов в БД (domrf): {len(db_complexes)}")

    missing_in_db = set(config_complexes.keys()) - db_complexes
    extra_in_db = db_complexes - set(config_complexes.keys())

    if missing_in_db:
        for cn in sorted(missing_in_db):
            oids = config_complexes[cn]
            warn(f"В конфиге, но НЕТ в БД: {cn} (object_ids: {oids}) — нужен парсинг")

    if extra_in_db:
        for cn in sorted(extra_in_db):
            cnt = conn.execute(
                "SELECT COUNT(*) FROM apartment_prices "
                "WHERE site = 'domrf' AND complex_name = ?", (cn,)
            ).fetchone()[0]
            warn(f"В БД, но НЕТ в конфиге: {cn} ({cnt} записей) — удалён из конфига?")

    if not missing_in_db and not extra_in_db:
        info("Все комплексы из конфига есть в БД и наоборот")

    conn.close()


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(description="Валидация данных pik_eva")
    parser.add_argument("--apartments", action="store_true", help="Только квартиры")
    parser.add_argument("--storehouses", action="store_true", help="Только кладовки")
    parser.add_argument("--eva", action="store_true", help="Только EVA")
    args = parser.parse_args()

    run_all = not (args.apartments or args.storehouses or args.eva)

    print(f"{BOLD}╔{'═' * 58}╗{RESET}")
    print(f"{BOLD}║  ВАЛИДАЦИЯ ДАННЫХ PIK_EVA                              ║{RESET}")
    print(f"{BOLD}╚{'═' * 58}╝{RESET}")

    if run_all or args.apartments or args.storehouses:
        validate_configs()

    if run_all or args.apartments:
        validate_apartments_db()

    if run_all or args.storehouses:
        validate_storehouses_db()

    if run_all:
        validate_db_vs_xlsx()

    if run_all or args.eva:
        validate_eva()

    if run_all:
        validate_config_vs_db()

    # Итоги
    print(f"\n{BOLD}{'═' * 60}{RESET}")
    print(f"{BOLD}  ИТОГО{RESET}")
    print(f"{BOLD}{'═' * 60}{RESET}")
    print(f"  {RED}Ошибки: {len(errors)}{RESET}")
    print(f"  {YELLOW}Предупреждения: {len(warnings)}{RESET}")
    print(f"  {GREEN}OK: {len(infos)}{RESET}")

    if errors:
        print(f"\n  {RED}{BOLD}Критические ошибки:{RESET}")
        for e in errors:
            print(f"    {RED}• {e}{RESET}")

    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
