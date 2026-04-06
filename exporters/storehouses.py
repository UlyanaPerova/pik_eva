"""
Экспорт данных в xlsx с примечаниями (история цен, скидки).

Два листа:
  1. «Кладовки» — красивый, со строками-заголовками Город/ЖК,
     разделителями корпусов, секцией в примечаниях. Сортировка внутри каждого ЖК.
  2. «Все данные» — плоская таблица со всеми колонками,
     автофильтром, сортировкой, разделителями корпусов,
     группировкой строк по ЖК для скрытия.

Функции:
  - Скрытие ЖК через группировку строк (outline) на обоих листах
  - Разделители корпусов на обоих листах
  - История цен и скидки в примечаниях
  - Примечание «Добавлена от [дата]» на новых кладовках
  - Столбец «Исх. порядок» для сброса сортировки
"""
from __future__ import annotations

import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import json

from exporters.common import (
    HEADER_FILL, HEADER_FONT,
    CITY_FILL, CITY_FONT,
    COMPLEX_FILL, COMPLEX_FONT,
    DATA_FONT, DATA_ALIGN,
    LINK_FONT, NEW_ITEM_FILL,
    TOTAL_FONT, COMPLEX_TOTAL_FILL, GRAND_TOTAL_FILL,
    THIN_BORDER, BUILDING_BOTTOM,
    SITE_NAMES, SITE_FILE_KEYS,
    append_comment as _append_comment,
    natural_sort_key as _natural_sort_key,
    add_new_item_comment as _add_new_item_comment_common,
    add_price_comment as _add_price_comment_common,
    add_ppm_comment as _add_ppm_comment_common,
)

from parsers.base import (
    StorehouseItem, get_price_history, get_first_seen_date,
    get_all_known_ids, logger, PROJECT_DIR, DATA_DIR,
)

OUTPUT_DIR = PROJECT_DIR / "output"


def _load_or_create_baseline(items: list[StorehouseItem]) -> set[str]:
    """
    Загрузить baseline (ID кладовок из первого парсинга) для каждого сайта отдельно.
    Файлы: data/baseline_pik.json, data/baseline_akbarsdom.json и т.д.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    all_baseline: set[str] = set()

    # Группируем items по сайтам
    sites_items: dict[str, list[str]] = {}
    for it in items:
        sites_items.setdefault(it.site, []).append(it.item_id)

    for site, item_ids in sites_items.items():
        baseline_path = DATA_DIR / f"baseline_{site}.json"
        if baseline_path.exists():
            with open(baseline_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            all_baseline.update(data)
        else:
            # Первый запуск этого сайта — сохраняем текущие ID
            with open(baseline_path, "w", encoding="utf-8") as f:
                json.dump(item_ids, f)
            logger.info("Создан baseline %s: %d кладовок", site, len(item_ids))
            all_baseline.update(item_ids)

    return all_baseline

# ── Стили: импортированы из exporter_common ─────────────
# Локальные алиас�� для совместимости внутри файла
NEW_ITEM_LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
BUILDING_BOTTOM_ALL = BUILDING_BOTTOM


# ── Колонки ──────────────────────────────────────────

PRETTY_COLUMNS = [
    "Корпус",
    "Кол-во\nкладовок",
    "Номер\nкладовой",
    "Площадь\n(м²)",
    "Цена (₽)",
    "Цена/м²\n(₽)",
    "Ссылка",
    "Исх.\nпорядок",
]
PCOL = len(PRETTY_COLUMNS)

FLAT_COLUMNS = [
    "Город",
    "Застройщик",
    "ЖК",
    "Корпус",
    "Кол-во кладовок",
    "Номер кладовой",
    "Площадь (м²)",
    "Цена (₽)",
    "Цена/м² (₽)",
    "Ссылка",
    "Исх. порядок",
]
FCOL = len(FLAT_COLUMNS)


# _natural_sort_key — импортирована из exporter_common


def _sort_key(it: StorehouseItem):
    try:
        num = int(it.item_number) if it.item_number else 999999
    except ValueError:
        num = 999999
    developer = getattr(it, 'developer', '') or ''
    return (
        it.city.lower(),
        developer.lower(),
        it.complex_name.lower(),
        _natural_sort_key(it.building),
        num,
    )


def export_xlsx(
    items: list[StorehouseItem],
    conn: sqlite3.Connection,
    filename: str | None = None,
    previously_known: set[str] | None = None,
) -> Path:
    """Создать xlsx с двумя листами."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        sites_in_data = sorted(set(
            SITE_FILE_KEYS.get(it.site, it.site) for it in items
        ))
        suffix = "_".join(sites_in_data) if sites_in_data else "ALL"
        filename = f"storehouses_{suffix}.xlsx"

    output_path = OUTPUT_DIR / filename

    # Определяем новые кладовки (которых раньше не было в БД)
    if previously_known is None:
        sites = set(it.site for it in items)
        previously_known = set()
        for site in sites:
            previously_known |= get_all_known_ids(conn, site)

    # Baseline — набор ID из самого первого парсинга
    baseline_ids = _load_or_create_baseline(items)

    wb = Workbook()

    ws_pretty = wb.active
    ws_pretty.title = "Кладовки"
    _fill_pretty_sheet(ws_pretty, items, conn, previously_known, baseline_ids)

    ws_flat = wb.create_sheet("Все данные")
    _fill_flat_sheet(ws_flat, items, conn, previously_known, baseline_ids)

    wb.save(output_path)
    logger.info("xlsx сохранён: %s (%d кладовок)", output_path, len(items))
    return output_path


# ══════════════════════════════════════════════════════
#  ЛИСТ 1: «Кладовки» — красивый
# ══════════════════════════════════════════════════════

def _fill_pretty_sheet(ws, items, conn, previously_known, baseline_ids) -> None:
    sorted_items = sorted(items, key=_sort_key)

    city_groups = defaultdict(lambda: defaultdict(list))
    for item in sorted_items:
        developer = getattr(item, 'developer', '') or ''
        city_groups[item.city][(developer, item.site, item.complex_name)].append(item)

    count_key = lambda it: (it.site, it.complex_name, it.building)
    counts = Counter(count_key(it) for it in items)

    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 1

    for city in sorted(city_groups.keys()):
        # ГОРОД
        ws.cell(row=row, column=1, value=city.upper()).font = CITY_FONT
        ws.cell(row=row, column=1).fill = CITY_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
        for c in range(1, PCOL + 1):
            ws.cell(row=row, column=c).fill = CITY_FILL
        ws.row_dimensions[row].height = 30
        row += 1

        for (developer, site, complex_name) in sorted(city_groups[city].keys()):
            complex_items = city_groups[city][(developer, site, complex_name)]
            display_name = developer or SITE_NAMES.get(site, site)
            total_in_complex = len(complex_items)

            # ЖК
            jk_text = f"{display_name}  —  {complex_name}  ({total_in_complex} кладовок)"
            jk_cell = ws.cell(row=row, column=1, value=jk_text)
            jk_cell.font = COMPLEX_FONT
            jk_cell.fill = COMPLEX_FILL
            jk_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
            for c in range(1, PCOL + 1):
                ws.cell(row=row, column=c).fill = COMPLEX_FILL
            ws.row_dimensions[row].height = 25
            jk_row = row
            row += 1

            # Шапка
            for col_idx, col_name in enumerate(PRETTY_COLUMNS, start=1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            row += 1

            block_start = row

            # Данные
            for i, item in enumerate(complex_items):
                is_last = i == len(complex_items) - 1
                cur_bld = complex_items[i].building.split("||")[0].strip()
                next_bld = complex_items[i + 1].building.split("||")[0].strip() if not is_last else ""
                is_last_in_building = is_last or cur_bld != next_bld

                storehouse_count = counts[count_key(item)]
                try:
                    number_val = int(item.item_number) if item.item_number else ""
                except ValueError:
                    number_val = item.item_number or ""

                # Разделяем building||секция (если есть)
                building_display = item.building
                building_note = None
                if "||" in item.building:
                    parts = item.building.split("||", 1)
                    building_display = parts[0].strip()
                    building_note = parts[1].strip()

                # Пустые цены (дом.рф) — пишем пустую строку вместо 0
                display_price = item.price if item.price else ""
                display_ppm = item.price_per_meter if item.price_per_meter else ""

                row_data = [
                    building_display,
                    storehouse_count,
                    number_val,
                    item.area,
                    display_price,
                    display_ppm,
                    "Открыть",
                    row - block_start + 1,
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = DATA_FONT
                    cell.alignment = DATA_ALIGN
                    cell.border = THIN_BORDER

                # Примечание к корпусу (секция)
                if building_note:
                    _append_comment(ws.cell(row=row, column=1), building_note, "Parser")

                # Форматы
                ws.cell(row=row, column=4).number_format = '0.00' if item.site == 'domrf' else '0.0'
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=6).number_format = '#,##0'
                if isinstance(number_val, int):
                    ws.cell(row=row, column=3).number_format = '0'
                ws.cell(row=row, column=2).number_format = '0'
                ws.cell(row=row, column=8).number_format = '0'
                ws.cell(row=row, column=8).font = Font(name="Calibri", size=9, color="BBBBBB")

                # Ссылка
                lc = ws.cell(row=row, column=7)
                if item.url:
                    lc.hyperlink = item.url
                    lc.font = LINK_FONT
                lc.alignment = DATA_ALIGN

                # Разделитель корпусов — на ВСЕХ столбцах строки
                if is_last_in_building and not is_last:
                    for c in range(1, PCOL + 1):
                        ws.cell(row=row, column=c).border = BUILDING_BOTTOM

                # Секция в примечании
                section_name = getattr(item, '_section_name', None)
                if section_name:
                    bc = ws.cell(row=row, column=1)
                    _append_comment(bc, section_name, "Парсер кладовок")
                    bc.comment.width = 150
                    bc.comment.height = 30

                # Новая кладовка — примечание на номере
                _add_new_item_comment(ws, row, 3, item, previously_known, conn, baseline_ids, total_cols=PCOL)

                # Примечания цен
                _add_price_comment(ws, row, 5, conn, item)
                _add_ppm_comment(ws, row, 6, conn, item)

                row += 1

            block_end = row - 1
            header_row = block_start - 1  # строка с шапкой колонок

            # Excel Table — сортировка и фильтр для каждого блока ЖК
            if block_end >= block_start:
                # Уникальное имя таблицы (без пробелов и спецсимволов)
                import re as _re
                safe_name = _re.sub(r'[^A-Za-z0-9а-яА-ЯёЁ]', '', complex_name)
                table_name = f"T_{site}_{safe_name}_{jk_row}"
                # Имена таблиц не могут содержать кириллицу в некоторых Excel
                table_name = _re.sub(r'[^A-Za-z0-9_]', '', table_name)
                if not table_name[0].isalpha():
                    table_name = "T" + table_name

                table_ref = (
                    f"A{header_row}:{get_column_letter(PCOL)}{block_end}"
                )
                tab = Table(displayName=table_name, ref=table_ref)
                tab.tableStyleInfo = TableStyleInfo(
                    name="TableStyleLight9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(tab)

            # Группировка ЖК (outline) — свёрнутые по умолчанию
            group_start = jk_row + 1
            group_end = block_end
            if group_end >= group_start:
                ws.row_dimensions.group(group_start, group_end, outline_level=1, hidden=True)

    # ── Итоги по ЖК ──
    row += 1
    ws.cell(row=row, column=1, value="ИТОГИ ПО ЖК").font = Font(
        name="Calibri", size=12, bold=True
    )
    row += 1
    for col_idx, col_name in enumerate(["Город", "Застройщик", "ЖК", "Всего"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    complex_counts = defaultdict(int)
    developer_map: dict[tuple, str] = {}
    for item in items:
        complex_counts[(item.city, item.site, item.complex_name)] += 1
        if item.developer:
            developer_map[(item.site, item.complex_name)] = item.developer

    for (city, site, cname), total in sorted(complex_counts.items()):
        display_name = developer_map.get((site, cname)) or SITE_NAMES.get(site, site)
        for col, val in [(1, city), (2, display_name), (3, cname), (4, total)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = TOTAL_FONT
            cell.fill = COMPLEX_TOTAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    total_all = sum(complex_counts.values())
    ws.cell(row=row, column=3, value="ВСЕГО").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    t = ws.cell(row=row, column=4, value=total_all)
    t.font = Font(name="Calibri", size=12, bold=True)
    t.fill = GRAND_TOTAL_FILL
    t.border = THIN_BORDER
    t.alignment = Alignment(horizontal="center")

    for i, w in enumerate([16, 14, 14, 12, 14, 14, 12, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════
#  ЛИСТ 2: «Все данные» — плоская таблица
# ══════════════════════════════════════════════════════

def _fill_flat_sheet(ws, items, conn, previously_known, baseline_ids) -> None:
    sorted_items = sorted(items, key=_sort_key)

    count_key = lambda it: (it.site, it.complex_name, it.building)
    counts = Counter(count_key(it) for it in items)

    # Заголовки
    for col_idx, col_name in enumerate(FLAT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    # Данные — считаем диапазоны ЖК для группировки
    prev_complex = None
    prev_building = None
    complex_ranges: list[tuple[int, int]] = []  # (start_row, end_row) per ЖК
    cur_complex_start = 2

    for i, item in enumerate(sorted_items):
        row = i + 2
        display_name = item.developer or SITE_NAMES.get(item.site, item.site)
        storehouse_count = counts[count_key(item)]

        try:
            number_val = int(item.item_number) if item.item_number else ""
        except ValueError:
            number_val = item.item_number or ""

        cur_complex = (item.city, item.site, item.complex_name)
        cur_building = (item.city, item.site, item.complex_name, item.building.split("||")[0].strip())

        # Фиксируем конец предыдущего ЖК
        if cur_complex != prev_complex and prev_complex is not None:
            complex_ranges.append((cur_complex_start, row - 1))
            cur_complex_start = row

        # Разделитель корпусов — жирная граница только на столбце «Корпус» (col 4)
        if cur_building != prev_building and prev_building is not None:
            if cur_complex == prev_complex:  # не между ЖК, а внутри
                prev_row = row - 1
                ws.cell(row=prev_row, column=4).border = BUILDING_BOTTOM

        prev_complex = cur_complex
        prev_building = cur_building

        # Разделяем building||секция (если есть)
        building_display = item.building
        building_note = None
        if "||" in item.building:
            parts = item.building.split("||", 1)
            building_display = parts[0].strip()
            building_note = parts[1].strip()

        # Пустые цены (дом.рф) — пишем пустую строку вместо 0
        display_price = item.price if item.price else ""
        display_ppm = item.price_per_meter if item.price_per_meter else ""

        row_data = [
            item.city,
            display_name,
            item.complex_name,
            building_display,
            storehouse_count,
            number_val,
            item.area,
            display_price,
            display_ppm,
            "Открыть",
            i + 1,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

        # Примечание к корпусу (секция)
        if building_note:
            _append_comment(ws.cell(row=row, column=4), building_note, "Parser")

        # Форматы
        ws.cell(row=row, column=7).number_format = '0.00' if item.site == 'domrf' else '0.0'
        ws.cell(row=row, column=8).number_format = '#,##0'
        ws.cell(row=row, column=9).number_format = '#,##0'
        if isinstance(number_val, int):
            ws.cell(row=row, column=6).number_format = '0'
        ws.cell(row=row, column=5).number_format = '0'
        ws.cell(row=row, column=11).number_format = '0'
        ws.cell(row=row, column=11).font = Font(name="Calibri", size=9, color="BBBBBB")

        # Ссылка
        lc = ws.cell(row=row, column=10)
        if item.url:
            lc.hyperlink = item.url
            lc.font = LINK_FONT
        lc.alignment = DATA_ALIGN

        # Секция
        section_name = getattr(item, '_section_name', None)
        if section_name:
            bc = ws.cell(row=row, column=4)
            _append_comment(bc, section_name, "Парсер кладовок")
            bc.comment.width = 150
            bc.comment.height = 30

        # Новая кладовка
        _add_new_item_comment(ws, row, 6, item, previously_known, conn, baseline_ids, total_cols=FCOL)

        # Примечания цен
        _add_price_comment(ws, row, 8, conn, item)
        _add_ppm_comment(ws, row, 9, conn, item)

    # Последний ЖК
    last_row = len(sorted_items) + 1
    if sorted_items:
        complex_ranges.append((cur_complex_start, last_row))

    # Жирная граница между ЖК
    for idx, (cs, ce) in enumerate(complex_ranges):
        if idx < len(complex_ranges) - 1:
            # Толстая нижняя граница на последней строке ЖК (на всех столбцах)
            thick_bottom = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="medium", color="4472C4"),
            )
            for c in range(1, FCOL + 1):
                ws.cell(row=ce, column=c).border = thick_bottom

    # Автофильтр (для скрытия ЖК — фильтр по столбцу «ЖК»)
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(FCOL)}{last_row}"

    # Ширина
    flat_widths = [14, 16, 18, 16, 14, 14, 12, 14, 14, 12, 8]
    for i, w in enumerate(flat_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Примечания ───────────────────────────────────────────

def _add_new_item_comment(ws, row, col, item, previously_known, conn,
                          baseline_ids, total_cols=None):
    """Примечание «Добавлена от [дата]» — делегирует в exporter_common."""
    _add_new_item_comment_common(
        ws, row, col, item, previously_known, conn, baseline_ids,
        get_first_seen_fn=get_first_seen_date,
        author="Парсер кладовок",
        total_cols=total_cols,
    )


def _add_price_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена» — делегирует в exporter_common."""
    _add_price_comment_common(
        ws, row, col, conn, item,
        get_price_history_fn=get_price_history,
        author="Парсер кладовок",
    )


def _add_ppm_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена/м²» — делегирует в exporter_common."""
    _add_ppm_comment_common(
        ws, row, col, conn, item,
        get_price_history_fn=get_price_history,
        author="Парсер кладовок",
    )
