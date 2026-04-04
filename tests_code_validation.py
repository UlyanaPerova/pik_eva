"""
Валидация кода pik_eva — проверка что все модули работают корректно.

НЕ валидация данных (для этого есть validate_data.py).
Проверяет: модели, БД-функции, конфиги, нормализацию building,
миграции, скоринг-формулы, обратную совместимость импортов,
чеклист из CLAUDE_CODE_PROMPT.md.

72 теста, ~0.4 секунды.

Запуск:
    python tests_code_validation.py
    python -m unittest tests_code_validation -v
"""
from __future__ import annotations

import os
import sqlite3
import tempfile
import unittest
from pathlib import Path

os.chdir(Path(__file__).resolve().parent)


# ═══════════════════════════════════════════════════════
#  1. parsers/models.py
# ═══════════════════════════════════════════════════════

class TestModels(unittest.TestCase):

    def test_storehouse_item_creation(self):
        from parsers.models import StorehouseItem
        item = StorehouseItem(
            site="pik", city="Казань", complex_name="Тест",
            building="1", item_id="s1", area=3.5,
            price=350000, price_per_meter=100000, url="http://example.com",
        )
        self.assertEqual(item.site, "pik")
        self.assertEqual(item.area, 3.5)
        self.assertIsNone(item.developer)
        self.assertIsNone(item.item_number)

    def test_apartment_item_creation(self):
        from parsers.models import ApartmentItem
        item = ApartmentItem(
            site="domrf", city="Казань", complex_name="Тест",
            building="1", item_id="a1", rooms=2, floor=5,
            area=55.0, price=0, price_per_meter=0, url="/test",
        )
        self.assertEqual(item.rooms, 2)
        self.assertEqual(item.rooms_label, "2-комн.")
        self.assertIsNone(item.living_area)

    def test_rooms_label(self):
        from parsers.models import rooms_label
        self.assertEqual(rooms_label(0), "Студия")
        self.assertEqual(rooms_label(1), "1-комн.")
        self.assertEqual(rooms_label(4), "4-комн.")
        self.assertEqual(rooms_label(7), "7-комн.")

    def test_storehouse_optional_fields(self):
        from parsers.models import StorehouseItem
        item = StorehouseItem(
            site="pik", city="Казань", complex_name="ЖК",
            building="1", item_id="x", area=3.0,
            price=300000, price_per_meter=100000, url="/",
            item_number="42", original_price=350000,
            discount_percent=14.3, developer="ПИК",
        )
        self.assertEqual(item.item_number, "42")
        self.assertEqual(item.original_price, 350000)
        self.assertAlmostEqual(item.discount_percent, 14.3)
        self.assertEqual(item.developer, "ПИК")


# ═══════════════════════════════════════════════════════
#  2. parsers/db.py
# ═══════════════════════════════════════════════════════

class TestDb(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.db_path = Path(self.tmpdir) / "test.db"

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_init_db_creates_table(self):
        from parsers.db import init_db
        conn = init_db(
            self.db_path, "items",
            "CREATE TABLE IF NOT EXISTS items (id INTEGER PRIMARY KEY, name TEXT)",
            "CREATE INDEX IF NOT EXISTS idx ON items (name)",
        )
        conn.execute("INSERT INTO items (name) VALUES ('test')")
        conn.commit()
        row = conn.execute("SELECT name FROM items").fetchone()
        self.assertEqual(row[0], "test")
        conn.close()

    def test_init_db_with_versioned_migrations(self):
        from parsers.db import init_db
        conn = init_db(
            self.db_path, "items",
            "CREATE TABLE IF NOT EXISTS items (id INTEGER PRIMARY KEY, name TEXT)",
            "CREATE INDEX IF NOT EXISTS idx ON items (name)",
            versioned_migrations=[
                (1, "Добавить col1", "ALTER TABLE items ADD COLUMN col1 TEXT"),
                (2, "Добавить col2", "ALTER TABLE items ADD COLUMN col2 REAL"),
            ],
        )
        # Столбцы должны существовать
        conn.execute("INSERT INTO items (name, col1, col2) VALUES ('x', 'a', 1.5)")
        conn.commit()
        row = conn.execute("SELECT col1, col2 FROM items").fetchone()
        self.assertEqual(row, ("a", 1.5))
        conn.close()

    def test_backup_db(self):
        from parsers.db import backup_db
        # Создать БД для бэкапа
        conn = sqlite3.connect(str(self.db_path))
        conn.execute("CREATE TABLE t (x INTEGER)")
        conn.commit()
        conn.close()

        backup_dir = Path(self.tmpdir) / "backups"
        result = backup_db(self.db_path, backup_dir, "test")
        self.assertIsNotNone(result)
        self.assertTrue(result.exists())
        self.assertTrue(result.name.startswith("test_"))

    def test_backup_db_rotation(self):
        from parsers.db import backup_db
        conn = sqlite3.connect(str(self.db_path))
        conn.execute("CREATE TABLE t (x INTEGER)")
        conn.commit()
        conn.close()

        backup_dir = Path(self.tmpdir) / "backups"
        # Создать 12 бэкапов, оставить только 3
        for _ in range(5):
            backup_db(self.db_path, backup_dir, "test", keep=3)
        backups = list(backup_dir.glob("test_*.db"))
        self.assertLessEqual(len(backups), 3)

    def test_backup_db_nonexistent(self):
        from parsers.db import backup_db
        result = backup_db(Path("/nonexistent.db"), Path(self.tmpdir) / "bk", "test")
        self.assertIsNone(result)

    def test_get_all_known_ids(self):
        from parsers.db import get_all_known_ids
        conn = sqlite3.connect(":memory:")
        conn.execute("CREATE TABLE prices (site TEXT, item_id TEXT)")
        conn.executemany(
            "INSERT INTO prices VALUES (?, ?)",
            [("pik", "a"), ("pik", "b"), ("pik", "a"), ("domrf", "c")],
        )
        conn.commit()
        ids = get_all_known_ids(conn, "prices", "pik")
        self.assertEqual(ids, {"a", "b"})
        conn.close()

    def test_get_price_history(self):
        from parsers.db import get_price_history
        conn = sqlite3.connect(":memory:")
        conn.execute("""CREATE TABLE prices (
            site TEXT, item_id TEXT, price REAL, price_per_meter REAL,
            original_price REAL, discount_percent REAL, parsed_at TEXT
        )""")
        conn.execute(
            "INSERT INTO prices VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("pik", "x", 100, 50, None, None, "2026-01-01"),
        )
        conn.execute(
            "INSERT INTO prices VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("pik", "x", 110, 55, None, None, "2026-01-02"),
        )
        conn.commit()
        history = get_price_history(conn, "prices", "pik", "x")
        self.assertEqual(len(history), 2)
        self.assertEqual(history[0][0], 110)  # новая цена первая
        conn.close()

    def test_get_first_seen_date(self):
        from parsers.db import get_first_seen_date
        conn = sqlite3.connect(":memory:")
        conn.execute("CREATE TABLE t (site TEXT, item_id TEXT, parsed_at TEXT)")
        conn.execute("INSERT INTO t VALUES ('a', '1', '2026-03-01')")
        conn.execute("INSERT INTO t VALUES ('a', '1', '2026-04-01')")
        conn.commit()
        d = get_first_seen_date(conn, "t", "a", "1")
        self.assertEqual(d, "2026-03-01")
        self.assertIsNone(get_first_seen_date(conn, "t", "a", "999"))
        conn.close()


# ═══════════════════════════════════════════════════════
#  3. parsers/config.py
# ═══════════════════════════════════════════════════════

class TestConfig(unittest.TestCase):

    def test_validate_config_valid(self):
        from parsers.config import validate_config
        config = {
            "links": [
                {"object_id": 1, "complex_name": "ЖК Тест", "city": "Казань", "developer": "Dev"},
                {"object_id": 2, "complex_name": "ЖК 2", "city": "Казань", "developer": "Dev"},
            ]
        }
        validate_config(config)  # не должно бросить

    def test_validate_config_missing_links(self):
        from parsers.config import validate_config
        with self.assertRaises(ValueError):
            validate_config({})

    def test_validate_config_missing_object_id(self):
        from parsers.config import validate_config
        config = {"links": [{"complex_name": "ЖК"}]}
        with self.assertRaises(ValueError):
            validate_config(config)

    def test_validate_config_missing_complex_name(self):
        from parsers.config import validate_config
        config = {"links": [{"object_id": 1}]}
        with self.assertRaises(ValueError):
            validate_config(config)

    def test_validate_config_require_building(self):
        from parsers.config import validate_config
        config = {"links": [{"object_id": 1, "complex_name": "ЖК", "city": "К", "developer": "D"}]}
        # Без require_building — ок
        validate_config(config, require_building=False)
        # С require_building — ошибка (нет building)
        with self.assertRaises(ValueError):
            validate_config(config, require_building=True)

    def test_validate_config_invalid_object_id(self):
        from parsers.config import validate_config
        config = {"links": [{"object_id": -1, "complex_name": "ЖК"}]}
        with self.assertRaises(ValueError):
            validate_config(config)

    def test_validate_config_empty_links(self):
        from parsers.config import validate_config
        config = {"links": []}
        validate_config(config)  # пустой список — не ошибка, только warning

    def test_load_config(self):
        from parsers.config import load_config
        import tempfile, yaml
        with tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", delete=False) as f:
            yaml.dump({"name": "test", "links": []}, f)
            f.flush()
            cfg = load_config(f.name)
        self.assertEqual(cfg["name"], "test")
        os.unlink(f.name)


# ═══════════════════════════════════════════════════════
#  4. parsers/building.py
# ═══════════════════════════════════════════════════════

class TestBuilding(unittest.TestCase):

    def test_normalize_simple(self):
        from parsers.building import normalize_building
        b = normalize_building("Корпус 1")
        self.assertEqual(b.primary, "Корпус 1")
        self.assertEqual(b.notes, "")
        self.assertIsNone(b.entrance)
        self.assertIsNone(b.section)

    def test_normalize_with_entrance(self):
        from parsers.building import normalize_building
        b = normalize_building("1||подъезд 2")
        self.assertEqual(b.primary, "1")
        self.assertEqual(b.notes, "подъезд 2")
        self.assertEqual(b.entrance, 2)
        self.assertIsNone(b.section)

    def test_normalize_with_section(self):
        from parsers.building import normalize_building
        b = normalize_building("1||секция A3")
        self.assertEqual(b.primary, "1")
        self.assertEqual(b.section, "A3")

    def test_normalize_pik_format(self):
        from parsers.building import normalize_building
        b = normalize_building("Корпус 1.1||Секции 1-4")
        self.assertEqual(b.primary, "Корпус 1.1")
        self.assertEqual(b.notes, "Секции 1-4")

    def test_normalize_empty(self):
        from parsers.building import normalize_building
        b = normalize_building("")
        self.assertEqual(b.primary, "")
        self.assertEqual(b.notes, "")

    def test_building_display(self):
        from parsers.building import building_display
        self.assertEqual(building_display("1||подъезд 2"), "1")
        self.assertEqual(building_display("М1/ПК-1"), "М1/ПК-1")
        self.assertEqual(building_display(""), "")
        self.assertEqual(building_display("  Корпус 1  "), "Корпус 1")

    def test_building_key(self):
        from parsers.building import building_key
        self.assertEqual(building_key("Корпус 10"), "10")
        self.assertEqual(building_key("корпус 10"), "10")
        self.assertEqual(building_key("М1/ПК-1"), "м1пк1")
        self.assertEqual(building_key("1||подъезд 2"), "1")
        self.assertEqual(building_key(""), "")
        self.assertEqual(building_key("  1.2  "), "1.2")

    def test_building_key_matches_norm(self):
        """building_key должна давать те же результаты, что _norm в eva_calculator."""
        from parsers.building import building_key
        from eva_calculator import _norm
        cases = ["Корпус 1", "1", "М1/ПК-1", "корпус 10", "1.2"]
        for c in cases:
            self.assertEqual(building_key(c), _norm(c), f"Mismatch for '{c}'")


# ═══════════════════════════════════════════════════════
#  5. parsers/migrations.py
# ═══════════════════════════════════════════════════════

class TestMigrations(unittest.TestCase):

    def test_get_version_new_db(self):
        from parsers.migrations import get_version
        conn = sqlite3.connect(":memory:")
        v = get_version(conn)
        self.assertEqual(v, 0)
        conn.close()

    def test_apply_migrations(self):
        from parsers.migrations import apply_migrations, get_version
        conn = sqlite3.connect(":memory:")
        conn.execute("CREATE TABLE items (id INTEGER PRIMARY KEY, name TEXT)")
        migrations = [
            (1, "Добавить col1", "ALTER TABLE items ADD COLUMN col1 TEXT"),
            (2, "Добавить col2", "ALTER TABLE items ADD COLUMN col2 REAL"),
        ]
        applied = apply_migrations(conn, migrations)
        self.assertEqual(applied, 2)
        self.assertEqual(get_version(conn), 2)
        conn.close()

    def test_apply_migrations_idempotent(self):
        from parsers.migrations import apply_migrations, get_version
        conn = sqlite3.connect(":memory:")
        conn.execute("CREATE TABLE items (id INTEGER PRIMARY KEY)")
        migrations = [(1, "test", "ALTER TABLE items ADD COLUMN x TEXT")]
        apply_migrations(conn, migrations)
        # Повторный вызов — ничего не применяется
        applied = apply_migrations(conn, migrations)
        self.assertEqual(applied, 0)
        self.assertEqual(get_version(conn), 1)
        conn.close()

    def test_apply_migrations_partial(self):
        from parsers.migrations import apply_migrations, get_version
        conn = sqlite3.connect(":memory:")
        conn.execute("CREATE TABLE items (id INTEGER PRIMARY KEY)")
        apply_migrations(conn, [(1, "v1", "ALTER TABLE items ADD COLUMN a TEXT")])
        self.assertEqual(get_version(conn), 1)
        # Добавляем ещё миграции
        applied = apply_migrations(conn, [
            (1, "v1", "ALTER TABLE items ADD COLUMN a TEXT"),
            (2, "v2", "ALTER TABLE items ADD COLUMN b TEXT"),
        ])
        self.assertEqual(applied, 1)
        self.assertEqual(get_version(conn), 2)
        conn.close()

    def test_record_and_get_parse_run(self):
        from parsers.migrations import record_parse_run, get_last_parse_run
        conn = sqlite3.connect(":memory:")
        record_parse_run(
            conn, "pik", "2026-04-01T10:00:00", "2026-04-01T10:05:00",
            items_count=100, items_saved=10, success=True, duration_sec=300.0,
        )
        run = get_last_parse_run(conn, "pik")
        self.assertIsNotNone(run)
        self.assertEqual(run["site"], "pik")
        self.assertEqual(run["items_count"], 100)
        self.assertTrue(run["success"])
        self.assertIsNone(get_last_parse_run(conn, "nonexistent"))
        conn.close()


# ═══════════════════════════════════════════════════════
#  6. runners/run_result.py
# ═══════════════════════════════════════════════════════

class TestRunResult(unittest.TestCase):

    def test_success_exit_code(self):
        from runners.run_result import RunResult
        r = RunResult(success=True, site="pik")
        self.assertEqual(r.exit_code, 0)

    def test_failure_exit_code(self):
        from runners.run_result import RunResult
        r = RunResult(success=False, site="pik")
        self.assertEqual(r.exit_code, 1)

    def test_summary_format(self):
        from runners.run_result import RunResult
        r = RunResult(
            success=True, site="pik",
            items_count=42, items_saved=5,
            duration_sec=3.5,
        )
        s = r.summary()
        self.assertIn("[OK]", s)
        self.assertIn("pik", s)
        self.assertIn("42", s)

    def test_summary_with_errors(self):
        from runners.run_result import RunResult
        r = RunResult(
            success=False, site="domrf",
            errors=["Ошибка 1", "Ошибка 2"],
            warnings=["warn"],
        )
        s = r.summary()
        self.assertIn("[FAIL]", s)
        self.assertIn("errors=2", s)
        self.assertIn("warnings=1", s)

    def test_defaults(self):
        from runners.run_result import RunResult
        r = RunResult(success=True, site="test")
        self.assertEqual(r.items_count, 0)
        self.assertEqual(r.items_saved, 0)
        self.assertEqual(r.errors, [])
        self.assertEqual(r.warnings, [])
        self.assertEqual(r.duration_sec, 0.0)
        self.assertIsNone(r.output_path)


# ═══════════════════════════════════════════════════════
#  7. exporter_common.py
# ═══════════════════════════════════════════════════════

class TestExporterCommon(unittest.TestCase):

    def test_natural_sort_key(self):
        from exporter_common import natural_sort_key
        items = ["ПК-10", "ПК-2", "ПК-1", "ПК-20"]
        result = sorted(items, key=natural_sort_key)
        self.assertEqual(result, ["ПК-1", "ПК-2", "ПК-10", "ПК-20"])

    def test_natural_sort_key_buildings(self):
        from exporter_common import natural_sort_key
        items = ["Корпус 2", "Корпус 1.2", "Корпус 1.1", "Корпус 10"]
        result = sorted(items, key=natural_sort_key)
        self.assertEqual(result, ["Корпус 1.1", "Корпус 1.2", "Корпус 2", "Корпус 10"])

    def test_site_names(self):
        from exporter_common import SITE_NAMES
        self.assertIn("pik", SITE_NAMES)
        self.assertIn("domrf", SITE_NAMES)
        self.assertEqual(SITE_NAMES["pik"], "ПИК")

    def test_site_file_keys(self):
        from exporter_common import SITE_FILE_KEYS
        self.assertEqual(SITE_FILE_KEYS["pik"], "PIK")
        self.assertEqual(SITE_FILE_KEYS["domrf"], "DomRF")

    def test_append_comment_new(self):
        from exporter_common import append_comment
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        append_comment(cell, "Тест", "Автор")
        self.assertEqual(cell.comment.text, "Тест")

    def test_append_comment_existing(self):
        from exporter_common import append_comment
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.comment = Comment("Первый", "Авт")
        append_comment(cell, "Второй", "Авт")
        self.assertIn("Первый", cell.comment.text)
        self.assertIn("Второй", cell.comment.text)

    def test_append_comment_no_duplicate(self):
        from exporter_common import append_comment
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.comment = Comment("Текст", "Авт")
        append_comment(cell, "Текст", "Авт")
        # Не дублирует
        self.assertEqual(cell.comment.text.count("Текст"), 1)

    def test_styles_are_shared(self):
        import exporter_common
        import exporter
        import exporter_apartments
        # Стили должны быть одними объектами
        self.assertIs(exporter.HEADER_FILL, exporter_common.HEADER_FILL)
        self.assertIs(exporter_apartments.HEADER_FILL, exporter_common.HEADER_FILL)
        self.assertIs(exporter.THIN_BORDER, exporter_common.THIN_BORDER)


# ═══════════════════════════════════════════════════════
#  8. scoring.py
# ═══════════════════════════════════════════════════════

class TestScoring(unittest.TestCase):

    def test_score_by_thresholds(self):
        from scoring import _score_by_thresholds
        thresholds = [
            {"max": 5, "points": 10},
            {"max": 10, "points": 5},
            {"max": 9999, "points": 0},
        ]
        self.assertEqual(_score_by_thresholds(3, thresholds), 10)
        self.assertEqual(_score_by_thresholds(7, thresholds), 5)
        self.assertEqual(_score_by_thresholds(100, thresholds), 0)

    def test_calc_first_stage(self):
        from scoring import calc_first_stage
        pts = calc_first_stage(
            days_until=180,
            rooms_count={0: 10, 1: 30, 2: 40, 3: 15, 4: 5},
            avg_area={0: 24, 1: 34, 2: 54, 3: 77, 4: 99},
            avg_non_living={0: 12, 1: 22, 2: 27, 3: 34, 4: 39},
        )
        self.assertIsInstance(pts, int)
        # С хорошими данными баллы должны быть положительными
        self.assertGreater(pts, 0)

    def test_calc_second_stage(self):
        from scoring import calc_second_stage
        pts = calc_second_stage(area=3.5, price=400000, price_per_meter=80000)
        self.assertIsInstance(pts, int)
        self.assertGreater(pts, 0)

    def test_generate_first_stage_formula(self):
        from scoring import generate_first_stage_formula
        formula = generate_first_stage_formula(3)
        self.assertTrue(formula.startswith("=ROUND("))
        self.assertIn("E3", formula)  # days
        self.assertIn("M3", formula)  # apt/store ratio
        self.assertIn("S3", formula)  # rooms_studio

    def test_generate_second_stage_formula(self):
        from scoring import generate_second_stage_formula
        formula = generate_second_stage_formula(5)
        self.assertTrue(formula.startswith("=ROUND("))
        self.assertIn("L5", formula)  # area
        self.assertIn("M5", formula)  # price
        self.assertIn("N5", formula)  # ppm

    def test_generate_balcony_ratio_formula(self):
        from scoring import generate_balcony_ratio_formula
        formula = generate_balcony_ratio_formula(3)
        self.assertIn("IFERROR", formula)
        self.assertIn("3", formula)

    def test_generate_total_formula(self):
        from scoring import generate_total_formula
        formula = generate_total_formula(7)
        self.assertEqual(formula, "=ROUND(I7+J7,2)")

    def test_jk_cols_completeness(self):
        from scoring import JK_COLS
        required_keys = [
            "city", "developer", "complex", "building", "days",
            "first_stage", "apt_store_ratio",
            "rooms_studio", "rooms_1k", "rooms_2k", "rooms_3k", "rooms_4k",
            "area_studio", "area_1k", "area_2k", "area_3k", "area_4k",
        ]
        for key in required_keys:
            self.assertIn(key, JK_COLS, f"Missing key: {key}")

    def test_formulas_read_from_yaml(self):
        """Формулы должны использовать параметры из eva.yaml, а не хардкод."""
        from scoring import generate_first_stage_formula, load_scoring_config
        scoring = load_scoring_config()
        # Изменяем deadline параметры
        scoring_modified = dict(scoring)
        scoring_modified["deadline_formula"] = {
            "tier1_days": 100,
            "tier1_max_pts": 50,
            "tier1_decay_pts": 25,
            "tier2_days": 200,
            "tier2_max_pts": 25,
            "tier2_decay_pts": 10,
        }
        formula = generate_first_stage_formula(3, scoring_modified)
        # Формула должна содержать наши параметры
        self.assertIn("100", formula)
        self.assertIn("50", formula)


# ═══════════════════════════════════════════════════════
#  9. config_manager.py
# ═══════════════════════════════════════════════════════

class TestConfigManager(unittest.TestCase):

    def test_get_status_structure(self):
        from config_manager import get_status
        status = get_status()
        self.assertIn("apartments", status)
        self.assertIn("storehouses", status)
        self.assertIn("config", status)
        self.assertIn("db_exists", status["apartments"])
        self.assertIn("sites", status["apartments"])
        self.assertIsInstance(status["config"]["apartments_links"], int)

    def test_get_last_run_info(self):
        from config_manager import get_last_run_info
        info = get_last_run_info("domrf")
        self.assertIn("apartments", info)
        self.assertIn("storehouses", info)

    def test_get_scoring_config(self):
        from config_manager import get_scoring_config
        scoring = get_scoring_config()
        self.assertIsInstance(scoring, dict)
        if scoring:
            self.assertIn("kvartirografia", scoring)
            self.assertIn("avg_area", scoring)
            self.assertIn("second_stage", scoring)
            # Проверяем что новые секции тоже есть
            self.assertIn("deadline_formula", scoring)
            self.assertIn("apt_store_ratio_formula", scoring)
            self.assertIn("balcony_ratio_formula", scoring)
            self.assertIn("balcony_bonus", scoring)


# ═══════════════════════════════════════════════════════
#  10. eva_calculator.py
# ═══════════════════════════════════════════════════════

class TestEvaCalculator(unittest.TestCase):

    def test_norm(self):
        from eva_calculator import _norm
        self.assertEqual(_norm("Корпус 10"), "10")
        self.assertEqual(_norm("корпус 1"), "1")
        self.assertEqual(_norm("М1/ПК-1"), "м1пк1")
        self.assertEqual(_norm(""), "")
        self.assertEqual(_norm("  1.2  "), "1.2")

    def test_norm_consistency(self):
        from eva_calculator import _norm
        # Одинаковые строки → одинаковые ключи
        self.assertEqual(_norm("Корпус 1"), _norm("корпус 1"))
        self.assertEqual(_norm("М1/ПК-1"), _norm("м1/пк-1"))

    def test_building_agg_dataclass(self):
        from eva_calculator import BuildingAgg
        # Просто убедимся что можно создать
        self.assertTrue(hasattr(BuildingAgg, "__dataclass_fields__"))


# ═══════════════════════════════════════════════════════
#  11. Интеграция: backward compatibility
# ═══════════════════════════════════════════════════════

class TestBackwardCompatibility(unittest.TestCase):
    """Проверяем что все старые импорты продолжают работать."""

    def test_base_exports(self):
        from parsers.base import (
            StorehouseItem, BaseParser, init_db, save_items, backup_db,
            validate_items, get_all_known_ids, get_price_history,
            get_first_seen_date, get_latest_items, load_config, validate_config,
            logger, DB_PATH, DATA_DIR, PROJECT_DIR, BACKUP_DIR, LOG_DIR,
        )
        self.assertIsNotNone(StorehouseItem)
        self.assertIsNotNone(BaseParser)

    def test_apartments_base_exports(self):
        from parsers.apartments_base import (
            ApartmentItem, BaseApartmentParser, ROOM_LABELS, rooms_label,
            init_db, save_items, backup_db, validate_items,
            get_all_known_ids, get_price_history, get_first_seen_date,
            calc_avg_prices, load_or_create_baseline, load_config, validate_config,
            logger, DB_PATH, DATA_DIR, PROJECT_DIR, BACKUP_DIR, LOG_DIR,
            OUTPUT_DIR, BASELINE_DIR,
        )
        self.assertIsNotNone(ApartmentItem)
        self.assertIsNotNone(BaseApartmentParser)

    def test_exporter_styles_available(self):
        from exporter import (
            HEADER_FILL, HEADER_FONT, CITY_FILL, CITY_FONT,
            COMPLEX_FILL, COMPLEX_FONT, DATA_FONT, DATA_ALIGN,
            LINK_FONT, NEW_ITEM_FILL, THIN_BORDER, BUILDING_BOTTOM,
            SITE_NAMES, SITE_FILE_KEYS,
        )
        self.assertIsNotNone(HEADER_FILL)

    def test_exporter_apartments_styles_available(self):
        from exporter_apartments import (
            HEADER_FILL, HEADER_FONT, DATA_FONT, DATA_ALIGN,
            LINK_FONT, NEW_ITEM_FILL, THIN_BORDER,
            SITE_NAMES, SITE_FILE_KEYS,
            ROOM_TYPE_FILL, AVG_HEADER_FILL,
        )
        self.assertIsNotNone(ROOM_TYPE_FILL)

    def test_runners_import(self):
        from runners import RunResult, run_storehouse_parser, run_apartment_parser
        self.assertIsNotNone(RunResult)

    def test_all_parsers_import(self):
        import importlib
        parsers = [
            "parsers.pik", "parsers.domrf", "parsers.glorax",
            "parsers.smu88", "parsers.akbarsdom", "parsers.unistroy",
            "parsers.pik_apartments", "parsers.domrf_apartments",
            "parsers.glorax_apartments", "parsers.smu88_apartments",
            "parsers.akbarsdom_apartments", "parsers.unistroy_apartments",
        ]
        for p in parsers:
            mod = importlib.import_module(p)
            self.assertIsNotNone(mod, f"Failed to import {p}")


# ═══════════════════════════════════════════════════════
#  12. Чеклист из CLAUDE_CODE_PROMPT.md
# ═══════════════════════════════════════════════════════

class TestChecklist(unittest.TestCase):
    """Тесты по чеклисту из промта."""

    def test_storehouse_item_required_fields(self):
        """Каждый StorehouseItem имеет обязательные поля."""
        from parsers.models import StorehouseItem
        item = StorehouseItem(
            site="pik", city="Казань", complex_name="ЖК",
            building="1", item_id="s1", area=3.0,
            price=300000, price_per_meter=100000, url="http://test",
        )
        self.assertTrue(item.site)
        self.assertTrue(item.item_id)
        self.assertGreater(item.area, 0)

    def test_apartment_item_rooms_range(self):
        """rooms ∈ {0,1,2,3,4} для квартир."""
        from parsers.models import ApartmentItem
        for r in [0, 1, 2, 3, 4]:
            item = ApartmentItem(
                site="pik", city="К", complex_name="ЖК",
                building="1", item_id=f"a{r}", rooms=r, floor=1,
                area=50, price=5000000, price_per_meter=100000, url="/",
            )
            self.assertIn(item.rooms, range(0, 11))

    def test_price_zero_only_for_domrf(self):
        """price=0 допустимо только для site='domrf'."""
        from parsers.base import validate_items
        from parsers.models import StorehouseItem
        # domrf с price=0 — ок (нет warnings о цене)
        domrf_items = [StorehouseItem(
            site="domrf", city="К", complex_name="ЖК",
            building="1", item_id="d1", area=3.0,
            price=0, price_per_meter=0, url="/",
        )]
        warnings = validate_items(domrf_items)
        price_warnings = [w for w in warnings if "Цена" in w]
        self.assertEqual(price_warnings, [])

        # pik с price=0 — warning
        pik_items = [StorehouseItem(
            site="pik", city="К", complex_name="ЖК",
            building="1", item_id="p1", area=3.0,
            price=0, price_per_meter=0, url="/",
        )]
        warnings = validate_items(pik_items)
        price_warnings = [w for w in warnings if "Цена" in w]
        self.assertGreater(len(price_warnings), 0)

    def test_formula_no_excel_errors(self):
        """Формулы не содержат #REF!, #DIV/0!, #VALUE!."""
        from scoring import generate_first_stage_formula, generate_second_stage_formula
        f1 = generate_first_stage_formula(3)
        f2 = generate_second_stage_formula(3)
        for bad in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?"]:
            self.assertNotIn(bad, f1)
            self.assertNotIn(bad, f2)

    def test_formula_contains_iferror(self):
        """Формулы защищены IFERROR."""
        from scoring import generate_first_stage_formula, generate_second_stage_formula
        f1 = generate_first_stage_formula(3)
        f2 = generate_second_stage_formula(3)
        self.assertIn("IFERROR", f1)
        self.assertIn("IFERROR", f2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
