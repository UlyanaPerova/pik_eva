"""
Экспорт данных в xlsx с примечаниями (история цен, скидки).

Два листа:
  1. «Кладовки» — красивый, со строками-заголовками Город/ЖК,
     разделителями корпусов, секцией в примечаниях. Сортировка внутри каждого ЖК.
  2. «Все данные» — плоская таблица со всеми колонками,
     автофильтром, сортировкой, разделителями корпусов,
     группировкой строк по ЖК для скрытия.

Функции:
  - Скрытие ЖК через группировку строк (outline) на обоих листах
  - Разделители корпусов на обоих листах
  - История цен и скидки в примечаниях
  - Примечание «Добавлена от [дата]» на новых кладовках
  - Столбец «Исх. порядок» для сброса сортировки
"""
from __future__ import annotations

import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from parsers.base import (
    StorehouseItem, get_price_history, get_first_seen_date,
    get_all_known_ids, logger, PROJECT_DIR,
)

OUTPUT_DIR = PROJECT_DIR / "output"

# ── Стили ────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

CITY_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CITY_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)

COMPLEX_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
COMPLEX_FONT = Font(name="Calibri", bold=True, color="2F5496", size=12)

DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
NEW_ITEM_FONT = Font(name="Calibri", size=11, color="2E7D32")

TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
COMPLEX_TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GRAND_TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

BUILDING_BOTTOM = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="808080"),
)

BUILDING_BOTTOM_ALL = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="medium", color="808080"),
)

SITE_NAMES = {
    "pik": "ПИК",
    "akbarsdom": "Ак Бар Дом",
    "smu88": "СМУ-88",
    "glorax": "GloraX",
    "unistroy": "УниСтрой",
}

SITE_FILE_KEYS = {
    "pik": "PIK",
    "akbarsdom": "AkBarsDom",
    "smu88": "SMU88",
    "glorax": "GloraX",
    "unistroy": "Unistroy",
}

# ── Колонки ──────────────────────────────────────────

PRETTY_COLUMNS = [
    "Корпус",
    "Кол-во\nкладовок",
    "Номер\nкладовой",
    "Площадь\n(м²)",
    "Цена (₽)",
    "Цена/м²\n(₽)",
    "Ссылка",
    "Исх.\nпорядок",
]
PCOL = len(PRETTY_COLUMNS)

FLAT_COLUMNS = [
    "Город",
    "Застройщик",
    "ЖК",
    "Корпус",
    "Кол-во кладовок",
    "Номер кладовой",
    "Площадь (м²)",
    "Цена (₽)",
    "Цена/м² (₽)",
    "Ссылка",
    "Исх. порядок",
]
FCOL = len(FLAT_COLUMNS)


def _sort_key(it: StorehouseItem):
    try:
        num = int(it.item_number) if it.item_number else 999999
    except ValueError:
        num = 999999
    return (
        it.city.lower(),
        it.site.lower(),
        it.complex_name.lower(),
        it.building.lower(),
        num,
    )


def export_xlsx(
    items: list[StorehouseItem],
    conn: sqlite3.Connection,
    filename: str | None = None,
) -> Path:
    """Создать xlsx с двумя листами."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        sites_in_data = sorted(set(
            SITE_FILE_KEYS.get(it.site, it.site) for it in items
        ))
        suffix = "_".join(sites_in_data) if sites_in_data else "ALL"
        filename = f"storehouses_{suffix}.xlsx"

    output_path = OUTPUT_DIR / filename

    # Определяем новые кладовки (которых раньше не было в БД)
    sites = set(it.site for it in items)
    previously_known: set[str] = set()
    for site in sites:
        previously_known |= get_all_known_ids(conn, site)

    wb = Workbook()

    ws_pretty = wb.active
    ws_pretty.title = "Кладовки"
    _fill_pretty_sheet(ws_pretty, items, conn, previously_known)

    ws_flat = wb.create_sheet("Все данные")
    _fill_flat_sheet(ws_flat, items, conn, previously_known)

    wb.save(output_path)
    logger.info("xlsx сохранён: %s (%d кладовок)", output_path, len(items))
    return output_path


# ══════════════════════════════════════════════════════
#  ЛИСТ 1: «Кладовки» — красивый
# ══════════════════════════════════════════════════════

def _fill_pretty_sheet(ws, items, conn, previously_known) -> None:
    sorted_items = sorted(items, key=_sort_key)

    city_groups = defaultdict(lambda: defaultdict(list))
    for item in sorted_items:
        city_groups[item.city][(item.site, item.complex_name)].append(item)

    count_key = lambda it: (it.site, it.complex_name, it.building)
    counts = Counter(count_key(it) for it in items)

    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 1

    for city in sorted(city_groups.keys()):
        # ГОРОД
        ws.cell(row=row, column=1, value=city.upper()).font = CITY_FONT
        ws.cell(row=row, column=1).fill = CITY_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
        for c in range(1, PCOL + 1):
            ws.cell(row=row, column=c).fill = CITY_FILL
        ws.row_dimensions[row].height = 30
        row += 1

        for (site, complex_name) in sorted(city_groups[city].keys()):
            complex_items = city_groups[city][(site, complex_name)]
            display_name = SITE_NAMES.get(site, site)
            total_in_complex = len(complex_items)

            # ЖК
            jk_text = f"{display_name}  —  {complex_name}  ({total_in_complex} кладовок)"
            jk_cell = ws.cell(row=row, column=1, value=jk_text)
            jk_cell.font = COMPLEX_FONT
            jk_cell.fill = COMPLEX_FILL
            jk_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
            for c in range(1, PCOL + 1):
                ws.cell(row=row, column=c).fill = COMPLEX_FILL
            ws.row_dimensions[row].height = 25
            jk_row = row
            row += 1

            # Шапка
            for col_idx, col_name in enumerate(PRETTY_COLUMNS, start=1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            row += 1

            block_start = row

            # Данные
            for i, item in enumerate(complex_items):
                is_last = i == len(complex_items) - 1
                is_last_in_building = (
                    is_last
                    or complex_items[i].building != complex_items[i + 1].building
                )

                storehouse_count = counts[count_key(item)]
                try:
                    number_val = int(item.item_number) if item.item_number else ""
                except ValueError:
                    number_val = item.item_number or ""

                row_data = [
                    item.building,
                    storehouse_count,
                    number_val,
                    item.area,
                    item.price,
                    item.price_per_meter,
                    "Открыть",
                    row - block_start + 1,
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = DATA_FONT
                    cell.alignment = DATA_ALIGN
                    cell.border = THIN_BORDER

                # Форматы
                ws.cell(row=row, column=4).number_format = '0.0'
                ws.cell(row=row, column=5).number_format = '#,##0'
                ws.cell(row=row, column=6).number_format = '#,##0'
                if isinstance(number_val, int):
                    ws.cell(row=row, column=3).number_format = '0'
                ws.cell(row=row, column=2).number_format = '0'
                ws.cell(row=row, column=8).number_format = '0'
                ws.cell(row=row, column=8).font = Font(name="Calibri", size=9, color="BBBBBB")

                # Ссылка
                lc = ws.cell(row=row, column=7)
                if item.url:
                    lc.hyperlink = item.url
                    lc.font = LINK_FONT
                lc.alignment = DATA_ALIGN

                # Разделитель корпусов — на ВСЕХ столбцах строки
                if is_last_in_building and not is_last:
                    for c in range(1, PCOL + 1):
                        ws.cell(row=row, column=c).border = BUILDING_BOTTOM

                # Секция в примечании
                section_name = getattr(item, '_section_name', None)
                if section_name:
                    bc = ws.cell(row=row, column=1)
                    bc.comment = Comment(section_name, "Парсер кладовок")
                    bc.comment.width = 150
                    bc.comment.height = 30

                # Новая кладовка — примечание на номере
                _add_new_item_comment(ws, row, 3, item, previously_known, conn)

                # Примечания цен
                _add_price_comment(ws, row, 5, conn, item)
                _add_ppm_comment(ws, row, 6, conn, item)

                row += 1

            block_end = row - 1

            # Группировка ЖК (outline)
            group_start = jk_row + 1
            group_end = block_end
            if group_end >= group_start:
                ws.row_dimensions.group(group_start, group_end, outline_level=1)

    # ── Итоги по ЖК ──
    row += 1
    ws.cell(row=row, column=1, value="ИТОГИ ПО ЖК").font = Font(
        name="Calibri", size=12, bold=True
    )
    row += 1
    for col_idx, col_name in enumerate(["Город", "Застройщик", "ЖК", "Всего"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    complex_counts = defaultdict(int)
    for item in items:
        complex_counts[(item.city, item.site, item.complex_name)] += 1

    for (city, site, cname), total in sorted(complex_counts.items()):
        display_name = SITE_NAMES.get(site, site)
        for col, val in [(1, city), (2, display_name), (3, cname), (4, total)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = TOTAL_FONT
            cell.fill = COMPLEX_TOTAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    total_all = sum(complex_counts.values())
    ws.cell(row=row, column=3, value="ВСЕГО").font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    t = ws.cell(row=row, column=4, value=total_all)
    t.font = Font(name="Calibri", size=12, bold=True)
    t.fill = GRAND_TOTAL_FILL
    t.border = THIN_BORDER
    t.alignment = Alignment(horizontal="center")

    for i, w in enumerate([16, 14, 14, 12, 14, 14, 12, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════
#  ЛИСТ 2: «Все данные» — плоская таблица
# ══════════════════════════════════════════════════════

def _fill_flat_sheet(ws, items, conn, previously_known) -> None:
    sorted_items = sorted(items, key=_sort_key)

    count_key = lambda it: (it.site, it.complex_name, it.building)
    counts = Counter(count_key(it) for it in items)

    ws.sheet_properties.outlinePr.summaryBelow = False

    # Заголовки
    for col_idx, col_name in enumerate(FLAT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    # Данные — считаем диапазоны ЖК для группировки
    prev_complex = None
    prev_building = None
    complex_ranges: list[tuple[int, int]] = []  # (start_row, end_row) per ЖК
    cur_complex_start = 2

    for i, item in enumerate(sorted_items):
        row = i + 2
        display_name = SITE_NAMES.get(item.site, item.site)
        storehouse_count = counts[count_key(item)]

        try:
            number_val = int(item.item_number) if item.item_number else ""
        except ValueError:
            number_val = item.item_number or ""

        cur_complex = (item.city, item.site, item.complex_name)
        cur_building = (item.city, item.site, item.complex_name, item.building)

        # Фиксируем конец предыдущего ЖК
        if cur_complex != prev_complex and prev_complex is not None:
            complex_ranges.append((cur_complex_start, row - 1))
            cur_complex_start = row

        # Разделитель корпусов — жирная нижняя граница на предыдущей строке
        if cur_building != prev_building and prev_building is not None:
            if cur_complex == prev_complex:  # не между ЖК, а внутри
                prev_row = row - 1
                for c in range(1, FCOL + 1):
                    ws.cell(row=prev_row, column=c).border = BUILDING_BOTTOM

        prev_complex = cur_complex
        prev_building = cur_building

        row_data = [
            item.city,
            display_name,
            item.complex_name,
            item.building,
            storehouse_count,
            number_val,
            item.area,
            item.price,
            item.price_per_meter,
            "Открыть",
            i + 1,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

        # Форматы
        ws.cell(row=row, column=7).number_format = '0.0'
        ws.cell(row=row, column=8).number_format = '#,##0'
        ws.cell(row=row, column=9).number_format = '#,##0'
        if isinstance(number_val, int):
            ws.cell(row=row, column=6).number_format = '0'
        ws.cell(row=row, column=5).number_format = '0'
        ws.cell(row=row, column=11).number_format = '0'
        ws.cell(row=row, column=11).font = Font(name="Calibri", size=9, color="BBBBBB")

        # Ссылка
        lc = ws.cell(row=row, column=10)
        if item.url:
            lc.hyperlink = item.url
            lc.font = LINK_FONT
        lc.alignment = DATA_ALIGN

        # Секция
        section_name = getattr(item, '_section_name', None)
        if section_name:
            bc = ws.cell(row=row, column=4)
            bc.comment = Comment(section_name, "Парсер кладовок")
            bc.comment.width = 150
            bc.comment.height = 30

        # Новая кладовка
        _add_new_item_comment(ws, row, 6, item, previously_known, conn)

        # Примечания цен
        _add_price_comment(ws, row, 8, conn, item)
        _add_ppm_comment(ws, row, 9, conn, item)

    # Последний ЖК
    last_row = len(sorted_items) + 1
    if sorted_items:
        complex_ranges.append((cur_complex_start, last_row))

    # Жирная граница между ЖК
    for idx, (cs, ce) in enumerate(complex_ranges):
        if idx < len(complex_ranges) - 1:
            # Толстая нижняя граница на последней строке ЖК (на всех столбцах)
            thick_bottom = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="medium", color="4472C4"),
            )
            for c in range(1, FCOL + 1):
                ws.cell(row=ce, column=c).border = thick_bottom

    # Группировка строк по ЖК (outline)
    for cs, ce in complex_ranges:
        if ce > cs:
            # Группируем все строки ЖК кроме первой (первая — summary)
            ws.row_dimensions.group(cs + 1, ce, outline_level=1)

    # Автофильтр
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(FCOL)}{last_row}"

    # Ширина
    flat_widths = [14, 16, 18, 16, 14, 14, 12, 14, 14, 12, 8]
    for i, w in enumerate(flat_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Примечания ───────────────────────────────────────────

def _add_new_item_comment(ws, row, col, item, previously_known, conn):
    """Примечание «Добавлена от [дата]» для новых кладовок."""
    if item.item_id in previously_known:
        # Не новая — была в БД до текущего парсинга
        return

    # Новая кладовка — ещё не сохранена в БД
    from datetime import datetime
    date_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    cell = ws.cell(row=row, column=col)
    cell.comment = Comment(
        f"Добавлена от {date_str}", "Парсер кладовок"
    )
    cell.comment.width = 200
    cell.comment.height = 30
    # Зелёный шрифт — выделить новую
    cell.font = NEW_ITEM_FONT


def _add_price_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена» — скидка + история."""
    lines = []

    if item.discount_percent and item.original_price:
        lines.append(
            f"Скидка {item.discount_percent:.0f}%, "
            f"цена без скидки: {item.original_price:,.0f} ₽"
        )

    history = get_price_history(conn, item.site, item.item_id)
    if len(history) > 1:
        lines.append("Предыдущие цены:")
        for price, ppm, orig_price, discount, date_str in history[1:]:
            date_short = date_str[:10]
            entry = f"• {price:,.0f} ₽ ({date_short})"
            if discount and orig_price:
                entry += f" [скидка {discount:.0f}%, без скидки: {orig_price:,.0f} ₽]"
            lines.append(entry)

    if lines:
        cell = ws.cell(row=row, column=col)
        cell.comment = Comment("\n".join(lines), "Парсер кладовок")
        cell.comment.width = 350
        cell.comment.height = max(80, len(lines) * 20)


def _add_ppm_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена/м²» — история."""
    history = get_price_history(conn, item.site, item.item_id)
    if len(history) <= 1:
        return

    lines = ["Предыдущие цены/м²:"]
    for price, ppm, orig_price, discount, date_str in history[1:]:
        date_short = date_str[:10]
        lines.append(f"• {ppm:,.0f} ₽/м² ({date_short})")

    cell = ws.cell(row=row, column=col)
    cell.comment = Comment("\n".join(lines), "Парсер кладовок")
    cell.comment.width = 300
    cell.comment.height = max(60, len(lines) * 18)
