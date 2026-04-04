"""
Экспорт данных квартир в xlsx с примечаниями и средними ценами.

Три листа:
  1. «Квартиры» — красивый, сгруппированный по Город → Застройщик/ЖК → Тип квартиры
  2. «Все данные» — плоская таблица со всеми колонками + автофильтр
  3. «Средние цены» — таблица средних цен по ЖК и типам квартир
"""
from __future__ import annotations

import re
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import json

from parsers.apartments_base import (
    ApartmentItem, rooms_label,
    get_price_history, get_first_seen_date,
    get_all_known_ids, load_or_create_baseline,
    calc_avg_prices,
    logger, OUTPUT_DIR, BASELINE_DIR,
)

from exporter_common import (
    HEADER_FILL, HEADER_FONT,
    CITY_FILL, CITY_FONT,
    COMPLEX_FILL, COMPLEX_FONT,
    DATA_FONT, DATA_ALIGN,
    LINK_FONT, NEW_ITEM_FILL,
    TOTAL_FONT, COMPLEX_TOTAL_FILL, GRAND_TOTAL_FILL,
    THIN_BORDER, BUILDING_BOTTOM,
    SITE_NAMES, SITE_FILE_KEYS,
    append_comment as _append_comment,
    natural_sort_key as _natural_sort_key,
    add_new_item_comment as _add_new_item_comment_common,
    add_price_comment as _add_price_comment_common,
    add_ppm_comment as _add_ppm_comment_common,
)

# ── Стили (специфические для квартир) ────────────────
ROOM_TYPE_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
ROOM_TYPE_FONT = Font(name="Calibri", bold=True, color="2F5496", size=11)

AVG_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
AVG_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
AVG_DATA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
AVG_TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


# ── Колонки ──────────────────────────────────────────
PRETTY_COLUMNS = [
    "Корпус",
    "Тип",
    "Кол-во",
    "Номер",
    "Этаж",
    "Площадь\n(м²)",
    "Жилая\nплощадь\n(м²)",
    "Цена (₽)",
    "Цена/м²\n(₽)",
    "Ссылка",
    "Исх.\nпорядок",
]
PCOL = len(PRETTY_COLUMNS)

FLAT_COLUMNS = [
    "Город",
    "Застройщик",
    "ЖК",
    "Корпус",
    "Тип",
    "Кол-во",
    "Номер",
    "Этаж",
    "Площадь (м²)",
    "Жилая площадь (м²)",
    "Цена (₽)",
    "Цена/м² (₽)",
    "Ссылка",
    "Исх. порядок",
    "ID",
]
FCOL = len(FLAT_COLUMNS)

SOLD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# _natural_sort_key — импортирована из exporter_common


def _sort_key(it: ApartmentItem):
    try:
        num = int(it.apartment_number) if it.apartment_number else 999999
    except ValueError:
        num = 999999
    developer = (getattr(it, "developer", None) or "").lower()
    return (
        it.city.lower(),
        developer,
        it.site.lower(),
        it.complex_name.lower(),
        it.rooms,
        _natural_sort_key(it.building),
        it.floor,
        num,
    )


def export_apartments_xlsx(
    items: list[ApartmentItem],
    conn: sqlite3.Connection,
    filename: str | None = None,
    previously_known: set[str] | None = None,
    merge_result=None,
) -> Path:
    """Создать xlsx с тремя листами."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if filename is None:
        sites_in_data = sorted(set(
            SITE_FILE_KEYS.get(it.site, it.site) for it in items
        ))
        suffix = "_".join(sites_in_data) if sites_in_data else "ALL"
        filename = f"apartments_{suffix}.xlsx"

    output_path = OUTPUT_DIR / filename

    if previously_known is None:
        sites = set(it.site for it in items)
        previously_known = set()
        for site in sites:
            previously_known |= get_all_known_ids(conn, site)

    baseline_ids = load_or_create_baseline(items)

    wb = Workbook()

    ws_pretty = wb.active
    ws_pretty.title = "Квартиры"
    _fill_pretty_sheet(ws_pretty, items, conn, previously_known, baseline_ids)

    ws_flat = wb.create_sheet("Все данные")
    _fill_flat_sheet(ws_flat, items, conn, previously_known, baseline_ids,
                     merge_result=merge_result)

    # Лист «Средние цены» — только если есть квартиры с ценами (не для domrf)
    has_priced = any(it.price > 0 for it in items)
    if has_priced:
        ws_avg = wb.create_sheet("Средние цены")
        _fill_avg_sheet(ws_avg, items)

    wb.save(output_path)
    logger.info("xlsx сохранён: %s (%d квартир)", output_path, len(items))
    return output_path


# ══════════════════════════════════════════════════════
#  ЛИСТ 1: «Квартиры» — красивый
# ══════════════════════════════════════════════════════

def _fill_pretty_sheet(ws, items, conn, previously_known, baseline_ids) -> None:
    sorted_items = sorted(items, key=_sort_key)

    # Группировка: город → (застройщик, ЖК, сайт) → тип квартиры → список
    city_groups = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for item in sorted_items:
        dev = (getattr(item, "developer", None) or "").lower()
        city_groups[item.city][(dev, item.complex_name, item.site)][item.rooms].append(item)

    # Подсчёт количества по (сайт, ЖК, корпус, тип)
    count_key = lambda it: (it.site, it.complex_name, it.building.split("||")[0].strip(), it.rooms)
    counts = Counter(count_key(it) for it in items)

    ws.sheet_properties.outlinePr.summaryBelow = False
    row = 1

    for city in sorted(city_groups.keys()):
        # ГОРОД
        ws.cell(row=row, column=1, value=city.upper()).font = CITY_FONT
        ws.cell(row=row, column=1).fill = CITY_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
        for c in range(1, PCOL + 1):
            ws.cell(row=row, column=c).fill = CITY_FILL
        ws.row_dimensions[row].height = 30
        row += 1

        for (dev_key, complex_name, site) in sorted(city_groups[city].keys()):
            rooms_groups = city_groups[city][(dev_key, complex_name, site)]
            first_item = next(iter(next(iter(rooms_groups.values()))))
            dev = getattr(first_item, "developer", None)
            total_in_complex = sum(len(v) for v in rooms_groups.values())

            # ЖК
            if dev:
                jk_text = f"{dev}  —  {complex_name}  ({total_in_complex} квартир)"
            else:
                jk_text = f"{complex_name}  ({total_in_complex} квартир)"
            jk_cell = ws.cell(row=row, column=1, value=jk_text)
            jk_cell.font = COMPLEX_FONT
            jk_cell.fill = COMPLEX_FILL
            jk_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
            for c in range(1, PCOL + 1):
                ws.cell(row=row, column=c).fill = COMPLEX_FILL
            ws.row_dimensions[row].height = 25
            jk_row = row
            row += 1

            # По типам квартир
            for rooms_num in sorted(rooms_groups.keys()):
                room_items = rooms_groups[rooms_num]
                rlabel = rooms_label(rooms_num)
                room_count = len(room_items)

                # Средняя цена по типу
                priced_items = [it for it in room_items if it.price > 0]
                avg_price = sum(it.price for it in priced_items) / len(priced_items) if priced_items else 0
                avg_ppm = sum(it.price_per_meter for it in priced_items) / len(priced_items) if priced_items else 0

                # Подзаголовок типа квартиры
                if avg_price > 0:
                    type_text = (f"{rlabel}  ({room_count} шт.)  |  "
                                 f"ср. цена: {avg_price:,.0f} ₽  |  ср. цена/м²: {avg_ppm:,.0f} ₽")
                else:
                    type_text = f"{rlabel}  ({room_count} шт.)"
                ws.cell(row=row, column=1, value=type_text).font = ROOM_TYPE_FONT
                ws.cell(row=row, column=1).fill = ROOM_TYPE_FILL
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=PCOL)
                for c in range(1, PCOL + 1):
                    ws.cell(row=row, column=c).fill = ROOM_TYPE_FILL
                row += 1

                # Шапка
                for col_idx, col_name in enumerate(PRETTY_COLUMNS, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = THIN_BORDER
                row += 1

                block_start = row

                # Данные
                for i, item in enumerate(room_items):
                    is_last = i == len(room_items) - 1
                    cur_bld = item.building.split("||")[0].strip()
                    next_bld = room_items[i + 1].building.split("||")[0].strip() if not is_last else ""
                    is_last_in_building = is_last or cur_bld != next_bld

                    apt_count_in_group = counts[count_key(item)]
                    try:
                        number_val = int(item.apartment_number) if item.apartment_number else ""
                    except ValueError:
                        number_val = item.apartment_number or ""

                    building_display = item.building
                    building_note = None
                    if "||" in item.building:
                        parts = item.building.split("||", 1)
                        building_display = parts[0].strip()
                        building_note = parts[1].strip()

                    display_price = item.price if item.price else ""
                    display_ppm = item.price_per_meter if item.price_per_meter else ""

                    living_area_val = item.living_area if item.living_area else ""

                    row_data = [
                        building_display,
                        item.rooms_label,
                        apt_count_in_group,
                        number_val,
                        item.floor,
                        item.area,
                        living_area_val,
                        display_price,
                        display_ppm,
                        "Открыть",
                        row - block_start + 1,
                    ]

                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.font = DATA_FONT
                        cell.alignment = DATA_ALIGN
                        cell.border = THIN_BORDER

                    if building_note:
                        _append_comment(ws.cell(row=row, column=1), building_note, "Parser")

                    # Форматы
                    area_fmt = '0.00' if item.site == 'domrf' else '0.0'
                    ws.cell(row=row, column=6).number_format = area_fmt
                    ws.cell(row=row, column=7).number_format = area_fmt  # living area
                    ws.cell(row=row, column=8).number_format = '#,##0'
                    ws.cell(row=row, column=9).number_format = '#,##0'
                    ws.cell(row=row, column=3).number_format = '0'
                    ws.cell(row=row, column=5).number_format = '0'
                    if isinstance(number_val, int):
                        ws.cell(row=row, column=4).number_format = '0'
                    ws.cell(row=row, column=11).number_format = '0'
                    ws.cell(row=row, column=11).font = Font(name="Calibri", size=9, color="BBBBBB")

                    # Ссылка
                    lc = ws.cell(row=row, column=10)
                    if item.url:
                        lc.hyperlink = item.url
                        lc.font = LINK_FONT
                    lc.alignment = DATA_ALIGN

                    # Разделитель корпусов
                    if is_last_in_building and not is_last:
                        for c in range(1, PCOL + 1):
                            ws.cell(row=row, column=c).border = BUILDING_BOTTOM

                    # Секция
                    section_name = getattr(item, '_section_name', None)
                    if section_name:
                        bc = ws.cell(row=row, column=1)
                        _append_comment(bc, section_name, "Парсер квартир")
                        bc.comment.width = 150
                        bc.comment.height = 30

                    # Новая квартира
                    _add_new_item_comment(ws, row, 4, item, previously_known, conn,
                                          baseline_ids, total_cols=PCOL)

                    # Примечания цен
                    _add_price_comment(ws, row, 8, conn, item)
                    _add_ppm_comment(ws, row, 9, conn, item)

                    row += 1

                block_end = row - 1
                header_row = block_start - 1  # строка с шапкой колонок

                # Excel Table — сортировка и фильтр для каждого блока типа квартиры
                if block_end >= block_start:
                    safe_complex = re.sub(r'[^A-Za-z0-9]', '', complex_name)
                    safe_site = re.sub(r'[^A-Za-z0-9]', '', site)
                    table_name = f"T_{safe_site}_{safe_complex}_r{rooms_num}_{jk_row}"
                    if not table_name[0].isalpha():
                        table_name = "T" + table_name

                    table_ref = (
                        f"A{header_row}:{get_column_letter(PCOL)}{block_end}"
                    )
                    tab = Table(displayName=table_name, ref=table_ref)
                    tab.tableStyleInfo = TableStyleInfo(
                        name="TableStyleLight9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    ws.add_table(tab)

            # Группировка ЖК (outline)
            group_start = jk_row + 1
            group_end = row - 1
            if group_end >= group_start:
                ws.row_dimensions.group(group_start, group_end, outline_level=1, hidden=True)

    # Ширины столбцов
    col_widths = [16, 10, 10, 10, 8, 12, 12, 14, 14, 12, 8]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════
#  ЛИСТ 2: «Все данные» — плоская таблица
# ══════════════════════════════════════════════════════

def _fill_flat_sheet(ws, items, conn, previously_known, baseline_ids,
                     merge_result=None) -> None:
    sorted_items = sorted(items, key=_sort_key)

    count_key = lambda it: (it.site, it.complex_name, it.building.split("||")[0].strip(), it.rooms)
    counts = Counter(count_key(it) for it in items)

    # Заголовки
    for col_idx, col_name in enumerate(FLAT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    prev_complex = None
    prev_building = None

    for i, item in enumerate(sorted_items):
        row = i + 2
        dev = getattr(item, "developer", None)
        display_name = dev if dev else ""
        apt_count = counts[count_key(item)]

        try:
            number_val = int(item.apartment_number) if item.apartment_number else ""
        except ValueError:
            number_val = item.apartment_number or ""

        cur_complex = (item.city, item.site, item.complex_name)
        cur_building = (item.city, item.site, item.complex_name,
                        item.building.split("||")[0].strip())

        # Жирная граница между корпусами
        if cur_building != prev_building and prev_building is not None:
            if cur_complex == prev_complex:
                prev_row = row - 1
                for c in range(1, FCOL + 1):
                    ws.cell(row=prev_row, column=c).border = BUILDING_BOTTOM

        # Жирная граница между ЖК
        if cur_complex != prev_complex and prev_complex is not None:
            prev_row = row - 1
            thick_bottom = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="medium", color="4472C4"),
            )
            for c in range(1, FCOL + 1):
                ws.cell(row=prev_row, column=c).border = thick_bottom

        prev_complex = cur_complex
        prev_building = cur_building

        building_display = item.building
        building_note = None
        if "||" in item.building:
            parts = item.building.split("||", 1)
            building_display = parts[0].strip()
            building_note = parts[1].strip()

        display_price = item.price if item.price else ""
        display_ppm = item.price_per_meter if item.price_per_meter else ""
        living_area_val = item.living_area if item.living_area else ""

        row_data = [
            item.city,
            display_name,
            item.complex_name,
            building_display,
            item.rooms_label,
            apt_count,
            number_val,
            item.floor,
            item.area,
            living_area_val,
            display_price,
            display_ppm,
            "Открыть",
            i + 1,
            item.item_id,
        ]

        # Применяем пользовательские правки (если есть)
        user_edits_for_item = {}
        if merge_result and hasattr(merge_result, 'user_edits'):
            user_edits_for_item = merge_result.user_edits.get(item.item_id, {})

        for col_idx, value in enumerate(row_data, start=1):
            # Если пользователь правил эту ячейку — сохраняем его значение
            if col_idx in user_edits_for_item:
                edit = user_edits_for_item[col_idx]
                value = edit.user_value

            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

        # Примечания о пользовательских правках
        for col_idx, edit in user_edits_for_item.items():
            cell = ws.cell(row=row, column=col_idx)
            from datetime import datetime as _dt
            date_str = _dt.now().strftime("%d.%m.%Y")
            comment_text = (
                f"{date_str}. Сохранено пользовательское значение.\n"
                f"Значение при парсинге: {edit.new_parser_value}"
            )
            _append_comment(cell, comment_text, "Smart Merge")
            cell.comment.width = 300
            cell.comment.height = 50

        # Подсветка новых квартир (зелёный) + комментарий
        is_new = merge_result and item.item_id in getattr(merge_result, 'new_ids', set())
        is_prev_new = merge_result and item.item_id in getattr(merge_result, 'prev_new_ids', set())
        if is_new:
            for c in range(1, FCOL + 1):
                ws.cell(row=row, column=c).fill = NEW_ITEM_FILL
            from datetime import datetime as _dt
            _c = ws.cell(row=row, column=7)  # Номер
            _append_comment(
                _c,
                f"Новая квартира от {_dt.now().strftime('%d.%m.%Y')}",
                "Smart Merge",
            )
            _c.comment.width = 200
            _c.comment.height = 30
        elif is_prev_new:
            # Был новым в прошлый раз — без цвета, но комментарий остаётся
            _c = ws.cell(row=row, column=7)
            _append_comment(_c, "Новая квартира (предыдущий парсинг)", "Smart Merge")
            _c.comment.width = 200
            _c.comment.height = 30

        if building_note:
            _append_comment(ws.cell(row=row, column=4), building_note, "Parser")

        # Форматы
        area_fmt = '0.00' if item.site == 'domrf' else '0.0'
        ws.cell(row=row, column=9).number_format = area_fmt
        ws.cell(row=row, column=10).number_format = area_fmt  # living area
        ws.cell(row=row, column=11).number_format = '#,##0'
        ws.cell(row=row, column=12).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = '0'
        ws.cell(row=row, column=8).number_format = '0'
        if isinstance(number_val, int):
            ws.cell(row=row, column=7).number_format = '0'
        ws.cell(row=row, column=14).number_format = '0'
        ws.cell(row=row, column=14).font = Font(name="Calibri", size=9, color="BBBBBB")

        # Ссылка
        lc = ws.cell(row=row, column=13)
        if item.url:
            lc.hyperlink = item.url
            lc.font = LINK_FONT
        lc.alignment = DATA_ALIGN

        # Секция
        section_name = getattr(item, '_section_name', None)
        if section_name:
            bc = ws.cell(row=row, column=4)
            _append_comment(bc, section_name, "Парсер квартир")
            bc.comment.width = 150
            bc.comment.height = 30

        # Новая квартира
        _add_new_item_comment(ws, row, 7, item, previously_known, conn,
                              baseline_ids, total_cols=FCOL)

        # Примечания цен
        _add_price_comment(ws, row, 11, conn, item)
        _add_ppm_comment(ws, row, 12, conn, item)

    # Проданные/снятые квартиры — добавляем в конец
    # sold_ids = впервые пропали → красный + комментарий
    # prev_sold_ids = пропали в прошлый раз → без цвета + комментарий
    all_sold = set()
    if merge_result:
        all_sold = getattr(merge_result, 'sold_ids', set()) | getattr(merge_result, 'prev_sold_ids', set())

    if all_sold and merge_result and merge_result.old_workbook:
        if "Все данные" in merge_result.old_workbook.sheetnames:
            old_ws = merge_result.old_workbook["Все данные"]
            id_col = None
            for col in range(1, old_ws.max_column + 1):
                if old_ws.cell(row=1, column=col).value == "ID":
                    id_col = col
                    break
            if id_col:
                sold_row = len(sorted_items) + 2
                from datetime import datetime as _dt
                date_str = _dt.now().strftime("%d.%m.%Y")

                for old_row in range(2, old_ws.max_row + 1):
                    old_id = old_ws.cell(row=old_row, column=id_col).value
                    if not old_id or str(old_id) not in all_sold:
                        continue

                    is_first_sold = str(old_id) in getattr(merge_result, 'sold_ids', set())

                    for col in range(1, FCOL + 1):
                        old_val = old_ws.cell(row=old_row, column=col).value
                        cell = ws.cell(row=sold_row, column=col, value=old_val)
                        cell.font = DATA_FONT
                        cell.alignment = DATA_ALIGN
                        cell.border = THIN_BORDER
                        if is_first_sold:
                            cell.fill = SOLD_FILL

                    # Комментарий
                    comment_cell = ws.cell(row=sold_row, column=7)  # Номер
                    if is_first_sold:
                        _append_comment(
                            comment_cell,
                            f"Снята с продажи от {date_str}", "Smart Merge"
                        )
                    else:
                        _append_comment(
                            comment_cell,
                            "Снята с продажи (предыдущий парсинг)", "Smart Merge"
                        )
                    comment_cell.comment.width = 250
                    comment_cell.comment.height = 30

                    sold_row += 1

    # Автофильтр
    last_row = ws.max_row
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(FCOL)}{last_row}"

    # Ширина
    flat_widths = [14, 16, 18, 16, 10, 10, 10, 8, 12, 12, 14, 14, 12, 8, 0]
    for i, w in enumerate(flat_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Скрываем столбец ID
    ws.column_dimensions[get_column_letter(FCOL)].hidden = True


# ══════════════════════════════════════════════════════
#  ЛИСТ 3: «Средние цены»
# ══════════════════════════════════════════════════════

def _fill_avg_sheet(ws, items: list[ApartmentItem]) -> None:
    """Лист со средними ценами по ЖК и типам квартир."""
    stats = calc_avg_prices(items)
    row = 1

    # ── Общие средние цены ──
    ws.cell(row=row, column=1, value="СРЕДНИЕ ЦЕНЫ ПО ТИПАМ КВАРТИР").font = Font(
        name="Calibri", size=14, bold=True
    )
    row += 2

    avg_cols = ["Тип квартиры", "Кол-во", "Ср. цена (₽)", "Ср. цена/м² (₽)",
                "Ср. площадь (м²)", "Мин. цена (₽)", "Макс. цена (₽)"]
    for col_idx, col_name in enumerate(avg_cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = AVG_HEADER_FILL
        cell.font = AVG_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    for rooms_num, data in stats["by_rooms"].items():
        rlabel = rooms_label(rooms_num)
        row_data = [
            rlabel,
            data["count"],
            data["avg_price"],
            data["avg_ppm"],
            data["avg_area"],
            data["min_price"],
            data["max_price"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            cell.fill = AVG_DATA_FILL

        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '0.0'
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=7).number_format = '#,##0'
        row += 1

    # Итого
    total = stats["total"]
    total_data = ["ВСЕГО", total["count"], total["avg_price"], total["avg_ppm"],
                  total["avg_area"], total["min_price"], total["max_price"]]
    for col_idx, value in enumerate(total_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER
        cell.fill = AVG_TOTAL_FILL

    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=5).number_format = '0.0'
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=7).number_format = '#,##0'
    row += 3

    # ── Средние цены по ЖК ──
    ws.cell(row=row, column=1, value="СРЕДНИЕ ЦЕНЫ ПО ЖК И ТИПАМ").font = Font(
        name="Calibri", size=14, bold=True
    )
    row += 2

    jk_cols = ["ЖК", "Тип квартиры", "Кол-во", "Ср. цена (₽)", "Ср. цена/м² (₽)",
               "Ср. площадь (м²)", "Мин. цена (₽)", "Макс. цена (₽)"]
    for col_idx, col_name in enumerate(jk_cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = AVG_HEADER_FILL
        cell.font = AVG_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    row += 1

    prev_jk = None
    for (complex_name, rooms_num), data in stats["by_complex_rooms"].items():
        rlabel = rooms_label(rooms_num)

        # Жирная граница между ЖК
        if prev_jk is not None and complex_name != prev_jk:
            prev_row = row - 1
            for c in range(1, len(jk_cols) + 1):
                cell = ws.cell(row=prev_row, column=c)
                cell.border = Border(
                    left=Side(style="thin", color="D9D9D9"),
                    right=Side(style="thin", color="D9D9D9"),
                    top=Side(style="thin", color="D9D9D9"),
                    bottom=Side(style="medium", color="548235"),
                )
        prev_jk = complex_name

        row_data = [
            complex_name, rlabel, data["count"],
            data["avg_price"], data["avg_ppm"], data["avg_area"],
            data["min_price"], data["max_price"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = '0.0'
        ws.cell(row=row, column=7).number_format = '#,##0'
        ws.cell(row=row, column=8).number_format = '#,##0'
        row += 1

    # Ширины
    avg_widths = [20, 14, 10, 16, 16, 14, 16, 16]
    for i, w in enumerate(avg_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Примечания ───────────────────────────────────────────

def _add_new_item_comment(ws, row, col, item, previously_known, conn,
                          baseline_ids, total_cols=None):
    """Примечание «Добавлена от [дата]» — делегирует в exporter_common."""
    _add_new_item_comment_common(
        ws, row, col, item, previously_known, conn, baseline_ids,
        get_first_seen_fn=get_first_seen_date,
        author="Парсер квартир",
        total_cols=total_cols,
    )


def _add_price_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена» — делегирует в exporter_common."""
    _add_price_comment_common(
        ws, row, col, conn, item,
        get_price_history_fn=get_price_history,
        author="Парсер квартир",
    )


def _add_ppm_comment(ws, row, col, conn, item):
    """Примечание к ячейке «Цена/м²» — делегирует в exporter_common."""
    _add_ppm_comment_common(
        ws, row, col, conn, item,
        get_price_history_fn=get_price_history,
        author="Парсер квартир",
    )
